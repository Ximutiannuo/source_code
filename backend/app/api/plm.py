from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_active_user
from app.services.plm_service import PLMService

router = APIRouter()

VERSION_SHEET_NAME = "BOM版本"
DETAIL_SHEET_NAME = "BOM明细"


class MaterialRead(BaseModel):
    id: int
    code: str
    name: str
    specification: Optional[str] = None
    category: Optional[str] = None
    unit: Optional[str] = None
    material_type: Optional[str] = None
    drawing_no: Optional[str] = None
    revision: Optional[str] = None
    safety_stock: Optional[float] = None
    current_stock: Optional[float] = None
    reserved_stock: Optional[float] = None
    incoming_stock: Optional[float] = None
    lead_time_days: Optional[int] = None
    description: Optional[str] = None

    class Config:
        from_attributes = True


class MaterialCreate(BaseModel):
    code: str
    name: str
    specification: Optional[str] = None
    category: Optional[str] = None
    unit: Optional[str] = "PCS"
    material_type: Optional[str] = "PART"
    drawing_no: Optional[str] = None
    revision: Optional[str] = "A"
    safety_stock: Optional[float] = 0
    current_stock: Optional[float] = 0
    reserved_stock: Optional[float] = 0
    incoming_stock: Optional[float] = 0
    lead_time_days: Optional[int] = 0
    description: Optional[str] = None


class MaterialUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    specification: Optional[str] = None
    category: Optional[str] = None
    unit: Optional[str] = None
    material_type: Optional[str] = None
    drawing_no: Optional[str] = None
    revision: Optional[str] = None
    safety_stock: Optional[float] = None
    current_stock: Optional[float] = None
    reserved_stock: Optional[float] = None
    incoming_stock: Optional[float] = None
    lead_time_days: Optional[int] = None
    description: Optional[str] = None


class BOMHeaderRead(BaseModel):
    id: int
    product_code: str
    version: str
    bom_type: Optional[str] = None
    status: Optional[str] = None
    description: Optional[str] = None
    is_active: bool
    product_family: Optional[str] = None
    business_unit: Optional[str] = None
    project_code: Optional[str] = None
    plant_code: Optional[str] = None
    discipline: Optional[str] = None
    source_system: Optional[str] = None
    source_file: Optional[str] = None
    sync_status: Optional[str] = None
    cad_document_no: Optional[str] = None
    released_by: Optional[str] = None
    last_synced_at: Optional[datetime] = None
    material: Optional[MaterialRead] = None


class DrawingDocumentLinkRead(BaseModel):
    id: int
    document_number: str
    document_name: str
    document_type: str
    status: str
    version: Optional[str] = None
    revision: Optional[str] = None
    material_code: Optional[str] = None
    product_code: Optional[str] = None

    class Config:
        from_attributes = True


class BOMItemRead(BaseModel):
    id: Optional[int] = None
    parent_item_code: Optional[str] = None
    child_item_code: str
    quantity: float = 1
    component_type: Optional[str] = None
    routing_link: Optional[str] = None
    find_number: Optional[str] = None
    item_level: int = 1
    item_category: Optional[str] = None
    procurement_type: Optional[str] = None
    loss_rate: float = 0
    unit_price: float = 0
    total_price: float = 0
    source_reference: Optional[str] = None
    drawing_document_id: Optional[int] = None
    drawing_mapping_status: Optional[str] = None
    drawing_validation_message: Optional[str] = None
    drawing_document: Optional[DrawingDocumentLinkRead] = None
    material: Optional[MaterialRead] = None


class BOMStatisticsRead(BaseModel):
    item_count: int
    leaf_count: int
    estimated_total_cost: float


class BOMDetailRead(BOMHeaderRead):
    items: List[BOMItemRead] = Field(default_factory=list)
    statistics: BOMStatisticsRead


class BOMNodeRead(BaseModel):
    id: int
    material_code: str
    material_name: str
    quantity: float
    unit: str
    level: int
    find_number: Optional[str] = None
    component_type: Optional[str] = None
    routing_link: Optional[str] = None
    item_category: Optional[str] = None
    procurement_type: Optional[str] = None
    loss_rate: float = 0
    unit_price: float = 0
    total_price: float = 0
    source_reference: Optional[str] = None
    children: List["BOMNodeRead"] = Field(default_factory=list)


class BOMItemUpsert(BaseModel):
    id: Optional[int] = None
    parent_item_code: Optional[str] = None
    child_item_code: str
    quantity: float = Field(default=1, ge=0)
    component_type: Optional[str] = "NORMAL"
    routing_link: Optional[str] = None
    find_number: Optional[str] = None
    item_level: Optional[int] = Field(default=None, ge=0)
    item_category: Optional[str] = None
    procurement_type: Optional[str] = None
    loss_rate: Optional[float] = Field(default=0, ge=0)
    unit_price: Optional[float] = Field(default=0, ge=0)
    total_price: Optional[float] = Field(default=None, ge=0)
    source_reference: Optional[str] = None
    drawing_document_id: Optional[int] = None
    material_name: Optional[str] = None
    specification: Optional[str] = None
    unit: Optional[str] = None
    drawing_no: Optional[str] = None
    revision: Optional[str] = None


class BOMItemDrawingMappingUpsert(BaseModel):
    bom_item_id: int
    drawing_document_id: Optional[int] = None


class BOMItemDrawingValidationRead(BaseModel):
    bom_item_id: int
    child_item_code: Optional[str] = None
    find_number: Optional[str] = None
    validation_status: str
    can_apply: bool
    message: str
    warnings: List[str] = Field(default_factory=list)
    errors: List[str] = Field(default_factory=list)
    current_document: Optional[DrawingDocumentLinkRead] = None
    candidate_document: Optional[DrawingDocumentLinkRead] = None


class BOMItemDrawingMappingResultRead(BaseModel):
    updated: int
    results: List[BOMItemDrawingValidationRead]


class BOMUpsert(BaseModel):
    id: Optional[int] = None
    product_code: str
    product_name: Optional[str] = None
    version: str = "v1.0"
    bom_type: str = "EBOM"
    status: str = "DRAFT"
    description: Optional[str] = None
    is_active: bool = True
    product_family: Optional[str] = None
    business_unit: Optional[str] = None
    project_code: Optional[str] = None
    plant_code: Optional[str] = None
    discipline: Optional[str] = None
    source_system: str = "MANUAL"
    source_file: Optional[str] = None
    sync_status: Optional[str] = "MANUAL"
    cad_document_no: Optional[str] = None
    items: List[BOMItemUpsert] = Field(default_factory=list)


class CADBOMSync(BOMUpsert):
    source_system: str = "SOLIDWORKS"
    source_file: Optional[str] = None


class ECNCreate(BaseModel):
    ecn_no: str
    title: Optional[str] = None
    description: Optional[str] = None
    change_type: str
    reason: Optional[str] = None
    impacts: List[Dict[str, Any]] = Field(default_factory=list)


VERSION_EXPORT_HEADERS = [
    "产品编码",
    "产品名称",
    "BOM版本",
    "BOM类型",
    "状态",
    "启用",
    "产品族",
    "事业部",
    "项目号",
    "工厂",
    "专业/维度",
    "来源系统",
    "来源文件",
    "同步状态",
    "CAD文档号",
    "说明",
]

DETAIL_EXPORT_HEADERS = [
    "产品编码",
    "BOM版本",
    "BOM类型",
    "父项编码",
    "子项编码",
    "子项名称",
    "规格型号",
    "单位",
    "层级",
    "位号",
    "数量",
    "组件类型",
    "物料维度",
    "采购类型",
    "工序关联",
    "损耗率",
    "单价",
    "总价",
    "图号",
    "版本/版次",
    "来源引用",
]

VERSION_HEADER_ALIASES = {
    "product_code": ["产品编码", "成品编码", "product_code"],
    "product_name": ["产品名称", "product_name"],
    "version": ["BOM版本", "版本", "version"],
    "bom_type": ["BOM类型", "bom_type"],
    "status": ["状态", "status"],
    "is_active": ["启用", "is_active"],
    "product_family": ["产品族", "product_family"],
    "business_unit": ["事业部", "business_unit"],
    "project_code": ["项目号", "project_code"],
    "plant_code": ["工厂", "plant_code"],
    "discipline": ["专业/维度", "discipline"],
    "source_system": ["来源系统", "source_system"],
    "source_file": ["来源文件", "source_file"],
    "sync_status": ["同步状态", "sync_status"],
    "cad_document_no": ["CAD文档号", "cad_document_no"],
    "description": ["说明", "description"],
}

DETAIL_HEADER_ALIASES = {
    "product_code": ["产品编码", "成品编码", "product_code"],
    "version": ["BOM版本", "版本", "version"],
    "bom_type": ["BOM类型", "bom_type"],
    "parent_item_code": ["父项编码", "parent_item_code"],
    "child_item_code": ["子项编码", "child_item_code"],
    "material_name": ["子项名称", "物料名称", "material_name"],
    "specification": ["规格型号", "specification"],
    "unit": ["单位", "unit"],
    "item_level": ["层级", "item_level"],
    "find_number": ["位号", "行号", "find_number"],
    "quantity": ["数量", "用量", "quantity"],
    "component_type": ["组件类型", "component_type"],
    "item_category": ["物料维度", "物料分类", "item_category"],
    "procurement_type": ["采购类型", "procurement_type"],
    "routing_link": ["工序关联", "routing_link"],
    "loss_rate": ["损耗率", "loss_rate"],
    "unit_price": ["单价", "unit_price"],
    "total_price": ["总价", "total_price"],
    "drawing_no": ["图号", "drawing_no"],
    "revision": ["版本/版次", "revision"],
    "source_reference": ["来源引用", "source_reference"],
}


def _normalize_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _parse_float(value: Any, default: float = 0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是", "启用", "active"}


def _build_header_map(headers: List[Any], aliases: Dict[str, List[str]]) -> Dict[str, int]:
    normalized_headers = {str(value or "").strip(): index for index, value in enumerate(headers)}
    result: Dict[str, int] = {}
    for field, names in aliases.items():
        for name in names:
            if name in normalized_headers:
                result[field] = normalized_headers[name]
                break
    return result


def _payload_to_dict(model: BaseModel, exclude_unset: bool = False) -> Dict[str, Any]:
    try:
        return model.model_dump(exclude_unset=exclude_unset)
    except AttributeError:
        return model.dict(exclude_unset=exclude_unset)


def _row_value(row: Tuple[Any, ...], col_map: Dict[str, int], field: str) -> Any:
    index = col_map.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def _parse_version_sheet(rows: List[Tuple[Any, ...]]) -> Dict[Tuple[str, str, str], Dict[str, Any]]:
    if len(rows) < 2:
        return {}

    col_map = _build_header_map(list(rows[0]), VERSION_HEADER_ALIASES)
    bundles: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    for row in rows[1:]:
        product_code = _normalize_string(_row_value(row, col_map, "product_code"))
        if not product_code:
            continue
        version = _normalize_string(_row_value(row, col_map, "version")) or "v1.0"
        bom_type = _normalize_string(_row_value(row, col_map, "bom_type")) or "EBOM"
        key = (product_code, version, bom_type)
        bundles[key] = {
            "product_code": product_code,
            "product_name": _normalize_string(_row_value(row, col_map, "product_name")) or product_code,
            "version": version,
            "bom_type": bom_type,
            "status": _normalize_string(_row_value(row, col_map, "status")) or "DRAFT",
            "is_active": _parse_bool(_row_value(row, col_map, "is_active"), default=True),
            "product_family": _normalize_string(_row_value(row, col_map, "product_family")),
            "business_unit": _normalize_string(_row_value(row, col_map, "business_unit")),
            "project_code": _normalize_string(_row_value(row, col_map, "project_code")),
            "plant_code": _normalize_string(_row_value(row, col_map, "plant_code")),
            "discipline": _normalize_string(_row_value(row, col_map, "discipline")),
            "source_system": _normalize_string(_row_value(row, col_map, "source_system")) or "EXCEL",
            "source_file": _normalize_string(_row_value(row, col_map, "source_file")),
            "sync_status": _normalize_string(_row_value(row, col_map, "sync_status")) or "SYNCED",
            "cad_document_no": _normalize_string(_row_value(row, col_map, "cad_document_no")),
            "description": _normalize_string(_row_value(row, col_map, "description")),
        }
    return bundles


def _init_import_bundle(
    key: Tuple[str, str, str],
    version_bundles: Dict[Tuple[str, str, str], Dict[str, Any]],
    filename: str,
) -> Dict[str, Any]:
    product_code, version, bom_type = key
    base_header = dict(version_bundles.get(key, {}))
    base_header.setdefault("product_code", product_code)
    base_header.setdefault("product_name", product_code)
    base_header.setdefault("version", version)
    base_header.setdefault("bom_type", bom_type)
    base_header.setdefault("status", "DRAFT")
    base_header.setdefault("is_active", True)
    base_header.setdefault("source_system", "EXCEL")
    base_header.setdefault("source_file", filename)
    base_header.setdefault("sync_status", "SYNCED")
    return {"header": base_header, "items": []}


@router.get("/materials", response_model=List[MaterialRead])
async def list_materials(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return PLMService.get_materials(db, skip, limit, search)


@router.post("/materials", response_model=MaterialRead)
async def create_material(
    material_in: MaterialCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return PLMService.create_material(db, _payload_to_dict(material_in))


@router.patch("/materials/{material_id}", response_model=MaterialRead)
async def update_material(
    material_id: int,
    material_in: MaterialUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    result = PLMService.update_material(db, material_id, _payload_to_dict(material_in, exclude_unset=True))
    if not result:
        raise HTTPException(status_code=404, detail="Material not found")
    return result


@router.get("/boms", response_model=List[BOMHeaderRead])
async def list_boms(
    skip: int = 0,
    limit: int = 100,
    bom_type: Optional[str] = None,
    source_system: Optional[str] = None,
    project_code: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return PLMService.get_boms(db, skip, limit, bom_type, source_system, project_code)


@router.post("/boms", response_model=BOMDetailRead)
async def upsert_bom(
    bom_in: BOMUpsert,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = _payload_to_dict(bom_in)
    items = payload.pop("items", [])
    payload["released_by"] = getattr(current_user, "full_name", None) or getattr(current_user, "username", None)
    header = PLMService.upsert_bom(db, payload, items)
    result = PLMService.get_bom_detail(db, header.id)
    if not result:
        raise HTTPException(status_code=404, detail="BOM not found")
    return result


@router.patch("/boms/{bom_id}", response_model=BOMDetailRead)
async def update_bom_header(
    bom_id: int,
    bom_in: BOMUpsert,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = _payload_to_dict(bom_in)
    items = payload.pop("items", [])
    payload["id"] = bom_id
    payload["released_by"] = getattr(current_user, "full_name", None) or getattr(current_user, "username", None)
    header = PLMService.upsert_bom(db, payload, items)
    result = PLMService.get_bom_detail(db, header.id)
    if not result:
        raise HTTPException(status_code=404, detail="BOM not found")
    return result


@router.get("/boms/{bom_id}/detail", response_model=BOMDetailRead)
async def get_bom_detail(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    result = PLMService.get_bom_detail(db, bom_id)
    if not result:
        raise HTTPException(status_code=404, detail="BOM not found")
    return result


@router.post("/boms/{bom_id}/items", response_model=BOMDetailRead)
async def update_bom_items(
    bom_id: int,
    items: List[BOMItemUpsert],
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    try:
        PLMService.update_bom_items(db, bom_id, [_payload_to_dict(item) for item in items])
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    result = PLMService.get_bom_detail(db, bom_id)
    if not result:
        raise HTTPException(status_code=404, detail="BOM not found")
    return result


@router.post("/boms/{bom_id}/item-drawing-mappings/validate", response_model=BOMItemDrawingMappingResultRead)
async def validate_bom_item_drawing_mappings(
    bom_id: int,
    mappings: List[BOMItemDrawingMappingUpsert],
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return PLMService.validate_bom_item_drawing_mappings(
        db,
        bom_id,
        [_payload_to_dict(mapping) for mapping in mappings],
    )


@router.put("/boms/{bom_id}/item-drawing-mappings", response_model=BOMItemDrawingMappingResultRead)
async def save_bom_item_drawing_mappings(
    bom_id: int,
    mappings: List[BOMItemDrawingMappingUpsert],
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return PLMService.apply_bom_item_drawing_mappings(
        db,
        bom_id,
        [_payload_to_dict(mapping) for mapping in mappings],
    )


@router.get("/boms/{bom_id}/expand", response_model=BOMNodeRead)
async def get_bom_structure(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    result = PLMService.expand_bom(db, bom_id)
    if not result:
        raise HTTPException(status_code=404, detail="BOM not found")
    return result


@router.post("/boms/cad-sync", response_model=BOMDetailRead)
async def sync_bom_from_cad(
    sync_in: CADBOMSync,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = _payload_to_dict(sync_in)
    operator_name = getattr(current_user, "full_name", None) or getattr(current_user, "username", None)
    header = PLMService.sync_bom_from_cad(db, payload, operator_name=operator_name)
    result = PLMService.get_bom_detail(db, header.id)
    if not result:
        raise HTTPException(status_code=404, detail="BOM not found")
    return result


@router.get("/boms-export")
async def export_boms(
    bom_type: Optional[str] = None,
    source_system: Optional[str] = None,
    project_code: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from exc

    headers = PLMService.get_boms(db, 0, 10000, bom_type, source_system, project_code)

    workbook = Workbook()
    version_sheet = workbook.active
    version_sheet.title = VERSION_SHEET_NAME
    detail_sheet = workbook.create_sheet(DETAIL_SHEET_NAME)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="155E75", end_color="155E75", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, title in enumerate(VERSION_EXPORT_HEADERS, 1):
        cell = version_sheet.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for col_idx, title in enumerate(DETAIL_EXPORT_HEADERS, 1):
        cell = detail_sheet.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    detail_row = 2
    for version_row, header in enumerate(headers, 2):
        material = header.get("material")
        header_values = [
            header.get("product_code"),
            getattr(material, "name", None) if material else "",
            header.get("version"),
            header.get("bom_type"),
            header.get("status"),
            "是" if header.get("is_active") else "否",
            header.get("product_family"),
            header.get("business_unit"),
            header.get("project_code"),
            header.get("plant_code"),
            header.get("discipline"),
            header.get("source_system"),
            header.get("source_file"),
            header.get("sync_status"),
            header.get("cad_document_no"),
            header.get("description"),
        ]
        for col_idx, value in enumerate(header_values, 1):
            cell = version_sheet.cell(row=version_row, column=col_idx, value=value)
            cell.border = border

        detail = PLMService.get_bom_detail(db, header["id"])
        for item in detail["items"] if detail else []:
            item_material = item.get("material")
            item_values = [
                header.get("product_code"),
                header.get("version"),
                header.get("bom_type"),
                item.get("parent_item_code"),
                item.get("child_item_code"),
                getattr(item_material, "name", None) if item_material else "",
                getattr(item_material, "specification", None) if item_material else "",
                getattr(item_material, "unit", None) if item_material else "",
                item.get("item_level"),
                item.get("find_number"),
                item.get("quantity"),
                item.get("component_type"),
                item.get("item_category"),
                item.get("procurement_type"),
                item.get("routing_link"),
                item.get("loss_rate"),
                item.get("unit_price"),
                item.get("total_price"),
                getattr(item_material, "drawing_no", None) if item_material else "",
                getattr(item_material, "revision", None) if item_material else "",
                item.get("source_reference"),
            ]
            for col_idx, value in enumerate(item_values, 1):
                cell = detail_sheet.cell(row=detail_row, column=col_idx, value=value)
                cell.border = border
            detail_row += 1

    for sheet in (version_sheet, detail_sheet):
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 28)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"bom_master_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/boms-import")
async def import_boms(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xlsm 文件")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed") from exc

    workbook = load_workbook(BytesIO(await file.read()), read_only=True, data_only=True)

    version_rows: List[Tuple[Any, ...]] = []
    if VERSION_SHEET_NAME in workbook.sheetnames:
        version_rows = list(workbook[VERSION_SHEET_NAME].iter_rows(values_only=True))
    version_bundles = _parse_version_sheet(version_rows)

    if DETAIL_SHEET_NAME not in workbook.sheetnames:
        if VERSION_SHEET_NAME not in workbook.sheetnames:
            raise HTTPException(status_code=400, detail="Excel 中未找到 BOM版本 或 BOM明细 工作表")
        raise HTTPException(status_code=400, detail="Excel 中缺少 BOM明细 工作表")

    detail_rows = list(workbook[DETAIL_SHEET_NAME].iter_rows(values_only=True))
    if len(detail_rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 中没有可导入的 BOM 明细")

    col_map = _build_header_map(list(detail_rows[0]), DETAIL_HEADER_ALIASES)
    required = {"product_code", "child_item_code"}
    missing = required - set(col_map.keys())
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必要列：{', '.join(sorted(missing))}")

    grouped: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    errors: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(detail_rows[1:], start=2):
        try:
            product_code = _normalize_string(_row_value(row, col_map, "product_code"))
            child_item_code = _normalize_string(_row_value(row, col_map, "child_item_code"))
            if not product_code or not child_item_code:
                continue

            version = _normalize_string(_row_value(row, col_map, "version")) or "v1.0"
            bom_type = _normalize_string(_row_value(row, col_map, "bom_type")) or "EBOM"
            key = (product_code, version, bom_type)

            if key not in grouped:
                grouped[key] = _init_import_bundle(key, version_bundles, filename)

            grouped[key]["items"].append(
                {
                    "parent_item_code": _normalize_string(_row_value(row, col_map, "parent_item_code")) or product_code,
                    "child_item_code": child_item_code,
                    "material_name": _normalize_string(_row_value(row, col_map, "material_name")) or child_item_code,
                    "specification": _normalize_string(_row_value(row, col_map, "specification")),
                    "unit": _normalize_string(_row_value(row, col_map, "unit")) or "PCS",
                    "item_level": int(_parse_float(_row_value(row, col_map, "item_level"), 0))
                    if "item_level" in col_map
                    else None,
                    "find_number": _normalize_string(_row_value(row, col_map, "find_number")),
                    "quantity": _parse_float(_row_value(row, col_map, "quantity"), 1),
                    "component_type": _normalize_string(_row_value(row, col_map, "component_type")) or "NORMAL",
                    "item_category": _normalize_string(_row_value(row, col_map, "item_category")),
                    "procurement_type": _normalize_string(_row_value(row, col_map, "procurement_type")),
                    "routing_link": _normalize_string(_row_value(row, col_map, "routing_link")),
                    "loss_rate": _parse_float(_row_value(row, col_map, "loss_rate"), 0),
                    "unit_price": _parse_float(_row_value(row, col_map, "unit_price"), 0),
                    "total_price": _parse_float(_row_value(row, col_map, "total_price"), 0)
                    if "total_price" in col_map
                    else None,
                    "drawing_no": _normalize_string(_row_value(row, col_map, "drawing_no")),
                    "revision": _normalize_string(_row_value(row, col_map, "revision")),
                    "source_reference": _normalize_string(_row_value(row, col_map, "source_reference")),
                }
            )
        except Exception as exc:
            errors.append({"row": row_idx, "error": str(exc)})

    if not grouped:
        raise HTTPException(status_code=400, detail="未识别到有效的 BOM 明细数据")

    operator_name = getattr(current_user, "full_name", None) or getattr(current_user, "username", None)
    imported: List[Dict[str, Any]] = []

    for bundle in grouped.values():
        try:
            bundle["header"]["released_by"] = operator_name
            header = PLMService.upsert_bom(db, bundle["header"], bundle["items"])
            imported.append(
                {
                    "id": header.id,
                    "product_code": header.product_code,
                    "version": header.version,
                    "bom_type": header.bom_type,
                }
            )
        except Exception as exc:
            errors.append({"row": "-", "error": str(exc)})

    return {
        "imported": len(imported),
        "errors": len(errors),
        "items_total": sum(len(bundle["items"]) for bundle in grouped.values()),
        "boms": imported,
        "error_details": errors,
    }


@router.post("/ecns")
async def create_ecn(
    ecn_in: ECNCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    data = _payload_to_dict(ecn_in)
    data["creator_id"] = current_user.id
    impacts = data.pop("impacts")
    return PLMService.create_ecn(db, data, impacts)


@router.post("/ecns/{ecn_id}/approve")
async def approve_ecn(
    ecn_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    result = PLMService.approve_ecn(db, ecn_id, current_user.id)
    if not result:
        raise HTTPException(status_code=404, detail="ECN not found")
    return result


try:
    BOMNodeRead.model_rebuild()
except AttributeError:
    BOMNodeRead.update_forward_refs()
