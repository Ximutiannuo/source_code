from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_active_user
from app.models.order import ManufacturingOrder
from app.models.procurement import ProcurementRequest
from app.services.material_planning_service import MaterialPlanningService
from app.services.manufacturing_order_service import ManufacturingOrderService
from app.services.procurement_request_service import ProcurementRequestService

router = APIRouter()


class BOMReferenceRead(BaseModel):
    id: int
    product_code: str
    version: str
    bom_type: Optional[str] = None

    class Config:
        from_attributes = True


class ProcessTemplateReferenceRead(BaseModel):
    id: int
    name: str

    class Config:
        from_attributes = True


class EquipmentReferenceRead(BaseModel):
    id: int
    code: str
    name: str
    model_number: Optional[str] = None
    workstation: Optional[str] = None
    status: str

    class Config:
        from_attributes = True


class WorkReportRead(BaseModel):
    id: int
    operator_id: Optional[int] = None
    operator_name: Optional[str] = None
    quantity: float
    scrap_qty: float
    work_hours: float
    downtime_minutes: int
    report_type: str
    remarks: Optional[str] = None
    report_time: Optional[datetime] = None
    created_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class QualityCheckRead(BaseModel):
    id: int
    inspector_id: Optional[int] = None
    inspector_name: Optional[str] = None
    check_type: str
    result: str
    checked_qty: float
    defect_qty: float
    rework_qty: float
    remarks: Optional[str] = None
    checked_at: Optional[datetime] = None
    created_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class ProductionStepSummaryRead(BaseModel):
    id: int
    step_code: str
    name: str
    sort_order: int
    target_qty: Optional[float] = None
    completed_qty: Optional[float] = None
    planned_work_hours: Optional[float] = None
    setup_hours: Optional[float] = None
    workstation_id: Optional[int] = None
    workstation_name: Optional[str] = None
    status: str
    equipment: Optional[EquipmentReferenceRead] = None

    class Config:
        from_attributes = True


class ProductionStepDetailRead(ProductionStepSummaryRead):
    reports: List[WorkReportRead] = Field(default_factory=list)
    quality_checks: List[QualityCheckRead] = Field(default_factory=list)


class ManufacturingOrderCreate(BaseModel):
    order_number: str = Field(min_length=1, max_length=100)
    customer_name: Optional[str] = None
    product_name: Optional[str] = None
    bom_id: Optional[int] = None
    process_template_id: Optional[int] = None
    quantity: int = Field(default=1, ge=1)
    due_date: Optional[datetime] = None
    priority: int = Field(default=3, ge=1, le=5)
    status: str = "PLANNED"
    notes: Optional[str] = None
    auto_generate_steps: bool = True


class ManufacturingOrderUpdate(BaseModel):
    order_number: Optional[str] = Field(default=None, min_length=1, max_length=100)
    customer_name: Optional[str] = None
    product_name: Optional[str] = None
    bom_id: Optional[int] = None
    process_template_id: Optional[int] = None
    quantity: Optional[int] = Field(default=None, ge=1)
    due_date: Optional[datetime] = None
    priority: Optional[int] = Field(default=None, ge=1, le=5)
    status: Optional[str] = None
    notes: Optional[str] = None


class ProductionStepStatusUpdate(BaseModel):
    status: str
    completed_qty: Optional[float] = Field(default=None, ge=0)


class ProductionStepPlanningUpdate(BaseModel):
    workstation_id: Optional[int] = Field(default=None)
    planned_work_hours: Optional[float] = Field(default=None, ge=0)
    setup_hours: Optional[float] = Field(default=None, ge=0)


class WorkReportCreate(BaseModel):
    quantity: float = Field(default=0, ge=0)
    scrap_qty: float = Field(default=0, ge=0)
    work_hours: Optional[float] = Field(default=None, ge=0)
    downtime_minutes: int = Field(default=0, ge=0)
    report_type: str = "MANUAL"
    remarks: Optional[str] = None


class QualityCheckCreate(BaseModel):
    check_type: str = "IPQC"
    result: str
    checked_qty: float = Field(default=0, ge=0)
    defect_qty: float = Field(default=0, ge=0)
    rework_qty: float = Field(default=0, ge=0)
    remarks: Optional[str] = None


class ManufacturingOrderSummaryRead(BaseModel):
    id: int
    order_number: str
    customer_name: Optional[str] = None
    product_name: Optional[str] = None
    quantity: int
    due_date: Optional[datetime] = None
    priority: int
    status: str
    notes: Optional[str] = None
    created_at: Optional[datetime] = None
    bom: Optional[BOMReferenceRead] = None
    process_template: Optional[ProcessTemplateReferenceRead] = None
    steps: List[ProductionStepSummaryRead] = Field(default_factory=list)

    class Config:
        from_attributes = True


class ManufacturingOrderDetailRead(ManufacturingOrderSummaryRead):
    steps: List[ProductionStepDetailRead] = Field(default_factory=list)


class WIPSummaryRead(BaseModel):
    orders_total: int
    orders_planned: int
    orders_released: int
    orders_in_progress: int
    orders_qc: int
    orders_completed: int
    orders_cancelled: int
    steps_total: int
    steps_planned: int
    steps_ready: int
    steps_in_progress: int
    steps_qc: int
    steps_completed: int
    steps_blocked: int
    quality_pass_count: int
    quality_fail_count: int
    quality_rework_count: int
    quality_hold_count: int
    rework_qty_total: float
    defect_qty_total: float
    planned_hours_total: float
    reported_hours_total: float
    downtime_minutes_total: int
    equipment_total: int
    equipment_active: int
    equipment_maintenance: int
    equipment_offline: int
    equipment_assigned: int
    overall_oee_rate: float
    reports_total: int
    reports_today: int


class EquipmentOEERead(BaseModel):
    id: int
    code: str
    name: str
    model_number: Optional[str] = None
    workstation: Optional[str] = None
    status: str
    assigned_steps: int
    orders_count: int
    planned_hours_total: float
    actual_hours_total: float
    runtime_hours_total: float
    theoretical_hours_total: float
    downtime_minutes_total: int
    good_qty_total: float
    scrap_qty_total: float
    availability_rate: float
    performance_rate: float
    quality_rate: float
    utilization_rate: float
    oee_rate: float


class EquipmentOEESummaryRead(BaseModel):
    equipment_total: int
    equipment_active: int
    equipment_maintenance: int
    equipment_offline: int
    equipment_assigned: int
    planned_hours_total: float
    actual_hours_total: float
    runtime_hours_total: float
    downtime_minutes_total: int
    overall_availability_rate: float
    overall_performance_rate: float
    overall_quality_rate: float
    overall_oee_rate: float
    items: List[EquipmentOEERead] = Field(default_factory=list)


class MaterialReadinessItemRead(BaseModel):
    material_code: str
    material_name: str
    unit: str
    material_type: Optional[str] = None
    material_category: Optional[str] = None
    lead_time_days: int = 0
    required_qty: float
    current_stock: float
    reserved_stock: float
    incoming_stock: float
    available_qty: float
    net_available_qty: float
    safety_stock: float
    shortage_qty: float
    shortage_with_safety_qty: float
    readiness_status: str
    shortage_reason: str
    impacted_order_count: int = 0
    impacted_orders: List[str] = Field(default_factory=list)


class OrderMaterialReadinessRead(BaseModel):
    order_id: int
    order_number: str
    bom_id: Optional[int] = None
    bom_version: Optional[str] = None
    kit_status: str
    required_items_total: int
    ready_items: int
    risk_items: int
    short_items: int
    shortage_qty_total: float
    kit_rate: float
    items: List[MaterialReadinessItemRead] = Field(default_factory=list)


class MaterialPlanningSummaryRead(BaseModel):
    orders_considered: int
    orders_without_bom: int
    materials_total: int
    ready_materials: int
    risk_materials: int
    short_materials: int
    shortage_qty_total: float
    impacted_orders: int
    items: List[MaterialReadinessItemRead] = Field(default_factory=list)


class ProcurementSuggestionItemRead(BaseModel):
    material_code: str
    material_name: str
    unit: str
    material_type: Optional[str] = None
    material_category: Optional[str] = None
    readiness_status: str
    shortage_reason: str
    procurement_mode: str
    suggested_action: str
    suggested_purchase_qty: float
    shortage_qty: float
    shortage_with_safety_qty: float
    current_stock: float
    reserved_stock: float
    incoming_stock: float
    net_available_qty: float
    safety_stock: float
    lead_time_days: int = 0
    earliest_due_date: Optional[datetime] = None
    suggested_order_date: Optional[datetime] = None
    urgency_level: str
    planning_note: str
    impacted_order_count: int = 0
    impacted_orders: List[str] = Field(default_factory=list)


class ProcurementSuggestionSummaryRead(BaseModel):
    orders_considered: int
    orders_without_bom: int
    items_total: int
    urgent_items: int
    high_items: int
    to_purchase_items: int
    to_expedite_items: int
    master_data_gap_items: int
    replenish_items: int
    suggested_purchase_qty_total: float
    impacted_orders: int
    items: List[ProcurementSuggestionItemRead] = Field(default_factory=list)


class OrderProcurementSuggestionRead(BaseModel):
    order_id: int
    order_number: str
    kit_status: str
    items_total: int
    urgent_items: int
    high_items: int
    to_purchase_items: int
    to_expedite_items: int
    master_data_gap_items: int
    replenish_items: int
    suggested_purchase_qty_total: float
    impacted_orders: int
    items: List[ProcurementSuggestionItemRead] = Field(default_factory=list)


class ProcurementRequestItemRead(BaseModel):
    id: int
    material_code: str
    material_name: str
    unit: Optional[str] = None
    material_type: Optional[str] = None
    material_category: Optional[str] = None
    readiness_status: Optional[str] = None
    shortage_reason: Optional[str] = None
    procurement_mode: Optional[str] = None
    suggested_action: Optional[str] = None
    urgency_level: Optional[str] = None
    requested_qty: float
    shortage_qty: float
    shortage_with_safety_qty: float
    current_stock: float
    reserved_stock: float
    incoming_stock: float
    net_available_qty: float
    safety_stock: float
    lead_time_days: int = 0
    earliest_due_date: Optional[datetime] = None
    suggested_order_date: Optional[datetime] = None
    impacted_order_count: int = 0
    impacted_orders: List[str] = Field(default_factory=list)
    planning_note: Optional[str] = None

    class Config:
        from_attributes = True


class ProcurementRequestRead(BaseModel):
    id: int
    request_no: str
    title: str
    source_scope: str
    source_order_id: Optional[int] = None
    source_order_number: Optional[str] = None
    status: str
    urgency_level: str
    total_items: int
    suggested_purchase_qty_total: float
    requester_id: Optional[int] = None
    requester_name: Optional[str] = None
    notes: Optional[str] = None
    submitted_at: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    items: List[ProcurementRequestItemRead] = Field(default_factory=list)

    class Config:
        from_attributes = True


class ProcurementRequestGenerate(BaseModel):
    source_scope: str = "GLOBAL"
    order_id: Optional[int] = None
    material_codes: List[str] = Field(default_factory=list)
    title: Optional[str] = None
    notes: Optional[str] = None


class ProcurementRequestStatusUpdate(BaseModel):
    status: str


class ProcurementRequestUpdate(BaseModel):
    title: Optional[str] = None
    source_scope: Optional[str] = None
    status: Optional[str] = None
    urgency_level: Optional[str] = None
    requester_name: Optional[str] = None
    notes: Optional[str] = None


def _normalize_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _parse_datetime_value(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    candidates = [text, text[:19], text[:16], text[:10]]
    for candidate in candidates:
        if not candidate:
            continue
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def _parse_priority_value(value: Any) -> int:
    if value in (None, ""):
        return 3
    if isinstance(value, str):
        normalized = value.strip().upper()
        priority_map = {
            "P1": 1,
            "P1 HIGHEST": 1,
            "P2": 2,
            "P3": 3,
            "P4": 4,
            "P5": 5,
            "P1 最高": 1,
            "P2 高": 2,
            "P3 标准": 3,
            "P4 低": 4,
            "P5 最低": 5,
        }
        if normalized in priority_map:
            return priority_map[normalized]
        if normalized.startswith("P") and len(normalized) >= 2 and normalized[1].isdigit():
            return max(1, min(5, int(normalized[1])))
    try:
        return max(1, min(5, int(float(value))))
    except (TypeError, ValueError):
        return 3


def _parse_order_status_value(value: Any) -> str:
    if value in (None, ""):
        return "PLANNED"
    normalized = str(value).strip().upper()
    status_aliases = {
        "PLANNED": "PLANNED",
        "RELEASED": "RELEASED",
        "IN_PROGRESS": "IN_PROGRESS",
        "QC": "QC",
        "COMPLETED": "COMPLETED",
        "CANCELLED": "CANCELLED",
        "PLAN": "PLANNED",
        "IN PROGRESS": "IN_PROGRESS",
        "计划中": "PLANNED",
        "已下达": "RELEASED",
        "生产中": "IN_PROGRESS",
        "质检中": "QC",
        "已完成": "COMPLETED",
        "已取消": "CANCELLED",
    }
    return status_aliases.get(normalized, normalized if normalized else "PLANNED")


def _parse_procurement_status_value(value: Any) -> str:
    if value in (None, ""):
        return "DRAFT"
    normalized = str(value).strip().upper()
    status_aliases = {
        "DRAFT": "DRAFT",
        "SUBMITTED": "SUBMITTED",
        "IN_PROGRESS": "IN_PROGRESS",
        "ORDERED": "ORDERED",
        "RECEIVED": "RECEIVED",
        "CANCELLED": "CANCELLED",
        "IN PROGRESS": "IN_PROGRESS",
        "草稿": "DRAFT",
        "已提交": "SUBMITTED",
        "处理中": "IN_PROGRESS",
        "已下单": "ORDERED",
        "已到货": "RECEIVED",
        "已取消": "CANCELLED",
    }
    return status_aliases.get(normalized, normalized if normalized else "DRAFT")


def _parse_urgency_value(value: Any) -> str:
    if value in (None, ""):
        return "MEDIUM"
    normalized = str(value).strip().upper()
    urgency_aliases = {
        "URGENT": "URGENT",
        "HIGH": "HIGH",
        "MEDIUM": "MEDIUM",
        "LOW": "LOW",
        "紧急": "URGENT",
        "高": "HIGH",
        "中": "MEDIUM",
        "低": "LOW",
    }
    return urgency_aliases.get(normalized, "MEDIUM")


def _build_header_map(headers: List[Any], aliases: Dict[str, List[str]]) -> Dict[str, int]:
    normalized_headers = {str(value or "").strip(): index for index, value in enumerate(headers)}
    result: Dict[str, int] = {}
    for field, names in aliases.items():
        for name in names:
            if name in normalized_headers:
                result[field] = normalized_headers[name]
                break
    return result


@router.get("/orders", response_model=List[ManufacturingOrderSummaryRead])
async def list_manufacturing_orders(
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return ManufacturingOrderService.list_orders(db, status=status, keyword=keyword)


@router.get("/orders/{order_id}", response_model=ManufacturingOrderDetailRead)
async def get_manufacturing_order(
    order_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    order = ManufacturingOrderService.get_order(
        db,
        order_id,
        include_reports=True,
        include_quality=True,
    )
    if not order:
        raise HTTPException(status_code=404, detail="Manufacturing order not found")
    return order


@router.patch("/orders/{order_id}", response_model=ManufacturingOrderDetailRead)
async def update_manufacturing_order(
    order_id: int,
    order_in: ManufacturingOrderUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = order_in.dict(exclude_unset=True)
    try:
        return ManufacturingOrderService.update_order(db=db, order_id=order_id, payload=payload)
    except ValueError as exc:
        detail = str(exc)
        status_code = 404 if "not found" in detail.lower() else 400
        raise HTTPException(status_code=status_code, detail=detail) from exc


@router.get("/orders/{order_id}/material-readiness", response_model=OrderMaterialReadinessRead)
async def get_order_material_readiness(
    order_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    result = MaterialPlanningService.get_order_material_readiness(db, order_id)
    if not result:
        raise HTTPException(status_code=404, detail="Manufacturing order not found")
    return result


@router.get("/orders/{order_id}/procurement-suggestions", response_model=OrderProcurementSuggestionRead)
async def get_order_procurement_suggestions(
    order_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    result = MaterialPlanningService.get_order_procurement_suggestions(db, order_id)
    if not result:
        raise HTTPException(status_code=404, detail="Manufacturing order not found")
    return result


@router.get("/wip-summary", response_model=WIPSummaryRead)
async def get_wip_summary(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return ManufacturingOrderService.get_wip_summary(db)


@router.get("/equipment-oee", response_model=EquipmentOEESummaryRead)
async def get_equipment_oee_summary(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return ManufacturingOrderService.get_equipment_oee_summary(db)


@router.get("/material-planning-summary", response_model=MaterialPlanningSummaryRead)
async def get_material_planning_summary(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return MaterialPlanningService.get_material_planning_summary(db)


@router.get("/procurement-suggestions", response_model=ProcurementSuggestionSummaryRead)
async def get_material_procurement_suggestions(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return MaterialPlanningService.get_material_procurement_suggestions(db)


@router.get("/procurement-requests", response_model=List[ProcurementRequestRead])
async def list_procurement_requests(
    status: Optional[str] = None,
    order_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    return ProcurementRequestService.list_requests(db, status=status, order_id=order_id)


@router.post("/procurement-requests/generate", response_model=ProcurementRequestRead)
async def generate_procurement_request(
    request_in: ProcurementRequestGenerate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = request_in.dict()
    try:
        return ProcurementRequestService.create_from_suggestions(
            db=db,
            current_user=current_user,
            source_scope=payload.get("source_scope", "GLOBAL"),
            order_id=payload.get("order_id"),
            material_codes=payload.get("material_codes", []),
            title=payload.get("title"),
            notes=payload.get("notes"),
        )
    except ValueError as exc:
        detail = str(exc)
        status_code = 404 if "not found" in detail.lower() else 400
        raise HTTPException(status_code=status_code, detail=detail) from exc


@router.patch("/procurement-requests/{request_id}/status", response_model=ProcurementRequestRead)
async def update_procurement_request_status(
    request_id: int,
    request_in: ProcurementRequestStatusUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    try:
        result = ProcurementRequestService.update_status(db, request_id, request_in.status)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not result:
        raise HTTPException(status_code=404, detail="Procurement request not found")
    return result


@router.patch("/procurement-requests/{request_id}", response_model=ProcurementRequestRead)
async def update_procurement_request(
    request_id: int,
    request_in: ProcurementRequestUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = request_in.dict(exclude_unset=True)
    try:
        result = ProcurementRequestService.update_request(db, request_id, payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not result:
        raise HTTPException(status_code=404, detail="Procurement request not found")
    return result


@router.post("/orders", response_model=ManufacturingOrderDetailRead)
async def create_manufacturing_order(
    order_in: ManufacturingOrderCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = order_in.dict()
    auto_generate_steps = payload.pop("auto_generate_steps", True)
    try:
        return ManufacturingOrderService.create_order(
            db=db,
            payload=payload,
            auto_generate_steps=auto_generate_steps,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.patch("/orders/{order_id}/steps/{step_id}/status", response_model=ProductionStepDetailRead)
async def update_production_step_status(
    order_id: int,
    step_id: int,
    step_in: ProductionStepStatusUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    try:
        return ManufacturingOrderService.update_step_status(
            db=db,
            order_id=order_id,
            step_id=step_id,
            status=step_in.status,
            completed_qty=step_in.completed_qty,
        )
    except ValueError as exc:
        detail = str(exc)
        status_code = 404 if "not found" in detail.lower() else 400
        raise HTTPException(status_code=status_code, detail=detail) from exc


@router.patch("/orders/{order_id}/steps/{step_id}/planning", response_model=ProductionStepDetailRead)
async def update_production_step_planning(
    order_id: int,
    step_id: int,
    step_in: ProductionStepPlanningUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    try:
        return ManufacturingOrderService.update_step_planning(
            db=db,
            order_id=order_id,
            step_id=step_id,
            workstation_id=step_in.workstation_id,
            planned_work_hours=step_in.planned_work_hours,
            setup_hours=step_in.setup_hours,
        )
    except ValueError as exc:
        detail = str(exc)
        status_code = 404 if "not found" in detail.lower() else 400
        raise HTTPException(status_code=status_code, detail=detail) from exc


@router.post("/orders/{order_id}/steps/{step_id}/reports", response_model=WorkReportRead)
async def create_production_step_report(
    order_id: int,
    step_id: int,
    report_in: WorkReportCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = report_in.dict()
    try:
        return ManufacturingOrderService.create_work_report(
            db=db,
            order_id=order_id,
            step_id=step_id,
            operator_id=current_user.id,
            quantity=payload["quantity"],
            scrap_qty=payload.get("scrap_qty", 0),
            work_hours=payload.get("work_hours"),
            downtime_minutes=payload.get("downtime_minutes", 0),
            report_type=payload.get("report_type", "MANUAL"),
            remarks=payload.get("remarks"),
        )
    except ValueError as exc:
        detail = str(exc)
        status_code = 404 if "not found" in detail.lower() else 400
        raise HTTPException(status_code=status_code, detail=detail) from exc


@router.post("/orders/{order_id}/steps/{step_id}/quality-checks", response_model=QualityCheckRead)
async def create_production_step_quality_check(
    order_id: int,
    step_id: int,
    quality_in: QualityCheckCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    payload = quality_in.dict()
    try:
        return ManufacturingOrderService.create_quality_check(
            db=db,
            order_id=order_id,
            step_id=step_id,
            inspector_id=current_user.id,
            result=payload["result"],
            checked_qty=payload.get("checked_qty", 0),
            defect_qty=payload.get("defect_qty", 0),
            rework_qty=payload.get("rework_qty", 0),
            check_type=payload.get("check_type", "IPQC"),
            remarks=payload.get("remarks"),
        )
    except ValueError as exc:
        detail = str(exc)
        status_code = 404 if "not found" in detail.lower() else 400
        raise HTTPException(status_code=status_code, detail=detail) from exc


@router.get("/orders-export")
async def export_manufacturing_orders(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """Export manufacturing orders to Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    orders = ManufacturingOrderService.list_orders(db, status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = "制造订单"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    STATUS_MAP = {
        "PLANNED": "计划中", "RELEASED": "已下达", "IN_PROGRESS": "生产中",
        "QC": "质检中", "COMPLETED": "已完成", "CANCELLED": "已取消",
    }
    PRIORITY_MAP = {1: "P1 最高", 2: "P2 高", 3: "P3 标准", 4: "P4 低", 5: "P5 最低"}

    headers = [
        "订单号", "客户名称", "产品名称", "数量", "优先级", "状态",
        "交期", "BOM编码", "BOM版本", "工艺模板", "工序数",
        "已完成工序", "进度(%)", "备注", "创建时间",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, order in enumerate(orders, 2):
        total_steps = len(order.steps) if order.steps else 0
        completed_steps = sum(1 for s in (order.steps or []) if s.status == "COMPLETED")
        progress = round(completed_steps / total_steps * 100) if total_steps > 0 else 0

        row_data = [
            order.order_number,
            order.customer_name or "",
            order.product_name or "",
            order.quantity,
            PRIORITY_MAP.get(order.priority, str(order.priority)),
            STATUS_MAP.get(order.status, order.status),
            order.due_date.strftime("%Y-%m-%d %H:%M") if order.due_date else "",
            order.bom.product_code if order.bom else "",
            order.bom.version if order.bom else "",
            order.process_template.name if order.process_template else "",
            total_steps,
            completed_steps,
            progress,
            order.notes or "",
            order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [16, 16, 20, 8, 12, 10, 18, 14, 10, 16, 8, 10, 10, 24, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Steps sheet
    ws2 = wb.create_sheet("工序明细")
    step_headers = [
        "订单号", "工序代码", "工序名称", "排序", "目标数量", "已完成数量",
        "计划工时", "准备工时", "设备", "状态",
    ]
    STEP_STATUS_MAP = {
        "PLANNED": "计划中", "READY": "待开工", "IN_PROGRESS": "生产中",
        "QC": "待质检", "COMPLETED": "已完成", "BLOCKED": "已阻塞",
    }
    for col_idx, header in enumerate(step_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    step_row = 2
    for order in orders:
        for step in (order.steps or []):
            step_data = [
                order.order_number,
                step.step_code,
                step.name,
                step.sort_order,
                float(step.target_qty or 0),
                float(step.completed_qty or 0),
                float(step.planned_work_hours or 0),
                float(step.setup_hours or 0),
                step.equipment.name if step.equipment else "",
                STEP_STATUS_MAP.get(step.status, step.status),
            ]
            for col_idx, value in enumerate(step_data, 1):
                cell = ws2.cell(row=step_row, column=col_idx, value=value)
                cell.border = thin_border
            step_row += 1

    step_col_widths = [16, 14, 20, 8, 10, 10, 10, 10, 16, 10]
    for col_idx, width in enumerate(step_col_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"manufacturing_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/orders-import")
async def import_manufacturing_orders(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """Import manufacturing orders from Excel (.xlsx). Expects columns: 订单号, 客户名称, 产品名称, 数量, 优先级(1-5), 交期(YYYY-MM-DD), 备注."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式文件")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="文件中无数据行")

    header_row = [str(c or "").strip() for c in rows[0]]
    col_map = {}
    expected_cols = {
        "订单号": "order_number",
        "客户名称": "customer_name",
        "产品名称": "product_name",
        "数量": "quantity",
        "优先级": "priority",
        "交期": "due_date",
        "备注": "notes",
    }
    for idx, col_name in enumerate(header_row):
        if col_name in expected_cols:
            col_map[expected_cols[col_name]] = idx

    if "order_number" not in col_map:
        raise HTTPException(status_code=400, detail="缺少必须列：订单号")

    PRIORITY_REVERSE = {"P1 最高": 1, "P2 高": 2, "P3 标准": 3, "P4 低": 4, "P5 最低": 5}

    created = []
    errors = []
    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            order_number = str(row[col_map["order_number"]] or "").strip()
            if not order_number:
                continue

            customer_name = str(row[col_map.get("customer_name", 999)] or "").strip() if "customer_name" in col_map else None
            product_name = str(row[col_map.get("product_name", 999)] or "").strip() if "product_name" in col_map else None

            raw_qty = row[col_map["quantity"]] if "quantity" in col_map else 1
            quantity = int(float(raw_qty or 1)) if raw_qty else 1
            if quantity < 1:
                quantity = 1

            raw_priority = row[col_map["priority"]] if "priority" in col_map else 3
            if isinstance(raw_priority, str):
                priority = PRIORITY_REVERSE.get(raw_priority.strip(), 3)
            else:
                priority = int(raw_priority or 3)
            priority = max(1, min(5, priority))

            due_date = None
            if "due_date" in col_map:
                raw_due = row[col_map["due_date"]]
                if isinstance(raw_due, datetime):
                    due_date = raw_due
                elif raw_due:
                    try:
                        due_date = datetime.strptime(str(raw_due).strip()[:10], "%Y-%m-%d")
                    except ValueError:
                        pass

            notes = str(row[col_map.get("notes", 999)] or "").strip() if "notes" in col_map else None

            payload = {
                "order_number": order_number,
                "customer_name": customer_name or None,
                "product_name": product_name or None,
                "quantity": quantity,
                "priority": priority,
                "due_date": due_date,
                "notes": notes or None,
                "status": "PLANNED",
            }
            order = ManufacturingOrderService.create_order(db, payload, auto_generate_steps=False)
            created.append({"row": row_idx, "order_number": order.order_number, "id": order.id})
        except Exception as exc:
            errors.append({"row": row_idx, "error": str(exc)})

    return {
        "imported": len(created),
        "errors": len(errors),
        "created": created,
        "error_details": errors,
    }
