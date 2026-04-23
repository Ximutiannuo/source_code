from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.manufacturing_orders import (
    _build_header_map,
    _normalize_string,
    _parse_datetime_value,
    _parse_order_status_value,
    _parse_priority_value,
    _parse_procurement_status_value,
    _parse_urgency_value,
)
from app.database import get_db
from app.dependencies import get_current_active_user
from app.models.procurement import ProcurementRequest
from app.models.order import ManufacturingOrder
from app.models.production_step import ProductionStep
from app.services.manufacturing_order_service import ManufacturingOrderService
from app.services.procurement_request_service import ProcurementRequestService
from app.utils.timezone import now as system_now

router = APIRouter()


def _parse_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _parse_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _parse_progress_percent(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        normalized = value.strip().replace("%", "")
    else:
        normalized = value
    try:
        progress = float(normalized)
    except (TypeError, ValueError):
        return None
    if 0 < progress <= 1:
        progress *= 100
    return max(0.0, min(100.0, progress))


def _parse_list(value: Any) -> List[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    raw = str(value).replace(";", ",")
    return [item.strip() for item in raw.split(",") if item.strip()]


def _apply_excel_header_style(ws, headers: List[str]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _sync_order_progress_from_import(
    db: Session,
    order: ManufacturingOrder,
    progress_percent: Optional[float],
    has_explicit_status: bool = False,
) -> None:
    if progress_percent is None:
        return

    ratio = Decimal(str(progress_percent)) / Decimal("100")
    quantity = max(1, int(order.quantity or 1))
    steps = list(order.steps or [])

    if not steps:
        step = ProductionStep(
            order_id=order.id,
            step_code="IMPORT_PROGRESS",
            name="导入进度",
            sort_order=0,
            target_qty=Decimal(str(quantity)),
            completed_qty=Decimal("0"),
            planned_work_hours=Decimal("0"),
            setup_hours=Decimal("0"),
            status="PLANNED",
        )
        db.add(step)
        db.flush()
        steps = [step]

    for step in steps:
        target_qty = Decimal(str(step.target_qty or quantity))
        if target_qty <= 0:
            target_qty = Decimal(str(quantity))
            step.target_qty = target_qty

        completed_qty = (target_qty * ratio).quantize(Decimal("0.01"))
        step.completed_qty = max(Decimal("0"), min(target_qty, completed_qty))

        if progress_percent >= 100:
            step.status = "COMPLETED"
        elif progress_percent <= 0:
            step.status = "PLANNED"
        else:
            step.status = "IN_PROGRESS"

    if not has_explicit_status and order.status != "CANCELLED":
        if progress_percent >= 100:
            order.status = "COMPLETED"
        elif progress_percent > 0:
            order.status = "IN_PROGRESS"
        else:
            order.status = "PLANNED"

    db.commit()


@router.post("/orders-import-upsert")
async def import_manufacturing_orders_upsert(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are supported")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="No data rows found in workbook")

    col_map = _build_header_map(
        list(rows[0]),
        {
            "order_number": ["订单号", "工单号", "Order Number", "order_number"],
            "customer_name": ["客户名称", "客户", "Customer", "customer_name"],
            "product_name": ["产品名称", "产品", "Product Name", "product_name"],
            "quantity": ["数量", "Quantity", "quantity"],
            "priority": ["优先级", "Priority", "priority"],
            "status": ["状态", "订单状态", "Status", "status"],
            "progress_percent": ["进度(%)", "执行进度", "进度", "Progress", "progress_percent"],
            "due_date": ["交期", "计划交期", "Due Date", "due_date"],
            "notes": ["备注", "Notes", "notes"],
        },
    )
    if "order_number" not in col_map:
        raise HTTPException(status_code=400, detail="Missing required column: order_number")

    created = []
    updated = []
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            order_number = _normalize_string(row[col_map["order_number"]]) if col_map["order_number"] < len(row) else None
            if not order_number:
                continue

            payload = {
                "order_number": order_number,
                "customer_name": _normalize_string(row[col_map["customer_name"]]) if "customer_name" in col_map and col_map["customer_name"] < len(row) else None,
                "product_name": _normalize_string(row[col_map["product_name"]]) if "product_name" in col_map and col_map["product_name"] < len(row) else None,
                "quantity": max(1, _parse_int(row[col_map["quantity"]])) if "quantity" in col_map and col_map["quantity"] < len(row) else 1,
                "priority": _parse_priority_value(row[col_map["priority"]]) if "priority" in col_map and col_map["priority"] < len(row) else 3,
                "status": _parse_order_status_value(row[col_map["status"]]) if "status" in col_map and col_map["status"] < len(row) else "PLANNED",
                "due_date": _parse_datetime_value(row[col_map["due_date"]]) if "due_date" in col_map and col_map["due_date"] < len(row) else None,
                "notes": _normalize_string(row[col_map["notes"]]) if "notes" in col_map and col_map["notes"] < len(row) else None,
            }
            progress_percent = (
                _parse_progress_percent(row[col_map["progress_percent"]])
                if "progress_percent" in col_map and col_map["progress_percent"] < len(row)
                else None
            )
            has_explicit_status = "status" in col_map and col_map["status"] < len(row) and row[col_map["status"]] not in (None, "")

            existing = db.query(ManufacturingOrder).filter(ManufacturingOrder.order_number == order_number).first()
            if existing:
                order = ManufacturingOrderService.update_order(db, existing.id, payload)
                _sync_order_progress_from_import(db, order, progress_percent, has_explicit_status=has_explicit_status)
                updated.append({"row": row_idx, "order_number": order.order_number, "id": order.id})
            else:
                order = ManufacturingOrderService.create_order(db, payload, auto_generate_steps=False)
                _sync_order_progress_from_import(db, order, progress_percent, has_explicit_status=has_explicit_status)
                created.append({"row": row_idx, "order_number": order.order_number, "id": order.id})
        except Exception as exc:
            errors.append({"row": row_idx, "error": str(exc)})

    return {
        "imported": len(created) + len(updated),
        "errors": len(errors),
        "created": created,
        "updated": updated,
        "error_details": errors,
    }


@router.get("/procurement-requests-export")
async def export_procurement_requests(
    status: Optional[str] = None,
    order_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    requests = ProcurementRequestService.list_requests(db, status=status, order_id=order_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "采购订单"
    request_headers = [
        "请购单号",
        "标题",
        "来源范围",
        "来源订单ID",
        "来源订单号",
        "状态",
        "紧急程度",
        "物料项数",
        "建议采购总量",
        "申请人",
        "备注",
        "提交时间",
        "完成时间",
        "创建时间",
        "更新时间",
    ]
    _apply_excel_header_style(ws, request_headers)

    for row_idx, request in enumerate(requests, start=2):
        row_data = [
            request.request_no,
            request.title,
            request.source_scope,
            request.source_order_id or "",
            request.source_order_number or "",
            request.status,
            request.urgency_level,
            request.total_items,
            request.suggested_purchase_qty_total,
            request.requester_name or "",
            request.notes or "",
            request.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if request.submitted_at else "",
            request.completed_at.strftime("%Y-%m-%d %H:%M:%S") if request.completed_at else "",
            request.created_at.strftime("%Y-%m-%d %H:%M:%S") if request.created_at else "",
            request.updated_at.strftime("%Y-%m-%d %H:%M:%S") if request.updated_at else "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, width in enumerate([18, 24, 12, 12, 16, 12, 12, 10, 14, 12, 26, 20, 20, 20, 20], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    item_ws = wb.create_sheet("采购明细")
    item_headers = [
        "请购单号",
        "物料编码",
        "物料名称",
        "单位",
        "物料类型",
        "物料分类",
        "齐套状态",
        "缺料原因",
        "采购模式",
        "建议动作",
        "紧急程度",
        "请购数量",
        "缺口数量",
        "含安全库存缺口",
        "现有库存",
        "预留库存",
        "在途库存",
        "净可用量",
        "安全库存",
        "提前期天数",
        "最早需求日期",
        "建议下单日期",
        "影响订单数",
        "影响订单",
        "计划备注",
    ]
    _apply_excel_header_style(item_ws, item_headers)

    row_idx = 2
    for request in requests:
        for item in request.items or []:
            row_data = [
                request.request_no,
                item.material_code,
                item.material_name,
                item.unit or "",
                item.material_type or "",
                item.material_category or "",
                item.readiness_status or "",
                item.shortage_reason or "",
                item.procurement_mode or "",
                item.suggested_action or "",
                item.urgency_level or "",
                item.requested_qty,
                item.shortage_qty,
                item.shortage_with_safety_qty,
                item.current_stock,
                item.reserved_stock,
                item.incoming_stock,
                item.net_available_qty,
                item.safety_stock,
                item.lead_time_days,
                item.earliest_due_date.strftime("%Y-%m-%d %H:%M:%S") if item.earliest_due_date else "",
                item.suggested_order_date.strftime("%Y-%m-%d %H:%M:%S") if item.suggested_order_date else "",
                item.impacted_order_count,
                ",".join(item.impacted_orders or []),
                item.planning_note or "",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                item_ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    for col_idx, width in enumerate([18, 16, 22, 10, 12, 12, 12, 16, 14, 12, 12, 12, 12, 14, 12, 12, 12, 12, 12, 12, 18, 18, 10, 20, 24], start=1):
        item_ws.column_dimensions[item_ws.cell(row=1, column=col_idx).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"procurement_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/procurement-requests-import")
async def import_procurement_requests(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are supported")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), read_only=True)
    request_ws = wb["采购订单"] if "采购订单" in wb.sheetnames else wb.active
    item_ws = wb["采购明细"] if "采购明细" in wb.sheetnames else None

    request_rows = list(request_ws.iter_rows(min_row=1, values_only=True))
    if len(request_rows) < 2:
        raise HTTPException(status_code=400, detail="No procurement request rows found in workbook")

    request_col_map = _build_header_map(
        list(request_rows[0]),
        {
            "request_no": ["请购单号", "Request No", "request_no"],
            "title": ["标题", "Title", "title"],
            "source_scope": ["来源范围", "Scope", "source_scope"],
            "source_order_id": ["来源订单ID", "Source Order ID", "source_order_id"],
            "status": ["状态", "Status", "status"],
            "urgency_level": ["紧急程度", "Urgency", "urgency_level"],
            "requester_name": ["申请人", "Requester", "requester_name"],
            "notes": ["备注", "Notes", "notes"],
            "submitted_at": ["提交时间", "Submitted At", "submitted_at"],
            "completed_at": ["完成时间", "Completed At", "completed_at"],
            "created_at": ["创建时间", "Created At", "created_at"],
            "updated_at": ["更新时间", "Updated At", "updated_at"],
        },
    )
    if "request_no" not in request_col_map:
        raise HTTPException(status_code=400, detail="Missing required column: request_no")

    request_items_map: Dict[str, List[Dict[str, Any]]] = {}
    if item_ws is not None:
        item_rows = list(item_ws.iter_rows(min_row=1, values_only=True))
        if len(item_rows) >= 2:
            item_col_map = _build_header_map(
                list(item_rows[0]),
                {
                    "request_no": ["请购单号", "Request No", "request_no"],
                    "material_code": ["物料编码", "Material Code", "material_code"],
                    "material_name": ["物料名称", "Material Name", "material_name"],
                    "unit": ["单位", "Unit", "unit"],
                    "material_type": ["物料类型", "Material Type", "material_type"],
                    "material_category": ["物料分类", "Material Category", "material_category"],
                    "readiness_status": ["齐套状态", "Readiness Status", "readiness_status"],
                    "shortage_reason": ["缺料原因", "Shortage Reason", "shortage_reason"],
                    "procurement_mode": ["采购模式", "Procurement Mode", "procurement_mode"],
                    "suggested_action": ["建议动作", "Suggested Action", "suggested_action"],
                    "urgency_level": ["紧急程度", "Urgency", "urgency_level"],
                    "requested_qty": ["请购数量", "Requested Qty", "requested_qty"],
                    "shortage_qty": ["缺口数量", "Shortage Qty", "shortage_qty"],
                    "shortage_with_safety_qty": ["含安全库存缺口", "Shortage With Safety Qty", "shortage_with_safety_qty"],
                    "current_stock": ["现有库存", "Current Stock", "current_stock"],
                    "reserved_stock": ["预留库存", "Reserved Stock", "reserved_stock"],
                    "incoming_stock": ["在途库存", "Incoming Stock", "incoming_stock"],
                    "net_available_qty": ["净可用量", "Net Available Qty", "net_available_qty"],
                    "safety_stock": ["安全库存", "Safety Stock", "safety_stock"],
                    "lead_time_days": ["提前期天数", "Lead Time Days", "lead_time_days"],
                    "earliest_due_date": ["最早需求日期", "Earliest Due Date", "earliest_due_date"],
                    "suggested_order_date": ["建议下单日期", "Suggested Order Date", "suggested_order_date"],
                    "impacted_order_count": ["影响订单数", "Impacted Order Count", "impacted_order_count"],
                    "impacted_orders": ["影响订单", "Impacted Orders", "impacted_orders"],
                    "planning_note": ["计划备注", "Planning Note", "planning_note"],
                },
            )
            for row in item_rows[1:]:
                request_no = (
                    _normalize_string(row[item_col_map["request_no"]])
                    if "request_no" in item_col_map and item_col_map["request_no"] < len(row)
                    else None
                )
                material_code = (
                    _normalize_string(row[item_col_map["material_code"]])
                    if "material_code" in item_col_map and item_col_map["material_code"] < len(row)
                    else None
                )
                material_name = (
                    _normalize_string(row[item_col_map["material_name"]])
                    if "material_name" in item_col_map and item_col_map["material_name"] < len(row)
                    else None
                )
                if not request_no or not material_code or not material_name:
                    continue

                request_items_map.setdefault(request_no, []).append(
                    {
                        "material_code": material_code,
                        "material_name": material_name,
                        "unit": _normalize_string(row[item_col_map["unit"]]) if "unit" in item_col_map and item_col_map["unit"] < len(row) else "PCS",
                        "material_type": _normalize_string(row[item_col_map["material_type"]]) if "material_type" in item_col_map and item_col_map["material_type"] < len(row) else None,
                        "material_category": _normalize_string(row[item_col_map["material_category"]]) if "material_category" in item_col_map and item_col_map["material_category"] < len(row) else None,
                        "readiness_status": _normalize_string(row[item_col_map["readiness_status"]]) if "readiness_status" in item_col_map and item_col_map["readiness_status"] < len(row) else None,
                        "shortage_reason": _normalize_string(row[item_col_map["shortage_reason"]]) if "shortage_reason" in item_col_map and item_col_map["shortage_reason"] < len(row) else None,
                        "procurement_mode": _normalize_string(row[item_col_map["procurement_mode"]]) if "procurement_mode" in item_col_map and item_col_map["procurement_mode"] < len(row) else None,
                        "suggested_action": _normalize_string(row[item_col_map["suggested_action"]]) if "suggested_action" in item_col_map and item_col_map["suggested_action"] < len(row) else None,
                        "urgency_level": _parse_urgency_value(row[item_col_map["urgency_level"]]) if "urgency_level" in item_col_map and item_col_map["urgency_level"] < len(row) else "LOW",
                        "requested_qty": _parse_float(row[item_col_map["requested_qty"]]) if "requested_qty" in item_col_map and item_col_map["requested_qty"] < len(row) else 0,
                        "shortage_qty": _parse_float(row[item_col_map["shortage_qty"]]) if "shortage_qty" in item_col_map and item_col_map["shortage_qty"] < len(row) else 0,
                        "shortage_with_safety_qty": _parse_float(row[item_col_map["shortage_with_safety_qty"]]) if "shortage_with_safety_qty" in item_col_map and item_col_map["shortage_with_safety_qty"] < len(row) else 0,
                        "current_stock": _parse_float(row[item_col_map["current_stock"]]) if "current_stock" in item_col_map and item_col_map["current_stock"] < len(row) else 0,
                        "reserved_stock": _parse_float(row[item_col_map["reserved_stock"]]) if "reserved_stock" in item_col_map and item_col_map["reserved_stock"] < len(row) else 0,
                        "incoming_stock": _parse_float(row[item_col_map["incoming_stock"]]) if "incoming_stock" in item_col_map and item_col_map["incoming_stock"] < len(row) else 0,
                        "net_available_qty": _parse_float(row[item_col_map["net_available_qty"]]) if "net_available_qty" in item_col_map and item_col_map["net_available_qty"] < len(row) else 0,
                        "safety_stock": _parse_float(row[item_col_map["safety_stock"]]) if "safety_stock" in item_col_map and item_col_map["safety_stock"] < len(row) else 0,
                        "lead_time_days": _parse_int(row[item_col_map["lead_time_days"]]) if "lead_time_days" in item_col_map and item_col_map["lead_time_days"] < len(row) else 0,
                        "earliest_due_date": _parse_datetime_value(row[item_col_map["earliest_due_date"]]) if "earliest_due_date" in item_col_map and item_col_map["earliest_due_date"] < len(row) else None,
                        "suggested_order_date": _parse_datetime_value(row[item_col_map["suggested_order_date"]]) if "suggested_order_date" in item_col_map and item_col_map["suggested_order_date"] < len(row) else None,
                        "impacted_order_count": _parse_int(row[item_col_map["impacted_order_count"]]) if "impacted_order_count" in item_col_map and item_col_map["impacted_order_count"] < len(row) else 0,
                        "impacted_orders": _parse_list(row[item_col_map["impacted_orders"]]) if "impacted_orders" in item_col_map and item_col_map["impacted_orders"] < len(row) else [],
                        "planning_note": _normalize_string(row[item_col_map["planning_note"]]) if "planning_note" in item_col_map and item_col_map["planning_note"] < len(row) else None,
                    }
                )

    created = []
    updated = []
    errors = []

    for row_idx, row in enumerate(request_rows[1:], start=2):
        try:
            request_no = (
                _normalize_string(row[request_col_map["request_no"]])
                if request_col_map["request_no"] < len(row)
                else None
            )
            if not request_no:
                continue

            item_payloads = request_items_map.get(request_no, [])
            payload = {
                "title": _normalize_string(row[request_col_map["title"]]) if "title" in request_col_map and request_col_map["title"] < len(row) else request_no,
                "source_scope": _normalize_string(row[request_col_map["source_scope"]]) if "source_scope" in request_col_map and request_col_map["source_scope"] < len(row) else "GLOBAL",
                "status": _parse_procurement_status_value(row[request_col_map["status"]]) if "status" in request_col_map and request_col_map["status"] < len(row) else "DRAFT",
                "urgency_level": _parse_urgency_value(row[request_col_map["urgency_level"]]) if "urgency_level" in request_col_map and request_col_map["urgency_level"] < len(row) else "MEDIUM",
                "requester_name": _normalize_string(row[request_col_map["requester_name"]]) if "requester_name" in request_col_map and request_col_map["requester_name"] < len(row) else None,
                "notes": _normalize_string(row[request_col_map["notes"]]) if "notes" in request_col_map and request_col_map["notes"] < len(row) else None,
            }
            source_order_id = _parse_int(row[request_col_map["source_order_id"]]) if "source_order_id" in request_col_map and request_col_map["source_order_id"] < len(row) else None
            source_order_id = source_order_id or None
            submitted_at = _parse_datetime_value(row[request_col_map["submitted_at"]]) if "submitted_at" in request_col_map and request_col_map["submitted_at"] < len(row) else None
            completed_at = _parse_datetime_value(row[request_col_map["completed_at"]]) if "completed_at" in request_col_map and request_col_map["completed_at"] < len(row) else None
            created_at = _parse_datetime_value(row[request_col_map["created_at"]]) if "created_at" in request_col_map and request_col_map["created_at"] < len(row) else None
            updated_at = _parse_datetime_value(row[request_col_map["updated_at"]]) if "updated_at" in request_col_map and request_col_map["updated_at"] < len(row) else None

            existing = db.query(ProcurementRequest).filter(ProcurementRequest.request_no == request_no).first()
            if existing:
                request = ProcurementRequestService.update_request(db, existing.id, payload)
                if item_payloads:
                    request = ProcurementRequestService.replace_request_items(db, existing.id, item_payloads)
                existing = db.query(ProcurementRequest).filter(ProcurementRequest.id == existing.id).first()
                if submitted_at is not None:
                    existing.submitted_at = submitted_at
                if completed_at is not None:
                    existing.completed_at = completed_at
                if created_at is not None:
                    existing.created_at = created_at
                if updated_at is not None:
                    existing.updated_at = updated_at
                db.commit()
                request = ProcurementRequestService.get_request(db, existing.id)
                updated.append({"row": row_idx, "request_no": request.request_no, "id": request.id})
            else:
                request = ProcurementRequest(
                    request_no=request_no,
                    title=payload["title"] or request_no,
                    source_scope=str(payload["source_scope"] or "GLOBAL").upper(),
                    source_order_id=source_order_id,
                    status=payload["status"] or "DRAFT",
                    urgency_level=payload["urgency_level"] or "MEDIUM",
                    total_items=len(item_payloads),
                    suggested_purchase_qty_total=round(sum(float(item.get("requested_qty") or 0) for item in item_payloads), 4),
                    requester_name=payload["requester_name"],
                    notes=payload["notes"],
                    submitted_at=submitted_at,
                    completed_at=completed_at,
                    created_at=created_at or system_now(),
                    updated_at=updated_at or system_now(),
                )
                db.add(request)
                db.commit()
                db.refresh(request)
                if item_payloads:
                    request = ProcurementRequestService.replace_request_items(db, request.id, item_payloads)
                created.append({"row": row_idx, "request_no": request.request_no, "id": request.id})
        except Exception as exc:
            errors.append({"row": row_idx, "error": str(exc)})

    return {
        "imported": len(created) + len(updated),
        "errors": len(errors),
        "created": created,
        "updated": updated,
        "error_details": errors,
    }
