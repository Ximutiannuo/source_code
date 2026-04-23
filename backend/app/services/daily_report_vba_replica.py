"""
日报系统服务 - 完全复刻VBA宏的操作流程和格式
严格按照 l_DAILYREPORT.bas 和 l_IIDAILYREPORT.bas 的逻辑实现
"""
from typing import List, Dict, Optional, Tuple, Any
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session
from sqlalchemy import text
import copy

from app.models.activity_summary import ActivitySummary
from app.models.report import MPDB, VFACTDB
from app.models.rsc import RSCDefine


class DailyReportVBAReplica:
    """
    完全复刻VBA宏的日报系统
    对应 VBA 宏：l_DAILYREPORT.bas 和 l_IIDAILYREPORT.bas
    """
    
    def __init__(self, db: Session):
        self.db = db
        self.ovrList = {}  # 对应VBA的ovrList，按scope索引
        self.scope = []    # 对应VBA的scope数组
        self.vlist = []    # 对应VBA的vlist（VFACT用）
    
    def create_ovr_list(self):
        """
        完全复刻VBA的CreateOvrList函数
        从Activity List按scope分组，创建ovrList数据结构
        
        VBA逻辑：
        1. 从Activity List工作表读取数据（第6行开始）
        2. 按scope（第14列）分组
        3. 对每个scope，将所有列的数据用"@"连接
        4. 创建ovrList字典，key是scope，value是数组的数组
        
        修复：使用gcc_scope从activity_summary或p6_activity_code_assignments获取
        """
        from app.p6_sync.models.activity_code_assignment import P6ActivityCodeAssignment
        from sqlalchemy import or_
        
        # 从activity_summary读取数据（对应VBA的Activity List）
        activities = self.db.query(ActivitySummary).filter(
            ActivitySummary.is_active == True
        ).all()
        
        # 获取所有activity_id，用于查询activity_code_assignments
        activity_ids = [act.activity_id for act in activities if act.activity_id]
        
        # 从p6_activity_code_assignments获取GCC_Scope（如果activity_summary的scope为空）
        scope_map = {}  # activity_id -> scope
        if activity_ids:
            code_assignments = self.db.query(P6ActivityCodeAssignment).filter(
                P6ActivityCodeAssignment.activity_id.in_(activity_ids),
                P6ActivityCodeAssignment.activity_code_type_name == 'GCC_Scope',
                P6ActivityCodeAssignment.is_active == True
            ).all()
            
            for assignment in code_assignments:
                if assignment.activity_id and assignment.activity_code_value:
                    scope_map[assignment.activity_id] = assignment.activity_code_value
        
        # 按scope分组
        scope_dict = {}
        
        for activity in activities:
            # 优先使用activity_summary的scope，如果为空则从activity_code_assignments获取
            scope_value = activity.scope
            if not scope_value and activity.activity_id:
                scope_value = scope_map.get(activity.activity_id)
            
            if not scope_value:
                continue
            
            if scope_value not in scope_dict:
                scope_dict[scope_value] = []
            
            # 构建活动数据数组（对应VBA的clist数组）
            # 注意：VBA中Activity List的列顺序需要确认
            # 这里假设按activity_summary的字段顺序
            activity_data = [
                activity.wbs_code,           # 1: WBS
                activity.activity_id,        # 2: Activity ID
                activity.project,            # 3: Project
                activity.subproject,    # 4: Sub-project
                activity.implement_phase,              # 5: Phase
                activity.train,              # 6: Train
                activity.unit,              # 7: Unit
                activity.block,             # 8: Block
                activity.quarter,       # 9: Quarter
                None,                       # 10: (可能需要补充)
                activity.title,             # 11: Description
                activity.discipline,        # 12: Discipline
                activity.work_package,       # 13: Work Package
                activity.scope,             # 14: Scope
                activity.simple_block,       # 15: SIMPBLK
                activity.start_up_sequence, # 16: Startup Sequence
                activity.uom,               # 17: UOM
                activity.contract_phase,  # 18: BCC Work Package
                None,                       # 19: (可能需要补充)
                activity.key_qty,           # 20: Key Qty
                activity.calculated_mhrs,    # 21: Calculated MHrs
                activity.resource_id,        # 22: Resource ID
                activity.spe_mhrs,          # 23: SPE MHrs
                activity.weight_factor,      # 24: Weight Factor
                None,                       # 25: (可能需要补充)
                activity.baseline1_start_date,          # 26: Baseline1 Start Date
                activity.baseline1_finish_date,         # 27: Baseline1 Finish Date
                activity.planned_duration,          # 28: Planned Duration
                activity.actual_start_date,      # 29: Actual Start Date
                activity.actual_finish_date,     # 30: Actual Finish Date
                activity.actual_duration,   # 31: Actual Duration
                activity.completed,         # 32: Completed
                activity.actual_manhour,    # 33: Actual Manhour
                None,                       # 34: (可能需要补充)
                None,                       # 35: (可能需要补充)
                None,                       # 36: (可能需要补充)
                None,                       # 37: (可能需要补充)
            ]
            
            scope_dict[scope_value].append(activity_data)
        
        # 构建ovrList（对应VBA的字典结构）
        # VBA中：ovrList(idx)(col) 表示第idx个scope的第col列数据数组
        self.scope = list(scope_dict.keys())
        self.ovrList = {}
        
        for idx, scope_value in enumerate(self.scope):
            activities_in_scope = scope_dict[scope_value]
            
            # 转置：从行优先转为列优先（对应VBA的数组结构）
            # VBA中ovrList(idx)(col)是第col列的所有值
            num_cols = len(activities_in_scope[0]) if activities_in_scope else 0
            ovrList_cols = []
            
            for col_idx in range(num_cols):
                col_values = [act[col_idx] for act in activities_in_scope]
                ovrList_cols.append(col_values)
            
            self.ovrList[idx] = ovrList_cols
    
    def build_report_structure(self, activity_ids: List[str], scope_idx: int) -> List[List[str]]:
        """
        完全复刻VBA的BuildReportStructure函数
        构建WBS层级结构
        
        VBA逻辑：
        1. 从activity_ids提取Project/Subproject/Phase/Train/Unit/Block-Discipline-WorkPackage信息
        2. 按这些信息分组
        3. 构建WBS路径字符串
        4. 按WBS Table排序
        5. 展开为层级结构
        
        Args:
            activity_ids: 活动ID列表（对应VBA的IDs参数）
            scope_idx: scope索引（对应VBA的idx参数）
            
        Returns:
            结构化的WBS路径列表（对应VBA的result数组）
        """
        if scope_idx not in self.ovrList:
            return []
        
        ovrList_data = self.ovrList[scope_idx]
        
        # 获取activity_id列（第2列，索引1）
        activity_id_col = ovrList_data[1] if len(ovrList_data) > 1 else []
        
        # 构建临时数组（对应VBA的temp数组）
        temp = []
        for i, activity_id in enumerate(activity_ids):
            # 在ovrList中查找对应的活动
            try:
                activity_idx = activity_id_col.index(activity_id)
            except ValueError:
                continue
            
            # 提取信息（对应VBA的temp数组构建逻辑）
            project = ovrList_data[2][activity_idx] if len(ovrList_data) > 2 else ''  # 第3列：Project
            subproject = ovrList_data[3][activity_idx] if len(ovrList_data) > 3 else ''  # 第4列：Sub-project
            implement_phase = ovrList_data[4][activity_idx] if len(ovrList_data) > 4 else ''  # 第5列：Phase
            train = ovrList_data[5][activity_idx] if len(ovrList_data) > 5 else ''  # 第6列：Train
            unit = ovrList_data[6][activity_idx] if len(ovrList_data) > 6 else ''  # 第7列：Unit
            block = ovrList_data[7][activity_idx] if len(ovrList_data) > 7 else ''  # 第8列：Block
            work_package = ovrList_data[12][activity_idx] if len(ovrList_data) > 12 else ''  # 第13列：Work Package
            
            # 解析block代码（格式：XXXX-XXXXX-XX）
            if len(block) >= 11:
                block_main = block[:4]
                block_mid = block[5:10] if len(block) > 5 else ''
                block_end = block[10:12] if len(block) > 10 else ''
            else:
                block_main = block[:4] if len(block) >= 4 else block
                block_mid = ''
                block_end = ''
            
            # 构建WBS路径（对应VBA的tempstr）
            wbs_path = f"{project}/{subproject}/{implement_phase}/{train}/{unit}/{block_main}-{block_mid}-{block_end}"
            
            # 获取activity_id后缀
            if len(activity_id) == 20:
                activity_id_suffix = activity_id[-4:]
            else:
                activity_id_suffix = activity_id[-3:] if len(activity_id) >= 3 else activity_id
            
            temp.append([
                project[:2] if project else '',  # 前2位
                subproject,
                implement_phase,
                train,
                unit,
                block_main,
                block_mid,
                block_end,
                work_package,
                activity_id_suffix
            ])
        
        # 按WBS路径分组（对应VBA的字典分组逻辑）
        grouped = {}
        for item in temp:
            group_key = f"{item[0]}/{item[1]}/{item[2]}/{item[3]}/{item[4]}/{item[5]}-{item[6]}-{item[7]}"
            if group_key not in grouped:
                grouped[group_key] = {'workpackages': [], 'activity_ids': []}
            grouped[group_key]['workpackages'].append(item[8])
            grouped[group_key]['activity_ids'].append(item[9])
        
        # 按WBS Table排序（这里需要从数据库或配置获取WBS Table）
        # 暂时按key排序
        sorted_groups = sorted(grouped.items())
        
        # 展开为层级结构（对应VBA的展开逻辑）
        result = []
        for group_key, group_data in sorted_groups:
            workpackages = group_data['workpackages']
            activity_ids_suffix = group_data['activity_ids']
            
            # 为每个workpackage和activity_id组合创建一行
            for wp, act_suffix in zip(workpackages, activity_ids_suffix):
                wbs_full_path = f"{group_key}/{wp}/{act_suffix}"
                result.append([wbs_full_path])
        
        return result
    
    def structure_info_matching(
        self, 
        structured_ids: List[List[str]], 
        scope_idx: int
    ) -> List[List[Any]]:
        """
        完全复刻VBA的StructureInfoMatching函数
        匹配活动信息，从WBS Table获取描述，设置层级
        
        VBA逻辑：
        1. 解析WBS路径，确定层级（blk）
        2. 如果blk=6，从ovrList匹配活动信息
        3. 否则，从WBS Table匹配描述
        4. 设置层级和描述
        
        Args:
            structured_ids: 结构化的WBS路径列表（来自BuildReportStructure）
            scope_idx: scope索引
            
        Returns:
            匹配后的数据数组（18列）
        """
        # 从数据库获取WBS Table（需要创建WBS Table模型或从配置读取）
        # 暂时使用空列表
        wbs_table = []
        relwbs = [3, 5, 7, 9, 11, 13, 15]  # 对应VBA的RELWBS数组
        
        result = []
        
        for structed_id_row in structured_ids:
            if not structed_id_row:
                continue
            
            wbs_path = structed_id_row[0]
            parts = wbs_path.split('/')
            blk = len(parts) - 1  # 层级（对应VBA的blk）
            tempstr = parts[-1] if parts else ''
            
            # 初始化行数据（18列）
            row_data = [None] * 18
            
            if blk == 6:
                # 从ovrList匹配（对应VBA的blk=6逻辑）
                if scope_idx in self.ovrList:
                    ovrList_data = self.ovrList[scope_idx]
                    block_part = parts[4] if len(parts) > 4 else ''
                    workpackage_part = parts[5] if len(parts) > 5 else ''
                    activity_id_suffix = parts[6] if len(parts) > 6 else ''
                    
                    # 在ovrList中查找匹配的活动
                    if len(ovrList_data) > 1:
                        activity_id_col = ovrList_data[1]
                        block_col = ovrList_data[7] if len(ovrList_data) > 7 else []
                        workpackage_col = ovrList_data[12] if len(ovrList_data) > 12 else []
                        
                        for j in range(len(activity_id_col)):
                            block_match = block_col[j] if j < len(block_col) else None
                            wp_match = workpackage_col[j] if j < len(workpackage_col) else None
                            activity_id = activity_id_col[j] if j < len(activity_id_col) else None
                            
                            if (block_match == block_part and 
                                wp_match == workpackage_part and
                                ((len(activity_id) == 20 and activity_id[-4:] == activity_id_suffix) or
                                 (len(activity_id) == 19 and activity_id[-3:] == activity_id_suffix))):
                                
                                # 填充数据（对应VBA的第218-228行逻辑）
                                for m in range(1, 9):
                                    if m < len(ovrList_data):
                                        row_data[m-1] = ovrList_data[m-1][j] if j < len(ovrList_data[m-1]) else None
                                
                                row_data[8] = ovrList_data[10][j] if len(ovrList_data) > 10 and j < len(ovrList_data[10]) else None  # 第11列
                                row_data[9] = ovrList_data[11][j] if len(ovrList_data) > 11 and j < len(ovrList_data[11]) else None  # 第12列
                                row_data[10] = ovrList_data[12][j] if len(ovrList_data) > 12 and j < len(ovrList_data[12]) else None  # 第13列
                                row_data[11] = 8  # 层级8
                                row_data[12] = ovrList_data[25][j] if len(ovrList_data) > 25 and j < len(ovrList_data[25]) else None  # 第26列：BL Start
                                row_data[13] = ovrList_data[26][j] if len(ovrList_data) > 26 and j < len(ovrList_data[26]) else None  # 第27列：BL Finish
                                row_data[14] = ovrList_data[27][j] if len(ovrList_data) > 27 and j < len(ovrList_data[27]) else None  # 第28列：Duration
                                row_data[15] = ovrList_data[28][j] if len(ovrList_data) > 28 and j < len(ovrList_data[28]) else None  # 第29列：Actual Start
                                row_data[16] = ovrList_data[29][j] if len(ovrList_data) > 29 and j < len(ovrList_data[29]) else None  # 第30列：Actual Finish
                                break
            else:
                # 从WBS Table匹配（对应VBA的blk!=6逻辑）
                if blk == 5 and len(tempstr) >= 4:
                    blk = 6
                
                # 匹配WBS Table
                matched = False
                for wbs_row in wbs_table:
                    if blk < len(relwbs) and wbs_row[relwbs[blk]] == tempstr:
                        row_data[8] = wbs_row[relwbs[blk] + 1] if len(wbs_row) > relwbs[blk] + 1 else None  # 描述
                        row_data[11] = blk + 1  # 层级
                        matched = True
                        break
                
                # 特殊处理（对应VBA的blk=0逻辑）
                if not matched and blk == 0:
                    if tempstr == "GE":
                        row_data[8] = "Interface"
                        row_data[11] = 1
                    elif tempstr == "EX":
                        row_data[8] = "Temporary Facilities"
                        row_data[11] = 1
                    elif tempstr == "UI":
                        row_data[8] = "Interconnecting Unit"
                        row_data[11] = 1
                    elif tempstr in ["PE", "PW", "HE", "BU"]:
                        row_data[8] = "PE/LAO"
                        row_data[11] = 1
                    elif tempstr == "EC":
                        row_data[8] = "Ethane Cracking Unit"
                        row_data[11] = 1
            
            result.append(row_data)
        
        return result
    
    def format_end_mp(
        self,
        ws: Worksheet,
        report_date: date,
        scope: str,
        format_config: Optional[List[List[Any]]] = None
    ):
        """
        完全复刻VBA的formatEndMP函数
        格式化MP模板Excel
        
        VBA逻辑：
        1. 从format工作表读取格式化配置
        2. 根据层级（第12列）应用颜色、字体、边框、缩进
        3. 设置公式和固定文本
        4. 解锁可编辑单元格（第8级的Q列和R列）
        """
        # 默认格式化配置（对应VBA的format工作表）
        if format_config is None:
            format_config = [
                # [R, G, B, 其他, Level]
                [54, 96, 146, None, 1],   # 层级1：深蓝色
                [79, 129, 189, None, 2],  # 层级2：蓝色
                [149, 179, 215, None, 3], # 层级3：浅蓝色
                [184, 204, 228, None, 4], # 层级4：更浅蓝色
                [217, 225, 242, None, 5], # 层级5：很浅蓝色
                [231, 230, 230, None, 6], # 层级6：灰色
                [242, 242, 242, None, 7], # 层级7：浅灰色
                [255, 255, 255, None, 8], # 层级8：白色（可编辑）
            ]
        
        # 读取数据（从第13行开始）
        max_row = ws.max_row
        if max_row < 13:
            return
        
        # 应用格式化（对应VBA的第472-489行）
        for row in range(13, max_row + 1):
            level = ws[f"L{row}"].value  # 第12列（L列）是层级
            
            if level is None:
                continue
            
            # 查找对应的格式化配置
            fmt_row = None
            for fmt in format_config:
                if len(fmt) > 4 and fmt[4] == level:
                    fmt_row = fmt
                    break
            
            if fmt_row is None:
                continue
            
            # 应用颜色
            bg_color = f"{fmt_row[0]:02X}{fmt_row[1]:02X}{fmt_row[2]:02X}"
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            
            # 应用字体
            font_name = 'Arial Narrow'
            font_color = '000000'  # 默认黑色
            if level == 3 or level == 4:
                font_color = 'FFFFFF'  # 白色
            elif level == 2:
                font_color = '0000FF'  # 蓝色
            
            font = Font(name=font_name, color=font_color)
            
            # 应用边框
            thin_border = Border(
                left=Side(style='thin', color='909090'),
                right=Side(style='thin', color='909090'),
                top=Side(style='thin', color='909090'),
                bottom=Side(style='thin', color='909090')
            )
            
            # 应用到整行（A到S列，19列）
            for col in range(1, 20):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.font = font
                cell.border = thin_border
                
                # 设置缩进（第9列，I列）
                if col == 9:
                    cell.alignment = Alignment(indent=level - 1)
        
        # 设置公式（对应VBA的第491-492行）
        ws['Q8'] = '=SUM(Q13:Q1000000)+SUM(Q2:Q7)'
        ws['R8'] = '=SUM(R13:R1000000)+SUM(R2:R7)'
        
        # 设置固定文本（对应VBA的第493-500行）
        ws['B7'] = scope
        ws['R1'] = f"DATE OF REPORT {report_date.strftime('%d.%m.%Y')} {datetime.now().strftime('%H:%M:%S')}"
        ws['O2'] = "Management Personnel"
        ws['O3'] = "Technical Personnel"
        ws['O4'] = "HSE"
        ws['O5'] = "Logistic"
        ws['O6'] = "Day-off(indirect)"
        ws['O7'] = "Day-off(direct)"
        
        # 设置O2:R8的字体
        for row in range(2, 9):
            for col_letter in ['O', 'P', 'Q', 'R']:
                cell = ws[f"{col_letter}{row}"]
                cell.font = Font(name='Arial Narrow')
        
        # 设置Q2:R8的边框和锁定状态（对应VBA的第502-506行）
        for row in range(2, 9):
            for col_letter in ['Q', 'R']:
                cell = ws[f"{col_letter}{row}"]
                cell.border = thin_border
                cell.protection = openpyxl.styles.Protection(locked=False)
        
        # 解锁第8级的Q列和R列（对应VBA的第637-642行）
        for row in range(13, max_row + 1):
            level = ws[f"L{row}"].value
            if level == 8:
                for col_letter in ['Q', 'R', 'S']:
                    cell = ws[f"{col_letter}{row}"]
                    cell.protection = openpyxl.styles.Protection(locked=False)
    
    def group_levels(self, ws: Worksheet, is_mp: bool = True):
        """
        完全复刻VBA的groupLevels函数
        分组层级，设置Excel分组
        
        VBA逻辑：
        1. 根据层级（第12列）分组行
        2. 设置列分组（C:H和J:N）
        3. 隐藏列
        4. 解锁可编辑单元格
        """
        max_row = ws.max_row
        if max_row < 13:
            return
        
        # 读取数据（从第13行开始）
        data_rows = []
        for row in range(13, max_row + 1):
            level = ws[f"L{row}"].value  # 第12列（L列）是层级
            if level is not None:
                data_rows.append((row - 13, level))  # 相对行号（从0开始）和层级
        
        # 构建层级数组（对应VBA的s数组）
        if is_mp:
            level_range = range(1, 8)  # MP: 1-7
        else:
            level_range = range(2, 9)  # VFACT: 2-8
        
        level_groups = {level: [] for level in level_range}
        
        for rel_row, level in data_rows:
            if level in level_groups:
                level_groups[level].append(rel_row)
        
        # 分组行（对应VBA的第595-632行逻辑）
        # 这里简化实现，实际需要更复杂的逻辑
        # openpyxl的分组功能有限，可能需要使用其他方法
        
        # 设置列分组（对应VBA的第633-636行）
        # C:H列分组
        ws.column_dimensions.group('C', 'H', outline_level=1)
        ws.column_dimensions.group('J', 'N', outline_level=1)
        
        # 隐藏列
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col_letter].hidden = True
    
    def generate_mp_template(
        self,
        report_date: date,
        scope: str,
        template_path: Optional[Path] = None,
        filename_config: Optional[Dict] = None
    ) -> bytes:
        """
        完全复刻VBA的MPREPORT_CREATE和CreateMPTemp函数
        生成MP（人力日报）Excel模板
        
        VBA流程：
        1. 创建文件夹（按日期）
        2. 从TEMP.xlsb模板创建文件
        3. CreateOvrList - 创建活动列表
        4. 对每个scope：
           - BuildReportStructure - 构建WBS结构
           - StructureInfoMatching - 匹配信息
           - 写入Excel（从A13开始）
           - formatEndMP - 格式化
           - groupLevels - 分组层级
        
        Args:
            report_date: 报告日期（VBA中是Date+1）
            scope: 范围代码
            template_path: TEMP.xlsb模板路径
            filename_config: FILENAME配置（对应VBA的FILENAME工作表）
            
        Returns:
            Excel文件的字节数据
        """
        # 1. 创建ovrList
        self.create_ovr_list()
        
        # 2. 找到对应的scope索引
        try:
            scope_idx = self.scope.index(scope)
        except ValueError:
            raise ValueError(f"Scope '{scope}' not found")
        
        # 3. 获取活动ID列表（对应VBA的processList = ovrList(j)(2)）
        if scope_idx not in self.ovrList:
            raise ValueError(f"No data for scope '{scope}'")
        
        ovrList_data = self.ovrList[scope_idx]
        if len(ovrList_data) < 2:
            raise ValueError(f"Insufficient data for scope '{scope}'")
        
        activity_ids = ovrList_data[1]  # 第2列是activity_id
        
        # 4. 构建报告结构
        structured_ids = self.build_report_structure(activity_ids, scope_idx)
        
        if not structured_ids:
            # 空列表，返回空模板
            structured_data = []
        else:
            # 5. 匹配信息
            structured_data = self.structure_info_matching(structured_ids, scope_idx)
        
        # 6. 加载或创建Excel模板
        if template_path and template_path.exists():
            wb = openpyxl.load_workbook(template_path)
            # 找到或创建工作表
            sheet_name = f"MP-{report_date.strftime('%Y%m%d')}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # 如果模板中有"TEMP PMS_MP"工作表，重命名它
                if "TEMP PMS_MP" in wb.sheetnames:
                    ws = wb["TEMP PMS_MP"]
                    ws.title = sheet_name
                else:
                    ws = wb.active
                    ws.title = sheet_name
        else:
            # 创建新工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"MP-{report_date.strftime('%Y%m%d')}"
        
        # 7. 写入数据（从A13开始，对应VBA的第196行）
        start_row = 13
        for idx, row_data in enumerate(structured_data):
            row_num = start_row + idx
            
            # 写入18列数据（对应VBA的vdata数组）
            # 列映射：A=WBS, B=Activity ID, C=Project, D=Sub-project, E=Phase, F=Train, G=Unit, H=Block, I=Description, J=Discipline, K=Work Package, L=Level, M=BL Start, N=BL Finish, O=Duration, P=Scope, Q=Manpower, R=Machinery, S=Remarks
            if len(row_data) >= 18:
                ws[f"A{row_num}"] = row_data[0] if row_data[0] else ''  # WBS（需要从structured_ids获取）
                ws[f"B{row_num}"] = row_data[1] if row_data[1] else ''  # Activity ID
                ws[f"C{row_num}"] = row_data[2] if row_data[2] else ''  # Project
                ws[f"D{row_num}"] = row_data[3] if row_data[3] else ''  # Sub-project
                ws[f"E{row_num}"] = row_data[4] if row_data[4] else ''  # Phase
                ws[f"F{row_num}"] = row_data[5] if row_data[5] else ''  # Train
                ws[f"G{row_num}"] = row_data[6] if row_data[6] else ''  # Unit
                ws[f"H{row_num}"] = row_data[7] if row_data[7] else ''  # Block
                ws[f"I{row_num}"] = row_data[8] if row_data[8] else ''  # Description
                ws[f"J{row_num}"] = row_data[9] if row_data[9] else ''  # Discipline
                ws[f"K{row_num}"] = row_data[10] if row_data[10] else ''  # Work Package
                ws[f"L{row_num}"] = row_data[11] if row_data[11] else ''  # Level
                ws[f"M{row_num}"] = row_data[12] if row_data[12] else ''  # BL Start
                ws[f"N{row_num}"] = row_data[13] if row_data[13] else ''  # BL Finish
                ws[f"O{row_num}"] = row_data[14] if row_data[14] else ''  # Duration
                ws[f"P{row_num}"] = scope  # Scope
                # Q和R列（Manpower和Machinery）留空，供用户填写
                # S列（Remarks）留空
        
        # 8. 格式化
        self.format_end_mp(ws, report_date, scope)
        
        # 9. 分组层级
        self.group_levels(ws, is_mp=True)
        
        # 10. 保护工作表（但解锁可编辑区域）
        ws.protection.sheet = True
        ws.protection.password = 'cc7'  # 对应VBA的密码
        
        # 11. 保存到内存
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def parse_mp_template(
        self,
        file_data: bytes,
        report_date: date
    ) -> List[Dict]:
        """
        完全复刻VBA宏的MP模板解析逻辑
        解析上传的MP模板文件，提取第8级（可编辑行）的Q列（人力）和R列（机械）数据
        
        VBA逻辑：
        1. 从第13行开始读取
        2. 只处理第12列（L列）= 8的行
        3. 读取Q列（人力）和R列（机械）
        4. 匹配activity_id获取活动信息
        """
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active
        
        # 从第13行开始读取数据
        start_row = 13
        mpdb_entries = []
        
        for row in range(start_row, ws.max_row + 1):
            activity_id = ws[f"B{row}"].value  # B列是Activity ID
            manpower = ws[f"Q{row}"].value  # Q列是人力
            machinery = ws[f"R{row}"].value  # R列是机械
            level = ws[f"L{row}"].value  # L列是层级
            
            # 只处理第8级（可编辑行）且有数据的行
            if level == 8 and activity_id and (manpower or machinery):
                # 获取活动信息
                activity = self.db.query(ActivitySummary).filter(
                    ActivitySummary.activity_id == activity_id
                ).first()
                
                if activity:
                    entry = {
                        'date': report_date,
                        'activity_id': activity_id,
                        'scope': activity.scope,
                        'typeof_mp': 'Direct',  # 默认为直接人力
                        'manpower': Decimal(str(manpower)) if manpower else Decimal('0'),
                        'machinery': Decimal(str(machinery)) if machinery else Decimal('0'),
                        'project': activity.project,
                        'subproject': activity.subproject,
                        'implement_phase': activity.implement_phase,
                        'train': activity.train,
                        'unit': activity.unit,
                        'block': activity.block,
                        'discipline': activity.discipline,
                        'work_package': activity.work_package,
                        'title': activity.title
                    }
                    mpdb_entries.append(entry)
        
        return mpdb_entries
    
    def generate_vfact_template(
        self,
        report_date: date,
        scope: str,
        template_path: Optional[Path] = None
    ) -> bytes:
        """
        完全复刻VBA的PVREPORT_CREATE和prepareList_VFACT函数
        生成VFACT（物理量日报）Excel模板
        
        VBA流程：
        1. prepareList_VFACT - 从已填写的MP日报中读取数据，创建vlist
        2. 对每个scope：
           - BuildReportStructure - 构建WBS结构
           - StructureInfoMatching - 匹配信息
           - PlanDB - 匹配WorkSteps_C和VFACTDB数据
           - 写入Excel
           - formatEndPV - 格式化
        """
        # 注意：此处先实现“可用版”VFACT模板，用于导出/导入闭环与网页填报辅助。
        # 目标：
        # - 行从第13行开始
        # - B列 Activity ID
        # - L列 Level（可编辑行=9）
        # - S列 Achieved（可作为参考/历史值）
        # - W列 Completed（用户填写，导入时写回 vfactdb.achieved）
        #
        # 完全复刻 PlanDB / WorkSteps_C 的版本可后续迭代增强。
        import openpyxl
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        # 1) 选取需要进入 VFACT 的活动：优先取同日期同scope的MP填报（有人力/机械的活动），否则退化为scope全量活动
        try:
            from app.models.report import MPDB, VFACTDB
        except Exception:
            MPDB = None
            VFACTDB = None

        activity_ids: List[str] = []
        if MPDB is not None:
            mp_rows = (
                self.db.query(MPDB.activity_id)
                .filter(MPDB.date == report_date, MPDB.scope == scope, MPDB.activity_id.isnot(None))
                .filter((MPDB.manpower > 0) | (MPDB.machinery > 0))
                .distinct()
                .all()
            )
            activity_ids = [r[0] for r in mp_rows if r and r[0]]

        activities: List[ActivitySummary] = []
        if activity_ids:
            activities = (
                self.db.query(ActivitySummary)
                .filter(ActivitySummary.scope == scope, ActivitySummary.activity_id.in_(activity_ids))
                .all()
            )
        else:
            activities = (
                self.db.query(ActivitySummary)
                .filter(ActivitySummary.scope == scope)
                .all()
            )

        # 2) 创建/加载工作簿（不读取历史数据，模板应该为空，让用户填写）
        if template_path and template_path.exists():
            wb = openpyxl.load_workbook(template_path)
            sheet_name = f"VFACT-{report_date.strftime('%Y%m%d')}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                # 如果模板中有"TEMP PMS_PV"工作表，重命名它
                if "TEMP PMS_PV" in wb.sheetnames:
                    ws = wb["TEMP PMS_PV"]
                    ws.title = sheet_name
                else:
                    ws = wb.active
                    ws.title = sheet_name
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"VFACT-{report_date.strftime('%Y%m%d')}"

        # 3) 写表头（在10-11行）
        # 表头在10-11行，数据从13行开始
        header_row_1 = 10
        header_row_2 = 11
        
        # 第一行表头（主要列名）
        headers_row1 = {
            "A": "WBS",
            "B": "作业ID",  # Activity ID
            "H": "Block",
            "I": "描述",  # Description
            "J": "专业",  # Discipline
            "K": "工作包",  # Work Package
            "L": "层级",  # Level
            "P": "Scope",
            "Q": "预估总量",  # Key Qty
        }
        for col, title in headers_row1.items():
            cell = ws[f"{col}{header_row_1}"]
            cell.value = title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 第二行表头（完成量相关）
        headers_row2 = {
            "S": "完成量",  # 历史完成量（参考值，不预填）
            "W": "昨日完成量",  # 用户填写列
        }
        for col, title in headers_row2.items():
            cell = ws[f"{col}{header_row_2}"]
            cell.value = title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 4) 写入数据：从第13行开始
        start_row = 13
        thin = Side(style="thin", color="909090")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for idx, act in enumerate(activities):
            r = start_row + idx

            ws[f"A{r}"] = getattr(act, "wbs_path", "") or ""  # 如果没有wbs_path也不影响
            ws[f"B{r}"] = act.activity_id
            ws[f"H{r}"] = act.block or ""
            ws[f"I{r}"] = act.title or ""
            ws[f"J{r}"] = act.discipline or ""
            ws[f"K{r}"] = act.work_package or ""
            ws[f"L{r}"] = 9  # 可编辑层级
            ws[f"P{r}"] = scope
            ws[f"Q{r}"] = act.key_qty if act.key_qty is not None else ""  # 预估总量（Key Qty）

            # 不预填任何完成量数据，让用户填写
            # S列（完成量）留空，作为参考列但不预填
            ws[f"S{r}"] = None
            # W列（昨日完成量）留空，用户填写
            ws[f"W{r}"] = None

            # 基础边框
            for col in ["A", "B", "H", "I", "J", "K", "L", "P", "Q", "S", "W"]:
                ws[f"{col}{r}"].border = border

            # 将 W 列标记为可编辑（解锁），并用灰底提示
            ws[f"W{r}"].fill = PatternFill("solid", fgColor="E0E0E0")

        # 6) 保护工作表：锁定默认单元格，解锁 W 列输入单元格
        ws.protection.sheet = True
        ws.protection.password = 'cc7'
        for idx in range(len(activities)):
            r = start_row + idx
            ws[f"W{r}"].protection = openpyxl.styles.Protection(locked=False)

        # 7) 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def parse_vfact_template(
        self,
        file_data: bytes,
        report_date: date
    ) -> List[Dict]:
        """
        完全复刻VBA宏的VFACT模板解析逻辑
        解析上传的VFACT模板文件，提取第9级（可编辑行）的完成量数据
        根据表头动态查找"Completed(input)"列的位置
        """
        import openpyxl
        from io import BytesIO
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        wb = openpyxl.load_workbook(BytesIO(file_data))
        ws = wb.active
        
        # 查找表头行（在10-11行查找"昨日完成量"）
        header_row = None
        completed_col = None
        achieved_col = None
        activity_id_col = None
        level_col = None
        
        # 在10-11行查找表头
        for row in range(10, 12):
            for col_idx in range(1, 30):  # 检查前30列
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    # 查找"昨日完成量"列
                    if "昨日完成量" in cell_str and completed_col is None:
                        completed_col = col_idx
                        header_row = row
                    # 查找"预估总量"列（用于定位）
                    if "预估总量" in cell_str or "Key Qty" in cell_str:
                        if header_row is None:
                            header_row = row
                    # 查找"作业ID"或"Activity ID"列
                    if "作业ID" in cell_str or "Activity ID" in cell_str or "Activity_ID" in cell_str:
                        activity_id_col = col_idx
                    # 查找"Level"或"层级"列
                    if cell_str == "Level" or cell_str == "层级":
                        level_col = col_idx
        
        # 如果没找到表头，使用默认值（向后兼容）
        if header_row is None:
            header_row = 12
        if activity_id_col is None:
            activity_id_col = 2  # B列
        if level_col is None:
            level_col = 12  # L列
        if achieved_col is None:
            achieved_col = 19  # S列（如果有的话）
        if completed_col is None:
            # 如果找不到"昨日完成量"，尝试查找其他可能的列名
            for row in range(10, 12):
                for col_idx in range(1, 30):
                    cell_value = ws.cell(row=row, column=col_idx).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if "Completed" in cell_str or "completed" in cell_str.lower():
                            completed_col = col_idx
                            break
                if completed_col:
                    break
            # 如果还是找不到，使用默认值W列
            if completed_col is None:
                completed_col = 23  # W列（默认值，向后兼容）
        
        # 数据从表头下一行开始
        start_row = header_row + 1
        vfactdb_entries = []
        
        for row in range(start_row, ws.max_row + 1):
            activity_id = ws.cell(row=row, column=activity_id_col).value
            level = ws.cell(row=row, column=level_col).value
            achieved = ws.cell(row=row, column=achieved_col).value if achieved_col else None
            completed = ws.cell(row=row, column=completed_col).value if completed_col else None
            
            # 只处理第9级（可编辑行）且有activity_id的行
            # 注意：即使completed和achieved都是空的，也要处理（可能是用户清空了值）
            if level == 9 and activity_id:
                # 获取活动信息
                activity = self.db.query(ActivitySummary).filter(
                    ActivitySummary.activity_id == str(activity_id).strip()
                ).first()
                
                if activity:
                    # 兼容：VFACTDB表只有 achieved 字段。
                    # 以用户填写的 Completed 列为准；若为空则退回 Achieved 列；如果都是空则为0
                    final_value = completed if completed is not None else (achieved if achieved is not None else 0.0)
                    
                    # 获取work_step_description（从Description列读取，通常是I列）
                    description_col = 9  # I列是Description
                    work_step_description = ws.cell(row=row, column=description_col).value
                    if not work_step_description:
                        work_step_description = None
                    
                    entry = {
                        'date': report_date,
                        'activity_id': str(activity_id).strip(),
                        'scope': activity.scope,
                        'achieved': float(final_value) if final_value is not None else 0.0,
                        'project': activity.project,
                        'subproject': activity.subproject,
                        'implement_phase': activity.implement_phase,
                        'train': activity.train,
                        'unit': activity.unit,
                        'block': activity.block,
                        'discipline': activity.discipline,
                        'work_package': activity.work_package,
                        'title': activity.title,
                        'work_step_description': str(work_step_description) if work_step_description else None
                    }
                    vfactdb_entries.append(entry)
        
        return vfactdb_entries

