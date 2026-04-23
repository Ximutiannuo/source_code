"""
用户批量导出与导入服务
复用 batch_create_users 和 export_users_to_xlsx 脚本逻辑，供 API 调用。
"""
from io import BytesIO
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import or_

try:
    import pandas as pd
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    pd = None
    openpyxl = None

from app.models.user import User


def generate_import_template_bytes(db: Session) -> bytes:
    """
    生成批量导入模板 Excel，含三张表：
    1. user_data：表头 + 示例行，供填写用户数据
    2. departments：部门清单（code, name）
    3. roles：角色清单（name）
    """
    if openpyxl is None or pd is None:
        raise RuntimeError("请安装 openpyxl 和 pandas")

    from app.models.department import Department
    from app.models.user import Role

    wb = openpyxl.Workbook()
    # 1. user_data 表
    ws_user = wb.active
    ws_user.title = "user_data"
    headers = [
        "username", "email", "password_default", "full_name",
        "department", "role.name", "responsible_for",
        "is_active", "is_superuser"
    ]
    for col, h in enumerate(headers, 1):
        ws_user.cell(row=1, column=col, value=h)
    # 示例行
    sample = [
        "zhangsan", "zhangsan@example.com", "Password123!", "张三",
        "", "", "采购对接",
        1, 0
    ]
    for col, v in enumerate(sample, 1):
        ws_user.cell(row=2, column=col, value=v)

    # 2. departments 表
    ws_dept = wb.create_sheet("departments")
    ws_dept.cell(row=1, column=1, value="code")
    ws_dept.cell(row=1, column=2, value="name")
    depts = db.query(Department).filter(Department.is_active.is_(True)).order_by(Department.sort_order, Department.name).all()
    for row_idx, d in enumerate(depts, 2):
        ws_dept.cell(row=row_idx, column=1, value=d.code or "")
        ws_dept.cell(row=row_idx, column=2, value=d.name or "")

    # 3. roles 表
    ws_role = wb.create_sheet("roles")
    ws_role.cell(row=1, column=1, value="name")
    roles = db.query(Role).filter(Role.is_active.is_(True)).order_by(Role.name).all()
    for row_idx, r in enumerate(roles, 2):
        ws_role.cell(row=row_idx, column=1, value=r.name or "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_users_to_xlsx_bytes(
    db: Session,
    search: Optional[str] = None,
    department_id: Optional[int] = None,
) -> tuple[bytes, int]:
    """
    根据筛选条件导出用户到 Excel，返回 (xlsx 字节内容, 用户数)。
    列：username, email, full_name, department, responsible_for, is_active
    """
    if pd is None:
        raise RuntimeError("请安装 pandas: pip install pandas")

    query = db.query(User).order_by(User.username)
    if search:
        query = query.filter(
            or_(
                User.username.like(f"%{search}%"),
                User.email.like(f"%{search}%"),
            )
        )
    if department_id is not None:
        query = query.filter(User.department_id == department_id)

    users = query.all()
    rows = []
    for u in users:
        dept = getattr(u, "department_id", None)
        dept_name = ""
        if dept:
            from app.models.department import Department
            d = db.query(Department).filter(Department.id == dept).first()
            dept_name = d.name if d else str(dept)
        rows.append({
            "username": u.username or "",
            "email": u.email or "",
            "full_name": u.full_name or "",
            "department": dept_name,
            "responsible_for": getattr(u, "responsible_for", None) or "",
            "is_active": u.is_active,
        })
    df = pd.DataFrame(rows)
    buf = BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf.getvalue(), len(users)


def batch_import_users_from_excel(db: Session, file_content: bytes) -> dict:
    """
    从 Excel 批量导入用户。
    支持单表格式：username, email, password_default, full_name, is_active, is_superuser, role.name, department, responsible_for
    返回 {"created": int, "updated": int, "role_assigned": int, "errors": list}
    """
    import openpyxl
    from app.models.user import User, Role
    from app.models.department import Department

    def _col_index(ws, header_name):
        headers = [cell.value for cell in ws[1]]
        for idx, h in enumerate(headers, start=1):
            if h and str(h).strip().lower() == str(header_name).strip().lower():
                return idx
        return None

    def _cell_value(row, col_1based):
        if col_1based is None or col_1based < 1:
            return None
        idx = col_1based - 1
        if idx >= len(row):
            return None
        cell = row[idx]
        return getattr(cell, "value", cell)

    def _bool_val(v, default=True):
        if v is None:
            return default
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            return bool(int(v))
        if isinstance(v, str) and v.isdigit():
            return bool(int(v))
        return default

    result = {"created": 0, "updated": 0, "role_assigned": 0, "errors": []}
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active
    col_username = _col_index(ws, "username")
    col_email = _col_index(ws, "email")
    col_password = _col_index(ws, "password_default") or _col_index(ws, "password")
    col_full_name = _col_index(ws, "full_name")
    col_active = _col_index(ws, "is_active")
    col_super = _col_index(ws, "is_superuser")
    col_role_name = _col_index(ws, "role.name")
    col_dept = _col_index(ws, "department")
    col_responsible_for = _col_index(ws, "responsible_for")

    if not col_username:
        result["errors"].append("Excel 缺少 username 列")
        return result

    for row_idx in range(2, ws.max_row + 1):
        row = list(ws[row_idx])
        username = _cell_value(row, col_username)
        if not username:
            continue
        username = str(username).strip()
        email = _cell_value(row, col_email)
        if not email or (isinstance(email, str) and str(email).strip().startswith("=")):
            email = f"{username}@cc7.cn"
        else:
            email = str(email).strip() or f"{username}@cc7.cn"
        password = _cell_value(row, col_password)
        if not password:
            result["errors"].append(f"第 {row_idx} 行 {username}: 缺少密码，已跳过")
            continue
        password = str(password)
        full_name = _cell_value(row, col_full_name)
        full_name = str(full_name).strip() if full_name else None
        is_active = _bool_val(_cell_value(row, col_active), True)
        is_superuser = _bool_val(_cell_value(row, col_super), False)
        role_name_val = _cell_value(row, col_role_name)
        role_name_val = str(role_name_val).strip() if role_name_val else None
        dept_val = _cell_value(row, col_dept)
        dept_val = str(dept_val).strip() if dept_val else None
        responsible_for = _cell_value(row, col_responsible_for)
        responsible_for = str(responsible_for).strip() if responsible_for else None

        department = None
        if dept_val:
            department = db.query(Department).filter(Department.code == dept_val).first()
            if not department:
                department = db.query(Department).filter(Department.name == dept_val).first()

        existing = db.query(User).filter(User.username == username).first()
        try:
            if existing:
                result["updated"] += 1
                user = existing
                if department and getattr(user, "department_id", None) != department.id:
                    user.department_id = department.id
                if responsible_for is not None and getattr(user, "responsible_for", None) != responsible_for:
                    user.responsible_for = responsible_for
            else:
                result["created"] += 1
                user = User(
                    username=username,
                    email=email,
                    full_name=full_name,
                    is_active=is_active,
                    is_superuser=is_superuser,
                    department_id=department.id if department else None,
                    responsible_for=responsible_for,
                )
                user.set_password(password)
                db.add(user)
                db.flush()

            if role_name_val:
                role = db.query(Role).filter(Role.name == role_name_val).first()
                if role and role not in user.roles:
                    user.roles.append(role)
                    result["role_assigned"] += 1
        except Exception as e:
            result["errors"].append(f"第 {row_idx} 行 {username}: {str(e)}")

    return result
