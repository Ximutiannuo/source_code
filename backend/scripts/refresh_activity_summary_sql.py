"""
刷新作业清单汇总表数据（SQL优化版）
使用SQL聚合替代内存字典，适合百万/千万级数据量
"""
import sys
import logging
from pathlib import Path
import time

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "backend"))

from sqlalchemy import text
from sqlalchemy.exc import OperationalError
from app.database import SessionLocal, get_default_engine
from app.utils.db import retry_on_deadlock, get_lock_diagnostics
import traceback
from datetime import timezone, timedelta

# GMT+3 时区
GMT3 = timezone(timedelta(hours=3))
UTC = timezone.utc


def load_preset_weight_factors():
    """
    从「original system/OWF Weight Factor.xlsx」加载预设权重因子。
    表结构：第1行表头，ACT ID 为 B 列，Weight Factor 为 I 列。
    返回 dict[activity_id -> weight_factor]，无文件或解析失败时返回空 dict。
    会在控制台打印未加载原因（便于排查重启后未执行预设步骤）。
    """
    owf_path = project_root / "original system" / "OWF Weight Factor.xlsx"
    try:
        import openpyxl
    except ImportError:
        print(f"  - 预设权重: 未加载（缺少 openpyxl），跳过")
        return {}
    if not owf_path.exists():
        print(f"  - 预设权重: 未加载（文件不存在: {owf_path}），跳过")
        return {}
    try:
        wb = openpyxl.load_workbook(owf_path, data_only=True)
        ws = wb.active
        result = {}
        # 第1行为表头，数据从第2行起；B=2, I=9
        for r in range(2, (ws.max_row or 0) + 1):
            act_id = ws.cell(r, 2).value
            wf = ws.cell(r, 9).value
            if act_id is None:
                continue
            act_id = str(act_id).strip()
            if not act_id:
                continue
            if wf is None or (isinstance(wf, str) and not wf.strip()):
                continue
            try:
                wf_val = float(wf)
            except (TypeError, ValueError):
                continue
            result[act_id] = wf_val
        if not result:
            print(f"  - 预设权重: 未加载（表内无有效 ACT ID/Weight Factor 数据），跳过")
        return result
    except Exception as e:
        print(f"  ⚠️ 读取预设权重表失败（将仅使用公式权重）: {e}")
        return {}

def format_time_gmt3(dt):
    """将datetime转换为GMT+3时区并格式化显示"""
    if dt is None:
        return None
    # 如果datetime没有时区信息，假设是UTC（因为数据库中的completed_at是UTC）
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    # 转换为GMT+3时区
    if dt.tzinfo != GMT3:
        dt = dt.astimezone(GMT3)
    return dt.strftime('%Y-%m-%d %H:%M:%S')

def ensure_gmt3(dt):
    """确保datetime是GMT+3时区（用于时间比较）
    
    注意：数据库中的completed_at是UTC，需要转换为GMT+3
    """
    if dt is None:
        return None
    # 如果datetime没有时区信息，假设是UTC（因为数据库中的completed_at是UTC）
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    # 转换为GMT+3时区
    if dt.tzinfo != GMT3:
        return dt.astimezone(GMT3)
    return dt


def check_indexes(db):
    """检查并建议创建必要的索引"""
    print("\n检查索引...")
    
    # 检查p6_activities索引
    p6_activity_indexes = db.execute(text("""
        SHOW INDEX FROM p6_activities WHERE Column_name = 'activity_id'
    """)).fetchall()
    
    if not p6_activity_indexes:
        print("  ⚠️ p6_activities.activity_id 缺少索引，建议创建：")
        print("     CREATE INDEX idx_p6_activities_activity_id ON p6_activities(activity_id);")
    else:
        print("  ✓ p6_activities.activity_id 索引存在")
    
    # 检查p6_activity_code_assignments索引
    p6_aca_indexes = db.execute(text("""
        SHOW INDEX FROM p6_activity_code_assignments WHERE Column_name = 'activity_object_id'
    """)).fetchall()
    
    if not p6_aca_indexes:
        print("  ⚠️ p6_activity_code_assignments.activity_object_id 缺少索引，建议创建：")
        print("     CREATE INDEX idx_p6_aca_activity_object_id ON p6_activity_code_assignments(activity_object_id);")
    else:
        print("  ✓ p6_activity_code_assignments.activity_object_id 索引存在")
    
    # 检查vfactdb索引
    vfact_indexes = db.execute(text("""
        SHOW INDEX FROM vfactdb WHERE Column_name = 'activity_id'
    """)).fetchall()
    
    if not vfact_indexes:
        print("  ⚠️ vfactdb.activity_id 缺少索引，建议创建：")
        print("     CREATE INDEX idx_vfactdb_activity_id ON vfactdb(activity_id) WHERE activity_id IS NOT NULL;")
    else:
        print("  ✓ vfactdb.activity_id 索引存在")
    
    # 检查mpdb索引
    mpdb_indexes = db.execute(text("""
        SHOW INDEX FROM mpdb WHERE Column_name = 'activity_id'
    """)).fetchall()
    
    if not mpdb_indexes:
        print("  ⚠️ mpdb.activity_id 缺少索引，建议创建：")
        print("     CREATE INDEX idx_mpdb_activity_id ON mpdb(activity_id) WHERE activity_id IS NOT NULL;")
    else:
        print("  ✓ mpdb.activity_id 索引存在")
    
    # 检查mpdb的date索引（用于MIN/MAX）
    mpdb_date_indexes = db.execute(text("""
        SHOW INDEX FROM mpdb WHERE Column_name = 'date'
    """)).fetchall()
    
    if not mpdb_date_indexes:
        print("  ⚠️ mpdb.date 缺少索引，建议创建复合索引：")
        print("     CREATE INDEX idx_mpdb_activity_date ON mpdb(activity_id, date) WHERE activity_id IS NOT NULL;")
    else:
        print("  ✓ mpdb.date 索引存在")

    # 检查增量刷新所需的 updated_at 索引
    print("检查增量刷新索引...")
    tables_to_check = ['vfactdb', 'mpdb', 'volume_control_quantity', 'p6_activities', 'rsc_defines']
    for table in tables_to_check:
        col = 'p6_last_update_date' if table == 'p6_activities' else ('work_package' if table == 'rsc_defines' else 'updated_at')
        idx_exists = db.execute(text(f"SHOW INDEX FROM {table} WHERE Column_name = '{col}'")).fetchall()
        if not idx_exists:
            print(f"  ⚠️ {table}.{col} 缺少索引，同步可能较慢。")
            print(f"     CREATE INDEX idx_{table}_{col} ON {table}({col});")
        else:
            print(f"  ✓ {table}.{col} 索引存在")


def refresh_facility_filter_options(db):
    """
    刷新 facility_filter_options 预聚合表（distinct 组合，用于 facility_filter API 加速）。
    与 activity_summary 联动：在 activity_summary 刷新完成后调用。
    使用 RENAME 原子替换，失败时仅记录日志，不中断主流程。
    """
    TABLE_NAME = "facility_filter_options"
    TABLE_NEW = "facility_filter_options_new"
    try:
        t0 = time.time()
        db.execute(text(f"DROP TABLE IF EXISTS {TABLE_NEW}"))
        db.execute(text(f"""
            CREATE TABLE {TABLE_NEW} AS
            SELECT DISTINCT
                a.scope, a.project, a.subproject, a.train, a.unit,
                a.simple_block, a.main_block, a.block, a.quarter,
                a.discipline, a.implement_phase, a.contract_phase, a.type, a.work_package,
                r.resource_id_name, r.bcc_kq_code, r.kq, r.cn_wk_report
            FROM activity_summary a
            LEFT JOIN rsc_defines r ON a.work_package = r.work_package AND (r.is_active IS NULL OR r.is_active = 1)
        """))
        db.execute(text(f"DROP TABLE IF EXISTS {TABLE_NAME}"))
        db.execute(text(f"RENAME TABLE {TABLE_NEW} TO {TABLE_NAME}"))
        db.commit()
        elapsed = time.time() - t0
        n = db.execute(text(f"SELECT COUNT(*) FROM {TABLE_NAME}")).scalar()
        print(f"  ✓ facility_filter_options 已刷新，行数: {n:,}，耗时: {elapsed:.2f}秒")
    except Exception as e:
        db.rollback()
        try:
            db.execute(text(f"DROP TABLE IF EXISTS {TABLE_NEW}"))
            db.commit()
        except Exception:
            pass
        logging.getLogger(__name__).warning("facility_filter_options 刷新失败（facility_filter 将回退至实时查询）: %s", e)
        print(f"  ⚠️ facility_filter_options 刷新失败: {e}")


@retry_on_deadlock(max_retries=3)
def refresh_activity_summary_sql(is_clear_mode: bool = False):
    """
    刷新activity_summary表数据（SQL优化版）
    
    Args:
        is_clear_mode: 是否为清空重建模式（True=全量重建，False=增量更新）
    """
    from app.services.system_task_service import SystemTaskService
    
    # 避让原则：如果检测到用户正在操作，则直接跳过本次同步
    if SystemTaskService.is_any_high_priority_task_active(exclude_tasks=["background_refresh"]):
        print("\n[避让] 检测到用户正在上传日报或执行其他高优先级任务，后台同步主动放弃，等待下次周期...")
        return {'success': True, 'row_count': 0, 'skipped': True}

    db = SessionLocal()
    start_time = time.time()
    project_total_mhrs = 0 # 初始化变量，防止后续打印报错
    
    # 设置自己的锁
    SystemTaskService.set_task_lock("background_refresh", True, updated_by="refresh_activity_summary_sql.py")
    
    try:
        # ... 原有逻辑 ...
        mode_str = "全量重建模式" if is_clear_mode else "增量更新模式"
        print("=" * 60)
        print(f"刷新作业清单汇总表数据（SQL优化版）- {mode_str}")
        print("数据来源：")
        print("  - 作业信息：从P6表（p6_activities, p6_activity_code_assignments）获取")
        if is_clear_mode:
            print("    * 包含所有作业（不限制GCC_Phase）")
        else:
            print("    * 只处理变更的作业（基于p6_last_update_date）")
        print("    * main_block: 从block中提取（SUBSTRING(block, 6, 5)）")
        print("  - 工程量：从volume_control_quantity获取")
        print("  - 资源计算：从rsc_defines获取（通过GCC_Work Package匹配work_package）")
        print("  - 实际数据：从vfactdb和mpdb聚合")
        print("  - calculated_mhrs、weight_factor和actual_weight_factor会在刷新时自动计算")
        print("    * calculated_mhrs = key_qty / norms * 10（只考虑 GCC_Phase='CT' AND !BCC_WORK PACKAGE='Add.3'）")
        print("    * weight_factor = (calculated_mhrs / 项目总人工时) * 254137500")
        print("    * actual_mhrs = sum(mpdb.manpower) grouped by activity_id")
        print("    * actual_weight_factor = ((10/norms) * completed) / 项目总人工时 * 254137500")
        print("=" * 60)
        
        # 步骤0: 检查并创建主表（如果不存在）
        print("\n步骤0: 检查activity_summary表是否存在...")
        table_exists = db.execute(text("""
            SELECT COUNT(*) 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() 
            AND table_name = 'activity_summary'
        """)).scalar() > 0
        
        if not table_exists:
            print("  ⚠️ activity_summary表不存在，正在创建...")
            db.execute(text("""
                CREATE TABLE activity_summary (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    activity_id VARCHAR(100) UNIQUE NOT NULL COMMENT 'ACT ID',
                    wbs_code VARCHAR(100) COMMENT 'WBS Code',
                    block VARCHAR(100) COMMENT 'Block',
                    title TEXT COMMENT 'TITLE',
                    discipline VARCHAR(100) COMMENT 'Discipline',
                    work_package VARCHAR(100) COMMENT 'Workpackage',
                    scope VARCHAR(100) COMMENT 'SCOPE',
                    implement_phase VARCHAR(100) COMMENT 'PHASE',
                    project VARCHAR(100) COMMENT 'Project',
                    subproject VARCHAR(100) COMMENT 'Sub-Project CODE',
                    train VARCHAR(100) COMMENT 'Train',
                    unit VARCHAR(100) COMMENT 'Unit',
                    main_block VARCHAR(100) COMMENT 'Main_Block',
                    quarter VARCHAR(100) COMMENT '!BCC_Quarter',
                    simple_block VARCHAR(100) COMMENT 'GCC_SIMPBLK',
                    start_up_sequence VARCHAR(100) COMMENT '!BCC_START-UP SEQUENCE',
                    key_qty DECIMAL(18, 2) COMMENT 'KEY QTY',
                    calculated_mhrs DECIMAL(18, 2) COMMENT 'Calculated MHrs',
                    resource_id VARCHAR(100) COMMENT 'Resource ID',
                    spe_mhrs DECIMAL(18, 2) COMMENT 'SPE MHrs',
                    uom VARCHAR(50) COMMENT 'UoM',
                    contract_phase VARCHAR(100) COMMENT '!BCC_WORK PACKAGE',
                    weight_factor DECIMAL(18, 2) COMMENT 'W.F',
                    actual_weight_factor DECIMAL(18, 2) COMMENT 'Actual Weight Factor (基于实际完成工时计算)',
                    start_date DATE COMMENT 'Start Date (P6)',
                    finish_date DATE COMMENT 'Finish Date (P6)',
                    planned_start_date DATE COMMENT 'Planned Start Date (P6)',
                    planned_finish_date DATE COMMENT 'Planned Finish Date (P6)',
                    planned_duration INT COMMENT 'Planned Duration (P6)',
                    at_completion_duration INT COMMENT 'At Completion Duration (P6)',
                    baseline1_start_date DATE COMMENT 'Baseline1 Start Date (P6)',
                    baseline1_finish_date DATE COMMENT 'Baseline1 Finish Date (P6)',
                    actual_start_date DATE COMMENT 'Actual Start Date (优先从P6获取，否则从MPDB聚合)',
                    actual_finish_date DATE COMMENT 'Actual Finish Date (优先从P6获取，否则从MPDB聚合)',
                    actual_duration INT COMMENT 'Actual Duration (优先从P6获取，否则从MPDB聚合计算)',
                    completed DECIMAL(18, 2) COMMENT 'Completed',
                    actual_manhour DECIMAL(18, 2) COMMENT 'Actual Manhour',
                    data_date DATETIME COMMENT 'Data Date (P6)',
                    type VARCHAR(50) COMMENT 'Activity Type (从P6获取，用于标识里程碑等)',
                    system_status VARCHAR(50) DEFAULT 'Not Started' COMMENT '本系统内的确认状态',
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
                    INDEX idx_activity_id (activity_id),
                    INDEX idx_wbs_code (wbs_code),
                    INDEX idx_block (block),
                    INDEX idx_discipline (discipline),
                    INDEX idx_work_package (work_package),
                    INDEX idx_scope (scope),
                    INDEX idx_project (project),
                    INDEX idx_subproject (subproject),
                    INDEX idx_train (train),
                    INDEX idx_unit (unit),
                    INDEX idx_main_block (main_block),
                    INDEX idx_quarter (quarter),
                    INDEX idx_simple_block (simple_block),
                    INDEX idx_actual_start_date (actual_start_date),
                    INDEX idx_actual_finish_date (actual_finish_date),
                    INDEX idx_type (type)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='作业清单汇总表（物化视图）'
            """))
            db.commit()
            print("  ✓ activity_summary表创建成功")
        else:
            print("  ✓ activity_summary表已存在")
        
        # 检查索引
        check_indexes(db)
        
        # 步骤0.5: 检查并添加缺失的字段（如果表已存在）
        if table_exists:
            print("\n步骤0.5: 检查表结构并添加缺失字段...")
            # 检查type字段是否存在
            type_column_exists = db.execute(text("""
                SELECT COUNT(*) 
                FROM information_schema.columns 
                WHERE table_schema = DATABASE() 
                AND table_name = 'activity_summary'
                AND column_name = 'type'
            """)).scalar() > 0
            
            if not type_column_exists:
                print("  ⚠️ 检测到缺失 type 字段，正在添加...")
                try:
                    db.execute(text("""
                        ALTER TABLE activity_summary 
                        ADD COLUMN type VARCHAR(50) COMMENT 'Activity Type (从P6获取，用于标识里程碑等)' AFTER actual_manhour
                    """))
                    db.commit()
                    print("  ✓ type 字段已添加")
                    
                    # 添加索引
                    try:
                        db.execute(text("""
                            ALTER TABLE activity_summary 
                            ADD INDEX idx_type (type)
                        """))
                        db.commit()
                        print("  ✓ type 字段索引已添加")
                    except Exception as e:
                        if "Duplicate key name" not in str(e):
                            print(f"  ⚠️ 添加 type 字段索引失败（可能已存在）: {e}")
                except Exception as e:
                    print(f"  ❌ 添加 type 字段失败: {e}")
                    raise
            else:
                print("  ✓ type 字段已存在")

            # 检查data_date字段是否存在
            data_date_column_exists = db.execute(text("""
                SELECT COUNT(*) 
                FROM information_schema.columns 
                WHERE table_schema = DATABASE() 
                AND table_name = 'activity_summary'
                AND column_name = 'data_date'
            """)).scalar() > 0
            
            if not data_date_column_exists:
                print("  ⚠️ 检测到缺失 data_date 字段，正在添加...")
                try:
                    db.execute(text("""
                        ALTER TABLE activity_summary 
                        ADD COLUMN data_date DATETIME COMMENT 'Data Date (P6)' AFTER actual_manhour
                    """))
                    db.commit()
                    print("  ✓ data_date 字段已添加")
                except Exception as e:
                    print(f"  ❌ 添加 data_date 字段失败: {e}")
            else:
                print("  ✓ data_date 字段已存在")
            
            # 补全既存数据的 data_date (针对增量模式下未触达的旧记录)
            # 优化：分批补全，防止锁冲突
            print("  - 正在检查并补全既存数据的 data_date (安全分批模式)...")
            try:
                # 1. 查出需要补全的 ID 列表
                null_date_ids_res = db.execute(text("""
                    SELECT a.activity_id FROM activity_summary a
                    INNER JOIN p6_activities p ON a.activity_id = p.activity_id
                    WHERE a.data_date IS NULL AND p.data_date IS NOT NULL
                """)).fetchall()
                # 按 activity_id 排序，统一加锁顺序，避免死锁
                null_date_ids = sorted([r[0] for r in null_date_ids_res if r[0]])
                
                if null_date_ids:
                    total_to_update = len(null_date_ids)
                    batch_size = 500
                    for i in range(0, total_to_update, batch_size):
                        batch_ids = null_date_ids[i:i + batch_size]
                        db.execute(text("""
                            UPDATE activity_summary a
                            INNER JOIN p6_activities p ON a.activity_id = p.activity_id
                            SET a.data_date = p.data_date
                            WHERE a.activity_id IN :ids
                        """), {"ids": batch_ids})
                        db.commit()
                        # 增加心跳
                        SystemTaskService.set_task_lock("background_refresh", True, updated_by="refresh_activity_summary_sql.py", quiet=True)
                    print(f"    ✓ 已成功补齐 {total_to_update} 条记录s的 data_date")
                else:
                    print("    ✓ 没有需要补全的 data_date")
            except Exception as e:
                print(f"  ⚠️ 补全 data_date 失败: {e}")
        
        # 步骤1: 创建临时表
        print("\n步骤1: 创建临时表...")
        try:
            # 放宽锁等待超时到 30 秒，后台同步任务不需要太激进
            db.execute(text("SET SESSION innodb_lock_wait_timeout = 30"))
            db.commit()
            
            db.execute(text("DROP TABLE IF EXISTS activity_summary_temp"))
            db.commit()
            
            db.execute(text("""
                CREATE TABLE activity_summary_temp LIKE activity_summary
            """))
            db.commit()
            print("  ✓ 临时表创建成功")
        except Exception as e:
            if "Lock wait timeout exceeded" in str(e) or "1205" in str(e):
                print(f"  ⚠️ 检测到锁等待超时: {e}")
                print("  建议：运行 python scripts/kill_long_queries.py")
                raise
            else:
                raise
        
        # 步骤1.5: 清理已删除或非活跃的作业（增量模式）
        # 优化：采用“查出 ID 后分批删除”的策略，彻底解决 Lock wait timeout 冲突
        if not is_clear_mode:
            print("\n步骤1.5: 清理已删除或非活跃的作业 (安全模式)...")
            cleanup_start = time.time()
            
            # 1. 查出标记为非活跃或物理删除的 ID 列表
            ids_to_delete_res = db.execute(text("""
                SELECT activity_id FROM activity_summary
                WHERE activity_id IN (
                    SELECT a.activity_id FROM activity_summary a
                    INNER JOIN p6_activities p ON a.activity_id = p.activity_id
                    WHERE p.is_active = 0
                ) OR activity_id NOT IN (
                    SELECT activity_id FROM p6_activities
                )
            """)).fetchall()
            
            # 按 activity_id 排序，与 API 侧一致，从根源上避免死锁（统一加锁顺序）
            ids_to_delete = sorted([r[0] for r in ids_to_delete_res if r[0]])
            
            if ids_to_delete:
                # 2. 分段删除，每段 100 个 ID，每段 commit 一次
                # 这样可以缩短单次事务持锁时间，把路让给 API 填报
                total_to_clean = len(ids_to_delete)
                batch_size = 100
                for i in range(0, total_to_clean, batch_size):
                    batch_ids = ids_to_delete[i:i + batch_size]
                    db.execute(text("DELETE FROM activity_summary WHERE activity_id IN :ids"), {"ids": batch_ids})
                    db.commit()
                
                cleanup_time = time.time() - cleanup_start
                print(f"  ✓ 成功清理 {total_to_clean} 条过时记录（分 { (total_to_clean+batch_size-1)//batch_size } 批完成，耗时: {cleanup_time:.2f}秒）")
            else:
                cleanup_time = time.time() - cleanup_start
                print(f"  ✓ 没有需要清理的记录（耗时: {cleanup_time:.2f}秒）")
        
        # 步骤2: 检查数据源
        print("\n步骤2: 检查数据源...")
        
        # 在增量模式下，获取上次刷新时间
        last_refresh_time = None
        if not is_clear_mode:
            # 尝试从p6_sync_logs获取上次同步时间
            try:
                # 使用SQL查询，直接使用字符串值
                # 注意：列名是 sync_type 和 sync_status，不是 entity_type 和 status
                last_sync_log = db.execute(text("""
                    SELECT completed_at 
                    FROM p6_sync_logs 
                    WHERE sync_type = 'activity' 
                        AND sync_status = 'completed' 
                    ORDER BY completed_at DESC 
                    LIMIT 1
                """)).fetchone()
                
                if last_sync_log and last_sync_log[0]:
                    last_refresh_time = last_sync_log[0]
                    # 数据库中的completed_at是UTC，需要转换为GMT+3
                    last_refresh_time = ensure_gmt3(last_refresh_time)
                    print(f"  - 上次同步时间（GMT+3）: {format_time_gmt3(last_refresh_time)}")
            except Exception as e:
                print(f"  ⚠️ 无法从p6_sync_logs获取上次同步时间: {e}")
            
            # 如果p6_sync_logs没有，尝试从activity_summary的updated_at获取
            if not last_refresh_time:
                try:
                    max_updated = db.execute(text("""
                        SELECT MAX(updated_at) FROM activity_summary
                    """)).scalar()
                    if max_updated:
                        last_refresh_time = max_updated
                        # updated_at可能没有时区信息，假设是UTC或系统时区，转换为GMT+3
                        last_refresh_time = ensure_gmt3(last_refresh_time)
                        print(f"  - 上次刷新时间（GMT+3）: {format_time_gmt3(last_refresh_time)}")
                except Exception as e:
                    print(f"  ⚠️ 无法从activity_summary获取上次刷新时间: {e}")
            
            # 如果仍然没有找到，给出警告（增量模式需要上次刷新时间）
            if not last_refresh_time:
                print("  ⚠️ 警告：无法获取上次刷新时间，增量模式将处理所有记录（等同于全量重建）")
        
        p6_activities_count = db.execute(text("SELECT COUNT(*) FROM p6_activities WHERE is_active = 1")).scalar()
        print(f"  - p6_activities (is_active=1): {p6_activities_count} 条")
        
        # 步骤3: 确定需要处理的作业 ID 列表
        print("\n步骤3: 确定需要处理的作业 ID 列表...")
        target_ids = []
        if not is_clear_mode and last_refresh_time:
            last_refresh_time_gmt3 = ensure_gmt3(last_refresh_time)
            last_refresh_time_naive = last_refresh_time_gmt3.replace(tzinfo=None)
            last_time_with_overlap = last_refresh_time_naive - timedelta(minutes=10)
            
            print("  - 正在检索变动的作业 ID 列表...")
            changed_ids_query = db.execute(text("""
                SELECT activity_id FROM p6_activities WHERE is_active = 1 AND (p6_last_update_date > :last_time OR p6_last_update_date IS NULL)
                UNION
                SELECT activity_id FROM p6_activity_code_assignments WHERE is_active = 1 AND (p6_last_update_date > :last_time OR p6_last_update_date IS NULL) AND activity_id IS NOT NULL
                UNION
                SELECT activity_id FROM vfactdb WHERE updated_at > :last_time OR created_at > :last_time
                UNION
                SELECT activity_id FROM mpdb WHERE updated_at > :last_time OR created_at > :last_time
                UNION
                SELECT activity_id FROM volume_control_quantity WHERE updated_at > :last_time OR created_at > :last_time
            """), {"last_time": last_time_with_overlap}).fetchall()
            target_ids = [r[0] for r in changed_ids_query if r[0]]
            print(f"  - 增量模式：检测到 {len(target_ids)} 条变动作业")
        else:
            print("  - 全量模式：处理所有活跃作业...")
            target_ids_res = db.execute(text("SELECT activity_id FROM p6_activities WHERE is_active = 1")).fetchall()
            target_ids = [r[0] for r in target_ids_res if r[0]]

        if not target_ids:
            print("  - 没有检测到变动，跳过刷新。")
            return {'success': True, 'row_count': 0}

        # 步骤3.1: 准备聚合数据并插入临时表
        print("\n步骤3.1: 准备聚合数据...")
        SystemTaskService.check_and_abort_background_task(db)
        insert_start = time.time()
        
        # 优化：设置隔离级别为 READ UNCOMMITTED，彻底解决读取大表时的锁冲突
        db.execute(text("SET SESSION TRANSACTION ISOLATION LEVEL READ UNCOMMITTED"))
        db.commit()

        # 3.1 预先聚合 VFACTDB
        print("  - 预聚合 VFACTDB 数据...")
        db.execute(text("DROP TABLE IF EXISTS tmp_refresh_vfact_agg"))
        if not is_clear_mode and len(target_ids) < 50000: # 增量模式且变动不多时，只聚合变动部分
            db.execute(text("""
                CREATE TABLE tmp_refresh_vfact_agg
                SELECT activity_id, SUM(achieved) AS completed
                FROM vfactdb
                WHERE activity_id IN :ids
                GROUP BY activity_id
            """), {"ids": target_ids})
        else:
            db.execute(text("""
                CREATE TABLE tmp_refresh_vfact_agg
                SELECT activity_id, SUM(achieved) AS completed
                FROM vfactdb
                WHERE activity_id IS NOT NULL
                GROUP BY activity_id
            """))
        db.execute(text("ALTER TABLE tmp_refresh_vfact_agg ADD PRIMARY KEY (activity_id)"))

        # 3.2 预先聚合 MPDB
        print("  - 预聚合 MPDB 数据...")
        db.execute(text("DROP TABLE IF EXISTS tmp_refresh_mpdb_agg"))
        if not is_clear_mode and len(target_ids) < 50000:
            db.execute(text("""
                CREATE TABLE tmp_refresh_mpdb_agg
                SELECT activity_id, SUM(manpower) AS actual_manhour
                FROM mpdb
                WHERE activity_id IN :ids
                GROUP BY activity_id
            """), {"ids": target_ids})
        else:
            db.execute(text("""
                CREATE TABLE tmp_refresh_mpdb_agg
                SELECT activity_id, SUM(manpower) AS actual_manhour
                FROM mpdb
                WHERE activity_id IS NOT NULL
                GROUP BY activity_id
            """))
        db.execute(text("ALTER TABLE tmp_refresh_mpdb_agg ADD PRIMARY KEY (activity_id)"))

        # 3.3 预处理代码分配表 (Pivot 操作提前做，这是性能提升的关键)
        print("  - 预处理代码分配表 (Pivot 加速)...")
        db.execute(text("DROP TABLE IF EXISTS tmp_refresh_code_pivoted"))
        if not is_clear_mode and len(target_ids) < 50000:
            # 增量模式：只 Pivot 变动相关的代码
            db.execute(text("""
                CREATE TABLE tmp_refresh_code_pivoted
                SELECT 
                    p6a.object_id AS activity_object_id,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Block' THEN pca.activity_code_value END) AS block,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Discipline' THEN pca.activity_code_value END) AS discipline,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Work Package' THEN pca.activity_code_value END) AS work_package,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Scope' THEN pca.activity_code_value END) AS scope,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Phase' THEN pca.activity_code_value END) AS implement_phase,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Project' THEN pca.activity_code_value END) AS project,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Sub-project' THEN pca.activity_code_value END) AS subproject,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Train' THEN pca.activity_code_value END) AS train,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_Unit' THEN pca.activity_code_value END) AS unit,
                    MAX(CASE WHEN pca.activity_code_type_name = '!BCC_Quarter' THEN pca.activity_code_value END) AS quarter,
                    MAX(CASE WHEN pca.activity_code_type_name = 'GCC_SIMPBLK' THEN pca.activity_code_value END) AS simple_block,
                    MAX(CASE WHEN pca.activity_code_type_name = '!BCC_START-UP SEQUENCE' THEN pca.activity_code_value END) AS start_up_sequence,
                    MAX(CASE WHEN pca.activity_code_type_name = '!BCC_WORK PACKAGE' THEN pca.activity_code_value END) AS contract_phase
                FROM p6_activities p6a
                LEFT JOIN p6_activity_code_assignments pca ON p6a.object_id = pca.activity_object_id
                WHERE p6a.activity_id IN :ids AND pca.is_active = 1
                GROUP BY p6a.object_id
            """), {"ids": target_ids})
        else:
            db.execute(text("""
                CREATE TABLE tmp_refresh_code_pivoted
                SELECT 
                    activity_object_id,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Block' THEN activity_code_value END) AS block,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Discipline' THEN activity_code_value END) AS discipline,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Work Package' THEN activity_code_value END) AS work_package,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Scope' THEN activity_code_value END) AS scope,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Phase' THEN activity_code_value END) AS implement_phase,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Project' THEN activity_code_value END) AS project,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Sub-project' THEN activity_code_value END) AS subproject,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Train' THEN activity_code_value END) AS train,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_Unit' THEN activity_code_value END) AS unit,
                    MAX(CASE WHEN activity_code_type_name = '!BCC_Quarter' THEN activity_code_value END) AS quarter,
                    MAX(CASE WHEN activity_code_type_name = 'GCC_SIMPBLK' THEN activity_code_value END) AS simple_block,
                    MAX(CASE WHEN activity_code_type_name = '!BCC_START-UP SEQUENCE' THEN activity_code_value END) AS start_up_sequence,
                    MAX(CASE WHEN activity_code_type_name = '!BCC_WORK PACKAGE' THEN activity_code_value END) AS contract_phase
                FROM p6_activity_code_assignments
                WHERE is_active = 1
                GROUP BY activity_object_id
            """))
        db.execute(text("ALTER TABLE tmp_refresh_code_pivoted ADD PRIMARY KEY (activity_object_id)"))
        db.commit()

        # 3.5 执行分批聚合插入 (分批合并主数据)
        print(f"  - 开始分段合并主数据（共 {len(target_ids)} 条，每批 5000 条）...")
        batch_size = 5000
        processed_count = 0
        
        for i in range(0, len(target_ids), batch_size):
            batch = target_ids[i:i + batch_size]
            # 每批前检查：若有 API/用户正在对 vcq 或 activity_summary 持锁，本周期中止，避免 INSERT…SELECT 读 vcq 等锁导致 1205
            SystemTaskService.check_and_abort_background_task(db)
            
            # 增加心跳
            SystemTaskService.set_task_lock("background_refresh", True, updated_by="refresh_activity_summary_sql.py", quiet=True)

            # 增加内部重试逻辑
            for attempt in range(3):
                try:
                    db.execute(text("""
                        INSERT INTO activity_summary_temp (
                            activity_id, wbs_code, block, title, discipline, work_package, scope, implement_phase,
                            project, subproject, train, unit, main_block, quarter, 
                            simple_block, start_up_sequence,
                            key_qty, calculated_mhrs, resource_id, spe_mhrs, uom, contract_phase,
                            weight_factor, actual_weight_factor,
                            start_date, finish_date, planned_start_date, planned_finish_date, planned_duration, at_completion_duration,
                            baseline1_start_date, baseline1_finish_date,
                            actual_start_date, actual_finish_date, actual_duration, completed, actual_manhour,
                            data_date, type, system_status
                        )
                        SELECT 
                            p6a.activity_id, p6a.wbs_code, cp.block, p6a.name, cp.discipline, cp.work_package, cp.scope, cp.implement_phase,
                            COALESCE(cp.project, f.project), COALESCE(cp.subproject, f.subproject), COALESCE(cp.train, f.train), COALESCE(cp.unit, f.unit),
                            CASE WHEN cp.block IS NOT NULL AND LENGTH(cp.block) >= 10 THEN SUBSTRING(cp.block, 6, 5) ELSE f.main_block END,
                            COALESCE(cp.quarter, f.quarter), COALESCE(cp.simple_block, f.simple_block), COALESCE(cp.start_up_sequence, f.start_up_sequence),
                            vc.estimated_total,
                            CASE WHEN cp.implement_phase = 'CT' AND cp.contract_phase = 'Add.3' AND r.norms > 0 AND vc.estimated_total > 0 THEN 10 / r.norms * vc.estimated_total
                                 WHEN cp.implement_phase = 'CT' AND cp.contract_phase = 'Add.3' AND r.norms > 0 THEN 0 ELSE NULL END,
                            r.resource_id, r.norms, r.uom, cp.contract_phase,
                            NULL, NULL,
                            DATE(p6a.start_date), DATE(p6a.finish_date), DATE(p6a.planned_start_date), DATE(p6a.planned_finish_date),
                            CAST(p6a.planned_duration AS UNSIGNED), CAST(p6a.at_completion_duration AS UNSIGNED),
                            DATE(p6a.baseline1_start_date), DATE(p6a.baseline1_finish_date),
                            DATE(p6a.actual_start_date), DATE(p6a.actual_finish_date),
                            IF(p6a.actual_duration IS NOT NULL, CAST(p6a.actual_duration AS UNSIGNED), DATEDIFF(p6a.actual_finish_date, p6a.actual_start_date) + 1),
                            COALESCE(v.completed, 0), COALESCE(m.actual_manhour, 0),
                            p6a.data_date, p6a.type,
                            CASE WHEN p6a.actual_finish_date IS NOT NULL THEN 'Completed' WHEN v.completed > 0 OR m.actual_manhour > 0 THEN 'In Progress' ELSE 'Not Started' END
                        FROM p6_activities p6a
                        LEFT JOIN tmp_refresh_code_pivoted cp ON p6a.object_id = cp.activity_object_id
                        LEFT JOIN facilities f ON cp.block = f.block
                        LEFT JOIN volume_control_quantity vc ON p6a.activity_id = vc.activity_id
                        LEFT JOIN rsc_defines r ON cp.work_package = r.work_package
                        LEFT JOIN tmp_refresh_vfact_agg v ON p6a.activity_id = v.activity_id
                        LEFT JOIN tmp_refresh_mpdb_agg m ON p6a.activity_id = m.activity_id
                        WHERE p6a.activity_id IN :ids
                    """), {"ids": batch})
                    db.commit()
                    break
                except Exception as e:
                    db.rollback()
                    if ("1205" in str(e) or "1213" in str(e)) and attempt < 2:
                        print(f"    ⚠️ 批次同步冲突 (尝试 {attempt+1}/3)，等待 2 秒...")
                        time.sleep(2)
                        continue
                    raise e
            
            processed_count += len(batch)
            print(f"    - 已同步 {processed_count}/{len(target_ids)} 条作业...")

        insert_time = time.time() - insert_start
        print(f"  ✓ 聚合数据插入完成（耗时: {insert_time:.2f}秒）")
        row_count = processed_count

        # 步骤3.5: 根据 activity_status_records 修正状态和日期（确保用户手动确认的状态不被覆盖）
        print("\n步骤3.5: 根据用户手动确认的状态修正临时表数据...")
        db.execute(text("""
            UPDATE activity_summary_temp t
            INNER JOIN activity_status_records asr ON t.activity_id = asr.activity_id
            SET t.system_status = asr.status,
                t.actual_finish_date = COALESCE(asr.actual_finish_date, t.actual_finish_date)
        """))
        db.commit()
        print(f"  ✓ 成功插入并修正 {row_count} 条记录")
        
        # 步骤4: 更新/插入数据到activity_summary表
        if is_clear_mode:
            # 全量重建模式：原子性替换原表
            print("\n步骤4: 原子性替换原表...")
            replace_start = time.time()
            
            table_exists = db.execute(text("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE() 
                AND table_name = 'activity_summary'
            """)).scalar() > 0
            
            # 4.1: 查找所有引用 activity_summary 的外键约束（无论表是否存在）
            print("  - 查找外键约束...")
            foreign_keys = db.execute(text("""
                SELECT 
                    TABLE_NAME,
                    CONSTRAINT_NAME
                FROM information_schema.KEY_COLUMN_USAGE
                WHERE TABLE_SCHEMA = DATABASE()
                AND REFERENCED_TABLE_NAME = 'activity_summary'
                AND REFERENCED_COLUMN_NAME = 'activity_id'
            """)).fetchall()
            
            if foreign_keys:
                print(f"    - 找到 {len(foreign_keys)} 个外键约束")
            
            if table_exists:
                # 4.2: 删除所有外键约束
                if foreign_keys:
                    print("  - 删除外键约束...")
                    for table_name, constraint_name in foreign_keys:
                        try:
                            db.execute(text(f"ALTER TABLE {table_name} DROP FOREIGN KEY {constraint_name}"))
                            print(f"    ✓ 删除 {table_name} 表的外键: {constraint_name}")
                        except Exception as e:
                            print(f"    ⚠️ 删除 {table_name} 表的外键 {constraint_name} 失败: {e}")
                    
                    db.commit()
                
                # 4.3: 删除原表
                print("  - 删除原表...")
                db.execute(text("DROP TABLE activity_summary"))
                db.commit()
                print("    ✓ 原表已删除")
            
            # 4.4: 重命名临时表
            print("  - 重命名临时表...")
            db.execute(text("RENAME TABLE activity_summary_temp TO activity_summary"))
            db.commit()
            print("    ✓ 临时表已重命名为 activity_summary")
            
            # 4.5: 清理不匹配的记录（在清空重建模式下）
            # 在创建外键约束之前，先清理 mpdb 和 vfactdb 表中不存在于新 activity_summary 表的记录
            if foreign_keys:
                print("  - 清理不匹配的记录（将不存在的 activity_id 设置为 NULL）...")
                for table_name, constraint_name in foreign_keys:
                    try:
                        # 将不存在于 activity_summary 的 activity_id 设置为 NULL
                        # 使用 LEFT JOIN 方式，性能更好
                        cleanup_result = db.execute(text(f"""
                            UPDATE {table_name} t
                            LEFT JOIN activity_summary a ON t.activity_id = a.activity_id
                            SET t.activity_id = NULL
                            WHERE t.activity_id IS NOT NULL
                            AND a.activity_id IS NULL
                        """))
                        cleaned_count = cleanup_result.rowcount
                        if cleaned_count > 0:
                            print(f"    ✓ {table_name}: 清理了 {cleaned_count} 条不匹配的记录（activity_id 设置为 NULL）")
                        else:
                            print(f"    ✓ {table_name}: 没有不匹配的记录")
                    except Exception as e:
                        print(f"    ⚠️ 清理 {table_name} 表失败: {e}")
                
                db.commit()
            
            # 4.6: 重新创建外键约束
            if foreign_keys:
                print("  - 重新创建外键约束...")
                for table_name, constraint_name in foreign_keys:
                    try:
                        # 构建新的外键名称（使用统一的命名规则）
                        new_fk_name = f"fk_{table_name}_activity_summary"
                        db.execute(text(f"""
                            ALTER TABLE {table_name}
                            ADD CONSTRAINT {new_fk_name}
                            FOREIGN KEY (activity_id)
                            REFERENCES activity_summary(activity_id)
                            ON DELETE SET NULL
                            ON UPDATE CASCADE
                        """))
                        print(f"    ✓ 创建 {table_name} 表的外键: {new_fk_name}")
                    except Exception as e:
                        print(f"    ⚠️ 创建 {table_name} 表的外键失败: {e}")
                db.commit()
            
            # 全量模式下也执行一次补全，确保万无一失
            print("  - 正在检查并补全 data_date...")
            try:
                db.execute(text("""
                    UPDATE activity_summary a
                    INNER JOIN p6_activities p ON a.activity_id = p.activity_id
                    SET a.data_date = p.data_date
                    WHERE a.data_date IS NULL AND p.data_date IS NOT NULL
                """))
                db.commit()
            except: pass

            replace_time = time.time() - replace_start
            print(f"  ✓ 表替换完成（耗时: {replace_time:.2f}秒）")
        else:
            # 增量更新模式：使用UPDATE/INSERT（避免表替换，大幅提升性能）
            print("\n步骤4: 增量更新activity_summary表（使用UPDATE/INSERT）...")
            SystemTaskService.check_and_abort_background_task(db)
            replace_start = time.time()
            
            # 检查activity_summary表是否存在
            table_exists = db.execute(text("""
                SELECT COUNT(*) 
                FROM information_schema.tables 
                WHERE table_schema = DATABASE() 
                AND table_name = 'activity_summary'
            """)).scalar() > 0
            
            if not table_exists:
                # 如果表不存在，直接重命名临时表（首次运行）
                print("  - activity_summary表不存在，直接重命名临时表...")
                db.execute(text("RENAME TABLE activity_summary_temp TO activity_summary"))
                db.commit()
                print("    ✓ 临时表已重命名为 activity_summary")
                replace_time = time.time() - replace_start
                print(f"  ✓ 表创建完成（耗时: {replace_time:.2f}秒）")
            else:
                # 补全既存数据的 data_date (针对增量模式下未触达的旧记录)
                # 优化：采用分批更新，避免锁超时
                print("  - 正在分批补全既存数据的 data_date...")
                try:
                    # 查出需要更新的 ID
                    null_date_ids_res = db.execute(text("""
                        SELECT a.activity_id FROM activity_summary a
                        INNER JOIN p6_activities p ON a.activity_id = p.activity_id
                        WHERE a.data_date IS NULL AND p.data_date IS NOT NULL
                    """)).fetchall()
                    # 按 activity_id 排序，统一加锁顺序，避免死锁
                    null_date_ids = sorted([r[0] for r in null_date_ids_res if r[0]])
                    
                    if null_date_ids:
                        for j in range(0, len(null_date_ids), 500):
                            batch_ids = null_date_ids[j:j+500]
                            db.execute(text("""
                                UPDATE activity_summary a
                                INNER JOIN p6_activities p ON a.activity_id = p.activity_id
                                SET a.data_date = p.data_date
                                WHERE a.activity_id IN :ids
                            """), {"ids": batch_ids})
                            db.commit()
                        print(f"    ✓ 已成功补齐 {len(null_date_ids)} 条旧记录的 data_date")
                except Exception as e:
                    print(f"  ⚠️ 补齐 data_date 失败: {e}")

                # 使用分批 INSERT ... ON DUPLICATE KEY UPDATE 进行增量更新 (防止锁超时)
                print("  - 使用分批 INSERT ... ON DUPLICATE KEY UPDATE 进行增量更新...")
                
                # 获取临时表中的所有作业 ID
                temp_ids_res = db.execute(text("SELECT activity_id FROM activity_summary_temp")).fetchall()
                # 按 activity_id 排序，统一加锁顺序，避免与 API 交叉加锁导致死锁
                temp_ids = sorted([r[0] for r in temp_ids_res if r[0]])
                
                if temp_ids:
                    total_to_upsert = len(temp_ids)
                    upsert_batch_size = 500
                    for j in range(0, total_to_upsert, upsert_batch_size):
                        batch_ids = temp_ids[j:j + upsert_batch_size]
                        # 增加心跳
                        SystemTaskService.set_task_lock("background_refresh", True, updated_by="refresh_activity_summary_sql.py", quiet=True)
                        for attempt in range(3):
                            try:
                                db.execute(text("""
                                    INSERT INTO activity_summary (
                                        activity_id, wbs_code, block, title, discipline, work_package, scope, implement_phase,
                                        project, subproject, train, unit, main_block, quarter, 
                                        simple_block, start_up_sequence,
                                        key_qty, calculated_mhrs, resource_id, spe_mhrs, uom, contract_phase,
                                        weight_factor, actual_weight_factor,
                                        start_date, finish_date, planned_start_date, planned_finish_date, planned_duration, at_completion_duration,
                                        baseline1_start_date, baseline1_finish_date,
                                        actual_start_date, actual_finish_date, actual_duration, completed, actual_manhour,
                                        data_date, type, system_status, updated_at
                                    )
                                    SELECT 
                                        activity_id, wbs_code, block, title, discipline, work_package, scope, implement_phase,
                                        project, subproject, train, unit, main_block, quarter, 
                                        simple_block, start_up_sequence,
                                        key_qty, calculated_mhrs, resource_id, spe_mhrs, uom, contract_phase,
                                        weight_factor, actual_weight_factor,
                                        start_date, finish_date, planned_start_date, planned_finish_date, planned_duration, at_completion_duration,
                                        baseline1_start_date, baseline1_finish_date,
                                        actual_start_date, actual_finish_date, actual_duration, completed, actual_manhour,
                                        data_date, type, system_status, NOW()
                                    FROM activity_summary_temp
                                    WHERE activity_id IN :batch_ids
                                    ON DUPLICATE KEY UPDATE
                                        wbs_code = VALUES(wbs_code),
                                        block = VALUES(block),
                                        title = VALUES(title),
                                        discipline = VALUES(discipline),
                                        work_package = VALUES(work_package),
                                        scope = VALUES(scope),
                                        implement_phase = VALUES(implement_phase),
                                        project = VALUES(project),
                                        subproject = VALUES(subproject),
                                        train = VALUES(train),
                                        unit = VALUES(unit),
                                        main_block = VALUES(main_block),
                                        quarter = VALUES(quarter),
                                        simple_block = VALUES(simple_block),
                                        start_up_sequence = VALUES(start_up_sequence),
                                        key_qty = VALUES(key_qty),
                                        calculated_mhrs = VALUES(calculated_mhrs),
                                        resource_id = VALUES(resource_id),
                                        spe_mhrs = VALUES(spe_mhrs),
                                        uom = VALUES(uom),
                                        contract_phase = VALUES(contract_phase),
                                        weight_factor = VALUES(weight_factor),
                                        actual_weight_factor = VALUES(actual_weight_factor),
                                        start_date = VALUES(start_date),
                                        finish_date = VALUES(finish_date),
                                        planned_start_date = VALUES(planned_start_date),
                                        planned_finish_date = VALUES(planned_finish_date),
                                        planned_duration = VALUES(planned_duration),
                                        at_completion_duration = VALUES(at_completion_duration),
                                        baseline1_start_date = VALUES(baseline1_start_date),
                                        baseline1_finish_date = VALUES(baseline1_finish_date),
                                        actual_start_date = VALUES(actual_start_date),
                                        actual_finish_date = VALUES(actual_finish_date),
                                        actual_duration = VALUES(actual_duration),
                                        completed = VALUES(completed),
                                        actual_manhour = VALUES(actual_manhour),
                                        data_date = VALUES(data_date),
                                        type = VALUES(type),
                                        system_status = VALUES(system_status),
                                        updated_at = NOW()
                                """), {"batch_ids": batch_ids})
                                db.commit()
                                break
                            except Exception as e:
                                db.rollback()
                                if "1205" in str(e) or "1213" in str(e):
                                    if attempt < 2:
                                        time.sleep(1 + attempt)
                                        continue
                                raise e
                        print(f"    - 已同步到主表 {min(j + upsert_batch_size, total_to_upsert)}/{total_to_upsert} 条...")
                
                # 保存变更的activity_id到临时表（用于步骤5的权重因子更新）
                db.execute(text("""
                    CREATE TEMPORARY TABLE IF NOT EXISTS temp_changed_activity_ids (
                        activity_id VARCHAR(100) PRIMARY KEY
                    )
                """))
                db.execute(text("TRUNCATE TABLE temp_changed_activity_ids"))
                db.execute(text("""
                    INSERT INTO temp_changed_activity_ids (activity_id)
                    SELECT activity_id FROM activity_summary_temp
                """))
                db.commit()
                
                replace_time = time.time() - replace_start
                print(f"  ✓ 增量更新完成（耗时: {replace_time:.2f}秒）")
        
        # 步骤5: 处理特殊情况并计算权重因子（weight_factor）
        print("\n步骤5: 处理特殊情况并计算权重因子...")
        weight_start = time.time()
        
        # 权重因子基数（默认值，可以从配置表读取）
        WEIGHT_FACTOR_BASE = 254137500
        
        # 5.1: 处理特殊情况 - PI08 的人工时计算
        # VBA 逻辑分析：
        # 1. 第 184-187 行：按 gcc_simpblk@workpackage 分组累加所有记录的 result(i, 21)，没有过滤条件
        # 2. 第 194-205 行：对于 PI08，从字典中查找同一 gcc_simpblk 下的 PI05 作业
        #    关键：VBA 代码中，字典包含了所有记录的累加值，不管 phase 和 bcc_work_package
        #    但是，只有匹配到 wkPKG 的记录才会有 result(i, 21) 的值
        # 3. 所以，PI08 计算时，应该使用所有 PI05 记录（不管 phase 和 bcc_work_package）
        #    但我们的业务需求是只计算满足条件的记录，所以这里保持过滤条件
        # 4. 但是，VBA 代码中，第 197 行使用的是 Left(k(j), 7) 来匹配 gcc_simpblk
        #    这意味着匹配逻辑可能不完全精确
        print("  - 处理特殊情况：PI08 人工时计算...")
        # 优化：先将 PI05 的聚合结果存入物理临时表，避免在 UPDATE 时对 activity_summary 加读锁，彻底解决死锁
        try:
            db.execute(text("DROP TABLE IF EXISTS tmp_pi05_totals"))
            db.execute(text("""
                CREATE TABLE tmp_pi05_totals AS
                SELECT 
                    simple_block,
                    COALESCE(SUM(COALESCE(calculated_mhrs, 0)), 0) AS pi05_total_mhrs
                FROM activity_summary
                WHERE work_package = 'PI05'
                    AND implement_phase = 'CT'
                    AND contract_phase = 'Add.3'
                GROUP BY simple_block
            """))
            db.execute(text("ALTER TABLE tmp_pi05_totals ADD INDEX idx_simple_block (simple_block)"))
            db.commit()
            
            # 分批更新 PI08，从临时表读取数据
            pi08_ids_res = db.execute(text("""
                SELECT activity_id FROM activity_summary 
                WHERE work_package = 'PI08' AND implement_phase = 'CT' AND contract_phase = 'Add.3'
            """)).fetchall()
            # 按 activity_id 排序，统一加锁顺序，避免与 API 交叉加锁导致死锁
            pi08_ids = sorted([r[0] for r in pi08_ids_res if r[0]])
            
            pi08_updated = 0
            if pi08_ids:
                for j in range(0, len(pi08_ids), 200): # 增加每批次数量
                    batch_ids = pi08_ids[j:j+200]
                    db.execute(text("""
                        UPDATE activity_summary a1
                        INNER JOIN tmp_pi05_totals ON a1.simple_block = tmp_pi05_totals.simple_block
                        SET a1.calculated_mhrs = tmp_pi05_totals.pi05_total_mhrs * 0.3
                        WHERE a1.activity_id IN :ids
                    """), {"ids": batch_ids})
                    db.commit()
                    pi08_updated += len(batch_ids)
            
            # 清理临时表
            db.execute(text("DROP TABLE IF EXISTS tmp_pi05_totals"))
            db.commit()
            
            if pi08_updated > 0:
                print(f"    ✓ 更新了 {pi08_updated} 条 PI08 作业的人工时（中间表分批模式）")
        except Exception as pi08_err:
            print(f"    ⚠️ 更新 PI08 失败: {pi08_err}")
            db.rollback()
        
        # 5.2: 按 project 分组计算项目总人工时
        # 注意：VBA 代码中计算项目总人工时时没有过滤条件，对所有记录都累加
        # 但根据业务需求，weight_factor 只对满足条件的记录计算
        # 所以项目总人工时也应该只计算满足条件的记录
        print("  - 按 project 分组计算项目总人工时...")
        project_totals = db.execute(text("""
            SELECT 
                project,
                COALESCE(SUM(calculated_mhrs), 0) AS total_mhrs
            FROM activity_summary
            WHERE implement_phase = 'CT' 
                AND contract_phase = 'Add.3'
                AND calculated_mhrs IS NOT NULL
                AND calculated_mhrs > 0
                AND project IS NOT NULL
            GROUP BY project
        """)).fetchall()
        
        if project_totals:
            print(f"    ✓ 找到 {len(project_totals)} 个项目")
            for proj_row in project_totals:
                project_name = proj_row[0]
                project_total_mhrs = float(proj_row[1])
                print(f"      - {project_name}: {project_total_mhrs:,.2f} 人工时")
        else:
            print("    ⚠️ 没有找到满足条件的项目")
        
        # 5.3: 更新权重因子（按 project 分组计算）
        print("  - 更新权重因子...")
        SystemTaskService.check_and_abort_background_task(db)
        if project_totals:
            updated_count = 0
            for proj_row in project_totals:
                project_name = proj_row[0]
                project_total_mhrs = float(proj_row[1])
                
                if project_total_mhrs > 0:
                    # 获取该项目下需要更新的 ID 列表
                    act_ids_res = db.execute(text("""
                        SELECT activity_id FROM activity_summary
                        WHERE implement_phase = 'CT' AND contract_phase = 'Add.3' AND project = :project_name
                    """), {"project_name": project_name}).fetchall()
                    # 按 activity_id 排序，统一加锁顺序，从根源上避免死锁
                    act_ids = sorted([r[0] for r in act_ids_res if r[0]])
                    
                    if act_ids:
                        # 分批更新，每批 200 条（减小批次以减少锁定时间），防止长事务死锁
                        for k in range(0, len(act_ids), 200):
                            batch_ids = act_ids[k:k+200]
                            # 增加心跳
                            SystemTaskService.set_task_lock("background_refresh", True, updated_by="refresh_activity_summary_sql.py", quiet=True)
                            # 增加内部重试逻辑，确保单批次失败不影响全局
                            for attempt in range(3):
                                try:
                                    db.execute(text("""
                                        UPDATE activity_summary
                                        SET weight_factor = (calculated_mhrs / :total_mhrs * :base)
                                        WHERE activity_id IN :ids
                                    """), {
                                        "total_mhrs": project_total_mhrs,
                                        "base": WEIGHT_FACTOR_BASE,
                                        "ids": batch_ids
                                    })
                                    db.commit()
                                    break
                                except Exception as e:
                                    db.rollback()
                                    if "1205" in str(e) or "1213" in str(e):
                                        if attempt < 2:
                                            time.sleep(1 + attempt)
                                            continue
                                    raise e
                            updated_count += len(batch_ids)
            
            weight_time = time.time() - weight_start
            print(f"  ✓ 成功分批更新 {updated_count} 条记录的权重因子（耗时: {weight_time:.2f}秒）")
            
            # 5.3b: 用预设权重表（OWF Weight Factor.xlsx）覆盖匹配 activity_id 的 weight_factor
            preset_wf = load_preset_weight_factors()
            if preset_wf:
                print(f"  - 应用预设权重因子（共 {len(preset_wf)} 条）...")
                try:
                    db.execute(text("DROP TABLE IF EXISTS tmp_preset_weight_factor"))
                    db.execute(text("""
                        CREATE TABLE tmp_preset_weight_factor (
                            activity_id VARCHAR(100) PRIMARY KEY,
                            weight_factor DECIMAL(18, 2) NOT NULL
                        )
                    """))
                    items = list(preset_wf.items())
                    for i in range(0, len(items), 200):
                        batch = items[i:i+200]
                        db.execute(text("""
                            INSERT INTO tmp_preset_weight_factor (activity_id, weight_factor)
                            VALUES (:aid, :wf)
                        """), [{"aid": aid, "wf": wf} for aid, wf in batch])
                    db.commit()
                    preset_updated = db.execute(text("""
                        UPDATE activity_summary a
                        INNER JOIN tmp_preset_weight_factor t ON a.activity_id = t.activity_id
                        SET a.weight_factor = t.weight_factor
                    """)).rowcount
                    db.commit()
                    db.execute(text("DROP TABLE IF EXISTS tmp_preset_weight_factor"))
                    db.commit()
                    print(f"    ✓ 已用预设表覆盖 {preset_updated} 条记录的 weight_factor")
                except Exception as preset_err:
                    print(f"    ⚠️ 应用预设权重失败: {preset_err}")
                    db.rollback()
                    try:
                        db.execute(text("DROP TABLE IF EXISTS tmp_preset_weight_factor"))
                        db.commit()
                    except Exception:
                        pass
            
            # 步骤5.1: 计算实际权重因子（actual_weight_factor）
            print("\n步骤5.1: 计算实际权重因子（actual_weight_factor）...")
            actual_weight_start = time.time()
            
            actual_weight_updated_count = 0
            if project_totals:
                # 按项目分组更新 actual_weight_factor
                for proj_row in project_totals:
                    project_name = proj_row[0]
                    project_total_mhrs = float(proj_row[1])
                    
                    if project_total_mhrs > 0:
                        # 获取需要更新的 ID
                        act_ids_res = db.execute(text("""
                            SELECT activity_id FROM activity_summary
                            WHERE implement_phase = 'CT' AND contract_phase = 'Add.3' AND project = :project_name
                              AND spe_mhrs IS NOT NULL AND spe_mhrs > 0 AND completed IS NOT NULL AND completed > 0
                        """), {"project_name": project_name}).fetchall()
                        # 按 activity_id 排序，统一加锁顺序，避免死锁
                        act_ids = sorted([r[0] for r in act_ids_res if r[0]])
                        
                        if act_ids:
                            # 分批更新，防止长事务死锁
                            for k in range(0, len(act_ids), 200):
                                batch_ids = act_ids[k:k+200]
                                # 增加心跳
                                SystemTaskService.set_task_lock("background_refresh", True, updated_by="refresh_activity_summary_sql.py", quiet=True)
                                for attempt in range(3):
                                    try:
                                        db.execute(text("""
                                            UPDATE activity_summary
                                            SET actual_weight_factor = ((10 / spe_mhrs * completed) / :total_mhrs * :base)
                                            WHERE activity_id IN :ids
                                        """), {
                                            "total_mhrs": project_total_mhrs,
                                            "base": WEIGHT_FACTOR_BASE,
                                            "ids": batch_ids
                                        })
                                        db.commit()
                                        break
                                    except Exception as e:
                                        db.rollback()
                                        if "1205" in str(e) or "1213" in str(e):
                                            if attempt < 2:
                                                time.sleep(1 + attempt)
                                                continue
                                        raise e
                                actual_weight_updated_count += len(batch_ids)
            
            actual_weight_time = time.time() - actual_weight_start
            print(f"  ✓ 成功分批更新 {actual_weight_updated_count} 条记录的实际权重因子（耗时: {actual_weight_time:.2f}秒）")

            # 步骤 6: 同步更新 volume_control_quantity
            print("\n步骤 6: 同步更新 volume_control_quantity 完工量 (安全分批模式)...")
            vcq_start = time.time()
            
            # 增量模式下，仅同步变动的作业到 VCQ
            # 全量模式下，同步所有作业
            if not is_clear_mode and target_ids:
                sync_ids = target_ids
            else:
                all_ids_res = db.execute(text("SELECT activity_id FROM activity_summary")).fetchall()
                sync_ids = [r[0] for r in all_ids_res if r[0]]
            
            if sync_ids:
                total_sync = len(sync_ids)
                # 使用较小的批次（1000条），确保即便是大批量同步也不会长时间持锁
                batch_size = 1000
                for j in range(0, total_sync, batch_size):
                    batch_ids = sync_ids[j:j+batch_size]
                    # 增加心跳
                    SystemTaskService.set_task_lock("background_refresh", True, updated_by="refresh_activity_summary_sql.py", quiet=True)
                    for attempt in range(3):
                        try:
                            # 直接从 vfactdb 聚合同步到 volume_control_quantity，以保持 20 位精度
                            # 避免从 activity_summary.completed (只有2位精度) 同步导致精度丢失
                            db.execute(text("""
                                INSERT INTO volume_control_quantity (activity_id, construction_completed, construction_completed_updated_at, updated_at)
                                SELECT activity_id, SUM(achieved), NOW(), NOW() 
                                FROM vfactdb 
                                WHERE activity_id IN :ids
                                GROUP BY activity_id
                                ON DUPLICATE KEY UPDATE 
                                    construction_completed = VALUES(construction_completed),
                                    construction_completed_updated_at = NOW(),
                                    updated_at = NOW()
                            """), {"ids": batch_ids})
                            db.commit()
                            break
                        except Exception as e:
                            db.rollback()
                            if "1205" in str(e) or "1213" in str(e):
                                if attempt < 2:
                                    time.sleep(1 + attempt)
                                    continue
                            raise e
                print(f"  ✓ 成功对齐 {total_sync} 条 volume_control_quantity 完工量（耗时: {time.time() - vcq_start:.2f}秒）")

        else:
            weight_time = time.time() - weight_start
            print(f"  ⚠️ 没有满足条件的项目，跳过权重因子计算（耗时: {weight_time:.2f}秒）")
            updated_count = 0
            
            # 直接从 vfactdb 同步到 volume_control_quantity (补漏)
            db.execute(text("""
                INSERT INTO volume_control_quantity (activity_id, construction_completed, construction_completed_updated_at, updated_at)
                SELECT activity_id, SUM(achieved), NOW(), NOW() 
                FROM vfactdb
                GROUP BY activity_id
                ON DUPLICATE KEY UPDATE construction_completed = VALUES(construction_completed), updated_at = NOW()
            """))
            db.commit()
        
        # 清理物理中间表
        print("\n清理中间计算表...")
        try:
            db.execute(text("DROP TABLE IF EXISTS tmp_refresh_vfact_agg"))
            db.execute(text("DROP TABLE IF EXISTS tmp_refresh_mpdb_agg"))
            db.execute(text("DROP TABLE IF EXISTS tmp_refresh_code_pivoted"))
            db.execute(text("DROP TABLE IF EXISTS activity_summary_temp"))
            db.execute(text("DROP TEMPORARY TABLE IF EXISTS temp_changed_activity_ids"))
            db.commit()
            print("  ✓ 中间表已清理")
        except Exception as e:
            pass

        # 步骤 7: 刷新 facility_filter_options 预聚合表（与 activity_summary 联动）
        print("\n步骤 7: 刷新 facility_filter_options 预聚合表...")
        refresh_facility_filter_options(db)

        total_time = time.time() - start_time
        
        print("\n" + "=" * 60)
        print("数据刷新完成（SQL版本）")
        print("=" * 60)
        print(f"总耗时: {total_time:.2f}秒")
        print(f"  - SQL聚合插入: {insert_time:.2f}秒")
        print(f"  - 表替换: {replace_time:.2f}秒")
        if project_total_mhrs > 0:
            print(f"  - 权重因子计算: {weight_time:.2f}秒")
        print(f"  - 处理记录数: {row_count} 条")
        if row_count > 0:
            print(f"  - 平均速度: {row_count / total_time:.0f} 条/秒")
        print("=" * 60)
        
        return {
            'success': True,
            'total_time': total_time,
            'insert_time': insert_time,
            'row_count': row_count
        }
        
    except InterruptedError as e:
        db.rollback()
        print("\n[避让] 检测到其他任务正在操作汇总表，本周期中止，等待下次调度: %s" % (e,))
        return {'success': True, 'row_count': 0, 'skipped': True}
    except Exception as e:
        db.rollback()
        print(f"\n❌ 错误: {e}")
        print(f"详细信息: {traceback.format_exc()}")
        
        # 清理物理中间表
        try:
            print("\n清理中间计算表...")
            db.execute(text("DROP TABLE IF EXISTS tmp_refresh_vfact_agg"))
            db.execute(text("DROP TABLE IF EXISTS tmp_refresh_mpdb_agg"))
            db.execute(text("DROP TABLE IF EXISTS tmp_refresh_code_pivoted"))
            db.execute(text("DROP TABLE IF EXISTS activity_summary_temp"))
            db.commit()
            print("  ✓ 中间计算表已清理，原表保持不变")
        except Exception as cleanup_error:
            print(f"  ⚠️ 清理中间计算表失败: {cleanup_error}")
        
        # 死锁/锁超时(1205,1213)：先输出当前锁/阻塞诊断到日志和 stdout，再重新抛出
        if isinstance(e, OperationalError) and e.orig and getattr(e.orig, 'args', None):
            if len(e.orig.args) >= 1 and e.orig.args[0] in (1205, 1213):
                try:
                    engine = get_default_engine()
                    diag = get_lock_diagnostics(engine)
                    msg = "\n[锁诊断] refresh_activity_summary_sql 失败 (1205/1213)，当前锁等待/阻塞信息（谁锁住了这些 activity_id）:\n" + diag
                    print(msg)
                    logging.getLogger(__name__).warning("[锁诊断] refresh_activity_summary_sql 整次刷新失败 (1205/1213)\n%s", diag)
                except Exception as dex:
                    print(f"[锁诊断] 无法输出锁诊断: {dex}")
                raise
        return {
            'success': False,
            'error': str(e)
        }
    finally:
        SystemTaskService.set_task_lock("background_refresh", False)
        db.close()


if __name__ == "__main__":
    result = refresh_activity_summary_sql()
    sys.exit(0 if result.get('success') else 1)

