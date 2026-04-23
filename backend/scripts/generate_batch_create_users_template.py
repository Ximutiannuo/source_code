"""
生成 batch_create_users 的 Excel 模板文件

执行方式：
  cd backend && python -m scripts.generate_batch_create_users_template
  或
  python backend/scripts/generate_batch_create_users_template.py [输出路径]

默认输出：D:/batch_create_users_template.xlsx
"""
import sys
import os
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "backend"))

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 默认输出路径
DEFAULT_OUTPUT = r"D:/batch_create_users_template.xlsx"


def generate_template(output_path: str):
    wb = openpyxl.Workbook()

    # 1. 单表格式（默认活动表）
    ws_single = wb.active
    ws_single.title = "users (单表)"
    headers_single = [
        "username",
        "email",
        "password_default",
        "full_name",
        "is_active",
        "is_superuser",
        "role.name",
        "department",
        "responsible_for",
    ]
    for col, h in enumerate(headers_single, start=1):
        cell = ws_single.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    # 示例行
    ws_single.cell(row=2, column=1, value="示例用户")
    ws_single.cell(row=2, column=2, value="example@cc7.cn")
    ws_single.cell(row=2, column=3, value="ChangeMe123")
    ws_single.cell(row=2, column=4, value="张三")
    ws_single.cell(row=2, column=5, value="1")
    ws_single.cell(row=2, column=6, value="0")
    ws_single.cell(row=2, column=7, value="计划经理")
    ws_single.cell(row=2, column=8, value="design")
    ws_single.cell(row=2, column=9, value="采购对接、设计审批")
    # 列宽
    for col in range(1, len(headers_single) + 1):
        ws_single.column_dimensions[get_column_letter(col)].width = 16

    # 2. 三表格式
    # new roles
    ws_roles = wb.create_sheet("new roles")
    headers_roles = ["id", "name", "description", "is_active"]
    for col, h in enumerate(headers_roles, start=1):
        cell = ws_roles.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    ws_roles.cell(row=2, column=1, value="1")
    ws_roles.cell(row=2, column=2, value="示例角色")
    ws_roles.cell(row=2, column=3, value="示例说明")
    ws_roles.cell(row=2, column=4, value="1")

    # new accounts
    ws_acc = wb.create_sheet("new accounts")
    headers_acc = ["id", "username", "email", "default_password", "full_name", "is_active", "is_superuser", "department", "responsible_for"]
    for col, h in enumerate(headers_acc, start=1):
        cell = ws_acc.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ws_acc.cell(row=2, column=1, value="1")
    ws_acc.cell(row=2, column=2, value="demo_user")
    ws_acc.cell(row=2, column=3, value="demo@cc7.cn")
    ws_acc.cell(row=2, column=4, value="ChangeMe123")
    ws_acc.cell(row=2, column=5, value="演示用户")
    ws_acc.cell(row=2, column=6, value="1")
    ws_acc.cell(row=2, column=7, value="0")
    ws_acc.cell(row=2, column=8, value="design")
    ws_acc.cell(row=2, column=9, value="设计对接")
    for col in range(1, len(headers_acc) + 1):
        ws_acc.column_dimensions[get_column_letter(col)].width = 14

    # user_roles
    ws_ur = wb.create_sheet("user_roles")
    headers_ur = ["user_id", "role_id"]
    for col, h in enumerate(headers_ur, start=1):
        cell = ws_ur.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws_ur.cell(row=2, column=1, value="1")
    ws_ur.cell(row=2, column=2, value="1")

    # 3. departments 清单说明
    ws_dept = wb.create_sheet("departments_list")
    ws_dept.cell(row=1, column=1, value="department.code")
    ws_dept.cell(row=1, column=2, value="department.name")
    dept_list = [
        ("design", "设计管理部"),
        ("procurement", "采购管理部"),
        ("planning", "计划管理部"),
        ("external_user", "外部用户"),
        ("project_leader", "项目领导"),
    ]
    for row_idx, (code, name) in enumerate(dept_list, start=2):
        ws_dept.cell(row=row_idx, column=1, value=code)
        ws_dept.cell(row=row_idx, column=2, value=name)
    ws_dept.column_dimensions["A"].width = 20
    ws_dept.column_dimensions["B"].width = 16

    wb.save(output_path)
    print(f"✅ 模板已生成: {output_path}")
    print("  - 单表格式：users (单表)，填 username, email, password_default, full_name, role.name, department, responsible_for")
    print("  - 三表格式：new roles / new accounts / user_roles")
    print("  - 部门参考：departments_list 表")
    print("  执行导入：python -m scripts.batch_create_users <excel路径> [--dry-run]")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="生成 batch_create_users 的 Excel 模板")
    parser.add_argument("output_path", type=str, nargs="?", default=DEFAULT_OUTPUT, help=f"输出路径，默认 {DEFAULT_OUTPUT}")
    args = parser.parse_args()

    output = args.output_path if os.path.isabs(args.output_path) else os.path.abspath(args.output_path)
    os.makedirs(os.path.dirname(output) or ".", exist_ok=True)

    print("=" * 60)
    print("生成批量创建用户模板")
    print("=" * 60)
    try:
        generate_template(output)
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
