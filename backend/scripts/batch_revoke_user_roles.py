"""
撤销由「批量创建用户」脚本根据指定 Excel 分配的角色。

用法：
  python backend/scripts/batch_revoke_user_roles.py "D:\new_roles and new_accounts.xlsx"

会读取同一 Excel 的 new roles / new accounts / user_roles，按与 batch_create_users 相同的
user_id/role_id 解析逻辑，只撤销这些 (用户, 角色) 的分配，然后你可修改 Excel 中的 role.id
再重新执行：
  python backend/scripts/batch_create_users.py "D:\new_roles and new_accounts.xlsx"

注意：只撤销本 Excel 中 user_roles 表里列出的分配，不会动其他用户或其它角色。
"""
import sys
import os

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app.database import load_env_with_fallback
load_env_with_fallback()

import openpyxl
from app.database import SessionLocal
from app.models.user import User, Role

SHEET_NEW_ROLES = "new roles"
SHEET_NEW_ACCOUNTS = "new accounts"
SHEET_USER_ROLES = "user_roles"


def _col_index(ws, header_name):
    headers = [cell.value for cell in ws[1]]
    for idx, h in enumerate(headers, start=1):
        if h and str(h).strip().lower() == str(header_name).strip().lower():
            return idx
    return None


def _cell_value(row, col_1based):
    if col_1based is None or col_1based < 1:
        return None
    idx = col_1based - 1
    if idx >= len(row):
        return None
    cell = row[idx]
    return getattr(cell, "value", cell)


def batch_revoke_user_roles(excel_path: str, dry_run: bool = False):
    """
    按 Excel 中 user_roles 的 (user_id, role_id) 撤销角色分配。
    user_id/role_id 解析逻辑与 batch_create_users 一致（先 Excel 表内 id，再数据库 id）。
    """
    print(f"正在读取: {excel_path}")
    if dry_run:
        print("【dry-run 模式】只打印将撤销的分配，不写入数据库。\n")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    db = SessionLocal()
    try:
        role_by_excel_id = {}
        user_by_excel_id = {}

        # ---------- 1. new roles：只解析 Excel id -> Role（不创建）----------
        if SHEET_NEW_ROLES in wb.sheetnames:
            ws_roles = wb[SHEET_NEW_ROLES]
            col_id = _col_index(ws_roles, "id")
            col_name = _col_index(ws_roles, "name")
            if col_name:
                for row_idx in range(2, ws_roles.max_row + 1):
                    row = list(ws_roles[row_idx])
                    excel_id = _cell_value(row, col_id)
                    name = _cell_value(row, col_name)
                    if not name:
                        continue
                    name = str(name).strip()
                    role = db.query(Role).filter(Role.name == name).first()
                    if role and excel_id is not None:
                        try:
                            role_by_excel_id[int(float(excel_id))] = role
                        except (TypeError, ValueError):
                            pass

        # ---------- 2. new accounts：只解析 Excel id -> User（不创建）----------
        if SHEET_NEW_ACCOUNTS not in wb.sheetnames:
            print(f"⚠️ 未找到工作表 '{SHEET_NEW_ACCOUNTS}'，无法解析用户。")
            return
        ws_acc = wb[SHEET_NEW_ACCOUNTS]
        col_id = _col_index(ws_acc, "id")
        col_username = _col_index(ws_acc, "username")
        if not col_username:
            print("⚠️ 'new accounts' 表缺少 username 列。")
            return
        for row_idx in range(2, ws_acc.max_row + 1):
            row = list(ws_acc[row_idx])
            excel_id = _cell_value(row, col_id)
            username = _cell_value(row, col_username)
            if not username:
                continue
            username = str(username).strip()
            user = db.query(User).filter(User.username == username).first()
            if user and excel_id is not None:
                try:
                    user_by_excel_id[int(float(excel_id))] = user
                except (TypeError, ValueError):
                    pass

        # ---------- 3. user_roles：撤销与 batch_create_users 相同的 (user, role) ----------
        if SHEET_USER_ROLES not in wb.sheetnames:
            print(f"⚠️ 未找到工作表 '{SHEET_USER_ROLES}'，无分配可撤销。")
            return
        ws_ur = wb[SHEET_USER_ROLES]
        col_user_id = _col_index(ws_ur, "user_id")
        col_role_id = _col_index(ws_ur, "role_id")
        if not col_user_id or not col_role_id:
            print("⚠️ 'user_roles' 表缺少 user_id 或 role_id 列。")
            return

        print("\n=== 撤销角色分配 (user_roles) ===")
        revoked = 0
        for row_idx in range(2, ws_ur.max_row + 1):
            row = list(ws_ur[row_idx])
            euid = _cell_value(row, col_user_id)
            erid = _cell_value(row, col_role_id)
            try:
                euid, erid = int(float(euid)), int(float(erid))
            except (TypeError, ValueError):
                continue
            user = user_by_excel_id.get(euid)
            role = role_by_excel_id.get(erid)
            if role is None:
                role = db.query(Role).filter(Role.id == erid).first()
            if not user:
                print(f"  跳过 user_id={euid}（未在 new accounts 中找到对应用户）")
                continue
            if not role:
                print(f"  跳过 role_id={erid}（未在 new roles 中且数据库中无该角色）")
                continue
            if role in user.roles:
                if not dry_run:
                    user.roles.remove(role)
                print(f"  ✅ 已撤销: {user.username} <- 角色 {role.name}")
                revoked += 1
            else:
                print(f"  ℹ️ {user.username} 当前未拥有角色 {role.name}，跳过")

        if dry_run:
            db.rollback()
            print(f"\n=== dry-run 完成（未写入数据库，将撤销 {revoked} 条）===")
        else:
            db.commit()
            print(f"\n=== 完成，已撤销 {revoked} 条角色分配 ===")
    except Exception as e:
        db.rollback()
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="按 Excel 撤销由 batch_create_users 分配的角色，便于修正 role.id 后重新分配。"
    )
    parser.add_argument("excel_path", type=str, help="与 batch_create_users 使用的同一 Excel 路径")
    parser.add_argument("--dry-run", action="store_true", help="试运行：只打印将撤销的项，不提交")
    args = parser.parse_args()

    if not os.path.isabs(args.excel_path):
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        excel_path = os.path.join(project_root, args.excel_path)
    else:
        excel_path = args.excel_path

    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        sys.exit(1)

    print("=" * 60)
    print("撤销批量分配的角色")
    print("=" * 60)
    batch_revoke_user_roles(excel_path, dry_run=args.dry_run)
