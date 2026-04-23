"""
批量创建用户并分配角色
支持两种 Excel 格式：
1. 三表格式：new roles / new accounts / user_roles（先建角色、再建用户、再按 user_roles 分配）
2. 单表格式：一个工作表，列含 username, email, password_default, full_name, role.id/role.name

数据库权限：本脚本需要能对 users、roles、user_roles 做 INSERT 的账号。若未设置 DATABASE_URL，
会使用 Vault 中的系统管理员账号；若出现 Access denied，请在环境变量或 backend/.env 中设置
DATABASE_URL（例如 mysql+pymysql://root:你的密码@主机:3306/projectcontrols?charset=utf8mb4）。
"""
import sys
import os

# 添加项目根目录到路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# 先加载 .env，确保 DATABASE_URL 优先生效（本脚本需建用户/角色，建议用 root 等有权限的账号）
from app.database import load_env_with_fallback
load_env_with_fallback()

import openpyxl
from app.database import SessionLocal
from app.models.user import User, Role
from app.models.department import Department

# 表名常量（与 Excel 中一致）
SHEET_NEW_ROLES = "new roles"
SHEET_NEW_ACCOUNTS = "new accounts"
SHEET_USER_ROLES = "user_roles"


def _col_index(ws, header_name):
    """取第一行表头中某列名对应的列索引（1-based），不存在返回 None。"""
    headers = [cell.value for cell in ws[1]]
    for idx, h in enumerate(headers, start=1):
        if h and str(h).strip().lower() == str(header_name).strip().lower():
            return idx
    return None


def _cell_value(row, col_1based):
    """取一行中某列的值，col 为 1-based。row 为 list/tuple，元素可为 Cell 或值。"""
    if col_1based is None or col_1based < 1:
        return None
    idx = col_1based - 1
    if idx >= len(row):
        return None
    cell = row[idx]
    return getattr(cell, "value", cell)


def _bool_val(v, default=True):
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(int(v))
    if isinstance(v, str) and v.isdigit():
        return bool(int(v))
    return default


def batch_create_from_three_sheets(db, wb):
    """
    从三张表执行：新建角色 -> 新建用户 -> 按 user_roles 分配角色。
    Excel 中 id 仅用于 sheet 间关联，数据库 id 由自增生成。
    """
    role_by_excel_id = {}  # excel role id -> Role
    user_by_excel_id = {}  # excel user id -> User

    # ---------- 1. new roles ----------
    if SHEET_NEW_ROLES not in wb.sheetnames:
        print(f"⚠️ 未找到工作表 '{SHEET_NEW_ROLES}'，跳过创建角色。")
    else:
        ws_roles = wb[SHEET_NEW_ROLES]
        col_id = _col_index(ws_roles, "id")
        col_name = _col_index(ws_roles, "name")
        col_desc = _col_index(ws_roles, "description")
        col_active = _col_index(ws_roles, "is_active")
        if not col_name:
            print("⚠️ 'new roles' 表缺少 name 列，跳过。")
        else:
            print("\n=== 1. 创建角色 (new roles) ===")
            for row_idx in range(2, ws_roles.max_row + 1):
                row = list(ws_roles[row_idx])
                excel_id = _cell_value(row, col_id)
                name = _cell_value(row, col_name)
                if not name:
                    continue
                name = str(name).strip()
                description = _cell_value(row, col_desc)
                if description is not None:
                    description = str(description).strip() or None
                is_active = _bool_val(_cell_value(row, col_active), True)

                existing = db.query(Role).filter(Role.name == name).first()
                if existing:
                    role = existing
                    print(f"  角色已存在: {name} (ID: {role.id})")
                else:
                    role = Role(name=name, description=description, is_active=is_active)
                    db.add(role)
                    db.flush()
                    print(f"  ✅ 创建角色: {name} (ID: {role.id})")
                if excel_id is not None:
                    try:
                        role_by_excel_id[int(float(excel_id))] = role
                    except (TypeError, ValueError):
                        pass

    # ---------- 2. new accounts ----------
    if SHEET_NEW_ACCOUNTS not in wb.sheetnames:
        print(f"⚠️ 未找到工作表 '{SHEET_NEW_ACCOUNTS}'，跳过创建用户。")
        return
    ws_acc = wb[SHEET_NEW_ACCOUNTS]
    col_id = _col_index(ws_acc, "id")
    col_username = _col_index(ws_acc, "username")
    col_email = _col_index(ws_acc, "email")
    col_password = _col_index(ws_acc, "default_password") or _col_index(ws_acc, "password_default")
    col_full_name = _col_index(ws_acc, "full_name")
    col_active = _col_index(ws_acc, "is_active")
    col_super = _col_index(ws_acc, "is_superuser")
    col_dept = _col_index(ws_acc, "department") or _col_index(ws_acc, "department.code")
    col_responsible_for = _col_index(ws_acc, "responsible_for")
    if not col_username:
        print("⚠️ 'new accounts' 表缺少 username 列。")
        return

    print("\n=== 2. 创建用户 (new accounts) ===")
    for row_idx in range(2, ws_acc.max_row + 1):
        row = list(ws_acc[row_idx])
        excel_id = _cell_value(row, col_id)
        username = _cell_value(row, col_username)
        if not username:
            continue
        username = str(username).strip()
        email = _cell_value(row, col_email)
        if email is not None:
            email = str(email).strip() or None
        if not email:
            email = f"{username}@cc7.cn"
        password = _cell_value(row, col_password)
        if not password:
            print(f"  第 {row_idx} 行 {username}: 跳过（无密码）")
            continue
        password = str(password)
        full_name = _cell_value(row, col_full_name)
        if full_name is not None:
            full_name = str(full_name).strip() or None
        is_active = _bool_val(_cell_value(row, col_active), True)
        is_superuser = _bool_val(_cell_value(row, col_super), False)
        dept_code = _cell_value(row, col_dept)
        dept_code = str(dept_code).strip() if dept_code else None
        department = None
        if dept_code:
            department = db.query(Department).filter(Department.code == dept_code).first()
            if not department:
                department = db.query(Department).filter(Department.name == dept_code).first()
            if not department:
                print(f"  ⚠️ {username}: 未找到部门 '{dept_code}'，department 留空")
        responsible_for = _cell_value(row, col_responsible_for)
        responsible_for = str(responsible_for).strip() if responsible_for else None

        existing = db.query(User).filter(User.username == username).first()
        if existing:
            user = existing
            new_dept_id = department.id if department else None
            changed = False
            if getattr(user, "department_id", None) != new_dept_id:
                user.department_id = new_dept_id
                changed = True
            if getattr(user, "responsible_for", None) != responsible_for:
                user.responsible_for = responsible_for
                changed = True
            if changed:
                print(f"  用户已存在，更新: {username}")
            else:
                print(f"  用户已存在: {username}")
        else:
            user = User(
                username=username,
                email=email,
                full_name=full_name,
                is_active=is_active,
                is_superuser=is_superuser,
                department_id=department.id if department else None,
                responsible_for=responsible_for,
            )
            user.set_password(password)
            db.add(user)
            db.flush()
            print(f"  ✅ 创建用户: {username} (ID: {user.id})")
        if excel_id is not None:
            try:
                user_by_excel_id[int(float(excel_id))] = user
            except (TypeError, ValueError):
                pass

    # ---------- 3. user_roles 分配 ----------
    if SHEET_USER_ROLES not in wb.sheetnames:
        print(f"\n⚠️ 未找到工作表 '{SHEET_USER_ROLES}'，跳过角色分配。")
    else:
        ws_ur = wb[SHEET_USER_ROLES]
        col_user_id = _col_index(ws_ur, "user_id")
        col_role_id = _col_index(ws_ur, "role_id")
        if not col_user_id or not col_role_id:
            print("\n⚠️ 'user_roles' 表缺少 user_id 或 role_id 列，跳过分配。")
        else:
            print("\n=== 3. 分配角色 (user_roles) ===")
            for row_idx in range(2, ws_ur.max_row + 1):
                row = list(ws_ur[row_idx])
                euid = _cell_value(row, col_user_id)
                erid = _cell_value(row, col_role_id)
                try:
                    euid, erid = int(float(euid)), int(float(erid))
                except (TypeError, ValueError):
                    continue
                user = user_by_excel_id.get(euid)
                role = role_by_excel_id.get(erid)
                if role is None:
                    role = db.query(Role).filter(Role.id == erid).first()
                if not user:
                    print(f"  跳过 user_id={euid}（未在 new accounts 中找到）")
                    continue
                if not role:
                    print(f"  跳过 role_id={erid}（未在 new roles 中且数据库中不存在该角色）")
                    continue
                if role in user.roles:
                    print(f"  ℹ️ {user.username} 已拥有角色 {role.name}")
                else:
                    user.roles.append(role)
                    print(f"  ✅ {user.username} -> 角色 {role.name}")


def batch_create_from_single_sheet(db, excel_path: str):
    """
    单表格式：第一行为表头，列含 username, email, password_default, full_name, is_active, is_superuser, role.id, role.name, department, responsible_for
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {}
    for idx, header in enumerate(headers, start=1):
        if header:
            col_map[str(header).strip()] = idx

    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        username = row[col_map.get("username", 2) - 1].value if "username" in col_map else None
        if not username:
            continue
        username = str(username).strip()
        email = row[col_map.get("email", 3) - 1].value
        if not email or (isinstance(email, str) and email.startswith("=")):
            email = f"{username}@cc7.cn"
        password = row[col_map.get("password_default", 4) - 1].value if "password_default" in col_map else None
        if not password:
            continue
        full_name = row[col_map.get("full_name", 5) - 1].value if "full_name" in col_map else None
        is_active = _bool_val(row[col_map.get("is_active", 6) - 1].value if "is_active" in col_map else None, True)
        is_superuser = _bool_val(row[col_map.get("is_superuser", 7) - 1].value if "is_superuser" in col_map else None, False)
        role_id = row[col_map.get("role.id", 8) - 1].value if "role.id" in col_map else None
        role_name = row[col_map.get("role.name", 9) - 1].value if "role.name" in col_map else None
        dept_val = row[col_map.get("department", 10) - 1].value if "department" in col_map else None
        dept_val = str(dept_val).strip() if dept_val else None
        responsible_for = row[col_map.get("responsible_for", 11) - 1].value if "responsible_for" in col_map else None
        responsible_for = str(responsible_for).strip() if responsible_for else None
        department = None
        if dept_val:
            department = db.query(Department).filter(Department.code == dept_val).first()
            if not department:
                department = db.query(Department).filter(Department.name == dept_val).first()

        existing_user = db.query(User).filter(User.username == username).first()
        if existing_user:
            user = existing_user
        else:
            user = User(
                username=username,
                email=email,
                full_name=full_name,
                is_active=is_active,
                is_superuser=is_superuser,
                department_id=department.id if department else None,
                responsible_for=responsible_for,
            )
            user.set_password(str(password))
            db.add(user)
            db.flush()
            print(f"✅ 创建用户: {username}")

        role = None
        if role_id:
            role = db.query(Role).filter(Role.id == int(role_id)).first()
        if not role and role_name:
            role = db.query(Role).filter(Role.name == role_name).first()
        if role and role not in user.roles:
            user.roles.append(role)
            print(f"  分配角色: {role.name}")
        if "department" in col_map:
            target_dept_id = department.id if department else None
            u = existing_user or user
            if getattr(u, "department_id", None) != target_dept_id:
                u.department_id = target_dept_id
                print(f"  设置部门: {department.code if department else '(空)'}")
        if "responsible_for" in col_map:
            u = existing_user or user
            if getattr(u, "responsible_for", None) != responsible_for:
                u.responsible_for = responsible_for
                print(f"  设置负责内容: {responsible_for or '(空)'}")


def batch_create_users(excel_path: str, dry_run: bool = False):
    """
    从 Excel 批量创建角色、用户并分配角色。
    - 若存在工作表 "new roles" 且 "new accounts"，则按三表流程执行。
    - 否则按单表格式（原逻辑）处理当前活动表。
    - dry_run=True 时：执行全部逻辑但不 commit，最后 rollback，不写入数据库。
    """
    print(f"正在读取: {excel_path}")
    if dry_run:
        print("【dry-run 模式】将执行逻辑但不提交，不会写入数据库。\n")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    db = SessionLocal()
    try:
        if SHEET_NEW_ACCOUNTS in wb.sheetnames:
            batch_create_from_three_sheets(db, wb)
        else:
            print("未检测到 'new accounts' 表，按单表格式处理。")
            batch_create_from_single_sheet(db, excel_path)
        if dry_run:
            db.rollback()
            print("\n=== dry-run 完成（已回滚，未写入数据库）===")
        else:
            db.commit()
            print("\n=== 完成 ===")
    except Exception as e:
        db.rollback()
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="批量创建角色、用户并分配角色。支持三表 Excel（new roles / new accounts / user_roles）或单表格式。"
    )
    parser.add_argument("excel_path", type=str, help="Excel 文件路径（绝对路径或相对项目根）")
    parser.add_argument("--dry-run", action="store_true", help="试运行：执行逻辑但不提交，不写入数据库")
    args = parser.parse_args()

    if not os.path.isabs(args.excel_path):
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        excel_path = os.path.join(project_root, args.excel_path)
    else:
        excel_path = args.excel_path

    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        sys.exit(1)

    if not os.getenv("DATABASE_URL"):
        print("提示：未设置 DATABASE_URL，将使用 Vault 中的系统管理员账号连接。")
        print("若出现 Access denied，请在环境变量或 backend/.env 中设置 DATABASE_URL（例如 root 等有建用户/角色权限的账号）。\n")

    print("=" * 60)
    print("批量创建角色与用户")
    print("=" * 60)
    batch_create_users(excel_path, dry_run=args.dry_run)
