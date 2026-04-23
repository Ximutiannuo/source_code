"""
将本系统涉及的所有数据库表结构导出到 Excel。
仅包含表结构元数据（数据库名、表名、列名、列类型、排序/编码、备注），不包含任何表内数据。

依赖: 在 backend 目录下需已安装 requirements.txt（含 sqlalchemy、pymysql、openpyxl）。
运行: 在 backend 目录执行
  python scripts/export_table_schema_to_excel.py -o table_schema_export.xlsx
  python scripts/export_table_schema_to_excel.py -o D:/output.xlsx --cwd   # 输出到指定路径
"""
import sys
import os
import argparse

# 添加项目根目录到路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# 先加载环境变量
from app.database import load_env_with_fallback
load_env_with_fallback()

from sqlalchemy import create_engine, text
import logging

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# Excel 表头（与需求一致）
HEADERS = [
    "数据库名称",
    "表名称",
    "列名称",
    "列字段",
    "列排序编码格式要求",
    "列备注",
]


def get_database_url():
    """从环境变量或 .env 获取 DATABASE_URL（连接主库 projectcontrols）。"""
    url = os.getenv("DATABASE_URL")
    if not url:
        env_file = os.path.join(os.path.dirname(__file__), '..', '.env')
        if os.path.exists(env_file):
            for enc in ('utf-8', 'gbk', 'gb2312'):
                try:
                    with open(env_file, 'r', encoding=enc, errors='replace') as f:
                        for line in f:
                            if line.strip().startswith('DATABASE_URL='):
                                raw = line.split('=', 1)[1].strip().strip('"').strip("'")
                                url = raw
                                break
                    if url:
                        break
                except Exception:
                    continue
    if not url:
        try:
            from app.services.secret_manager import get_secret_manager
            from app.config import settings
            sm = get_secret_manager()
            user = sm.get_role_username("SYSTEM_ADMIN")
            pwd = sm.get_role_password("SYSTEM_ADMIN")
            if user and pwd:
                host = os.getenv("DB_HOST", getattr(settings, "DB_HOST", "localhost"))
                port = getattr(settings, "DB_PORT", 3306)
                db_name = getattr(settings, "DB_NAME", "projectcontrols")
                import urllib.parse
                url = f"mysql+pymysql://{user}:{urllib.parse.quote_plus(pwd)}@{host}:{port}/{db_name}?charset=utf8mb4"
        except Exception as e:
            logger.debug("SecretManager 不可用: %s", e)
    if not url:
        url = "mysql+pymysql://root:password@localhost:3306/projectcontrols?charset=utf8mb4"
    return url


def get_schema_databases(engine):
    """
    获取本系统涉及的数据库列表。
    默认：projectcontrols、PRECOMCONTROL（来自 config）、ENG（MDR 同步源库，含 ENGDB 等表）。
    """
    from app.config import settings
    dbs = [settings.DB_NAME]
    if getattr(settings, "DB_PRECOMCONTROL_NAME", None):
        dbs.append(settings.DB_PRECOMCONTROL_NAME)
    # ENG：MDR 同步源库（ENG.ENGDB），与 projectcontrols 同实例
    dbs.append("ENG")
    # 排除重复并保持顺序
    seen = set()
    out = []
    for name in dbs:
        name = (name or "").strip()
        if name and name not in seen:
            seen.add(name)
            out.append(name)
    return out


def fetch_columns_metadata(engine, schema_list):
    """
    从 INFORMATION_SCHEMA 拉取列元数据。
    返回 list of dict: 数据库名称, 表名称, 列名称, 列字段, 列排序编码格式要求, 列备注
    """
    placeholders = ", ".join([f":db{i}" for i in range(len(schema_list))])
    sql = text(f"""
        SELECT
            c.TABLE_SCHEMA   AS db_name,
            c.TABLE_NAME     AS table_name,
            c.COLUMN_NAME    AS column_name,
            c.COLUMN_TYPE    AS column_type,
            c.ORDINAL_POSITION AS ordinal_position,
            c.CHARACTER_SET_NAME AS character_set_name,
            c.COLLATION_NAME AS collation_name,
            IFNULL(c.COLUMN_COMMENT, '') AS column_comment
        FROM INFORMATION_SCHEMA.COLUMNS c
        WHERE c.TABLE_SCHEMA IN ({placeholders})
        ORDER BY c.TABLE_SCHEMA, c.TABLE_NAME, c.ORDINAL_POSITION
    """)
    params = {f"db{i}": name for i, name in enumerate(schema_list)}
    with engine.connect() as conn:
        rows = conn.execute(sql, params).fetchall()
    result = []
    for row in rows:
        r = row._mapping if hasattr(row, "_mapping") else dict(row)
        pos = r.get("ordinal_position") or 0
        cs = r.get("character_set_name")
        coll = r.get("collation_name")
        if cs or coll:
            sort_encoding = f"排序: {pos}; 字符集: {cs or '-'}; 排序规则: {coll or '-'}"
        else:
            sort_encoding = f"排序: {pos}"
        result.append({
            "数据库名称": r.get("db_name") or "",
            "表名称": r.get("table_name") or "",
            "列名称": r.get("column_name") or "",
            "列字段": r.get("column_type") or "",
            "列排序编码格式要求": sort_encoding,
            "列备注": (r.get("column_comment") or "").strip(),
        })
    return result


def write_excel(rows, path):
    """将行数据写入 Excel，表头为 HEADERS。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("请安装 openpyxl: pip install openpyxl")
        raise
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表结构"
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            val = row.get(key, "")
            if val is None:
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24
    wb.save(path)
    logger.info("已写入: %s (共 %d 行)", path, len(rows))


def main():
    parser = argparse.ArgumentParser(description="导出本系统所有表结构到 Excel（仅元数据，不含数据）")
    parser.add_argument(
        "-o", "--output",
        default="table_schema_export.xlsx",
        help="输出 Excel 文件路径（默认: table_schema_export.xlsx）",
    )
    parser.add_argument(
        "--cwd",
        action="store_true",
        help="输出到当前工作目录而非脚本所在目录",
    )
    args = parser.parse_args()
    out_path = args.output
    if not os.path.isabs(out_path) and not args.cwd:
        out_path = os.path.join(os.path.dirname(__file__), '..', out_path)
    out_path = os.path.abspath(out_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    url = get_database_url()
    engine = create_engine(url, pool_pre_ping=True)
    schema_list = get_schema_databases(engine)
    logger.info("导出数据库: %s", schema_list)
    rows = fetch_columns_metadata(engine, schema_list)
    if not rows:
        logger.warning("未查询到任何列信息，请检查数据库名或连接权限")
    else:
        write_excel(rows, out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
