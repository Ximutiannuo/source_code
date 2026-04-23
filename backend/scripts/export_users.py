"""
导出用户清单到 Excel，用于配置部门信息

前置：请先执行 migrate_add_departments_and_user_department.py 创建 departments 表及 users.department_id。

导出列：id, username, email, full_name, is_active, department_id, department.code, department.name, roles
- department.code / department.name 为当前已配置的部门（若有）
- 请在 department.code 列填写部门代码（design, procurement 等），或 department.name 填写部门名称
- 配置完成后，使用 update_users_department.py 将部门信息写回数据库

执行方式：
  cd backend && python -m scripts.export_users [输出路径]
  或
  python backend/scripts/export_users.py users_export.xlsx
"""
import sys
import os
from pathlib import Path
from datetime import datetime

project_root = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "backend"))

from app.database import load_env_with_fallback

if not os.getenv("DATABASE_URL"):
    load_env_with_fallback()

import openpyxl
from openpyxl.styles import Font, Alignment
from app.database import SessionLocal
from app.models.user import User, Role
from app.models.department import Department


def export_users(output_path: str):
    db = SessionLocal()
    try:
        users = db.query(User).order_by(User.id).all()
        depts = {d.id: d for d in db.query(Department).all()}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "users"

        headers = [
            "id",
            "username",
            "email",
            "full_name",
            "is_active",
            "department_id",
            "department.code",
            "department.name",
            "responsible_for",
            "roles",
        ]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, user in enumerate(users, start=2):
            dept = depts.get(user.department_id) if getattr(user, "department_id", None) else None
            roles_str = ", ".join(r.name for r in user.roles) if user.roles else ""
            ws.cell(row=row_idx, column=1, value=user.id)
            ws.cell(row=row_idx, column=2, value=user.username)
            ws.cell(row=row_idx, column=3, value=user.email or "")
            ws.cell(row=row_idx, column=4, value=user.full_name or "")
            ws.cell(row=row_idx, column=5, value="是" if user.is_active else "否")
            ws.cell(row=row_idx, column=6, value=user.department_id if getattr(user, "department_id", None) else "")
            ws.cell(row=row_idx, column=7, value=dept.code if dept else "")
            ws.cell(row=row_idx, column=8, value=dept.name if dept else "")
            ws.cell(row=row_idx, column=9, value=getattr(user, "responsible_for", None) or "")
            ws.cell(row=row_idx, column=10, value=roles_str)

        # 部门清单说明（第二个工作表）
        ws2 = wb.create_sheet("departments_list")
        ws2.cell(row=1, column=1, value="department.code")
        ws2.cell(row=1, column=2, value="department.name")
        for row_idx, d in enumerate(db.query(Department).order_by(Department.sort_order, Department.id).all(), start=2):
            ws2.cell(row=row_idx, column=1, value=d.code)
            ws2.cell(row=row_idx, column=2, value=d.name)

        wb.save(output_path)
        print(f"✅ 已导出 {len(users)} 个用户到: {output_path}")
        print("  - 工作表 'users'：用户清单，请填写 department.code、responsible_for 列")
        print("  - 工作表 'departments_list'：部门清单，可供参考")
        print("  配置完成后运行: python -m scripts.update_users_department <excel_path>")
    finally:
        db.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="导出用户清单到 Excel，用于配置部门。"
    )
    default_name = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    parser.add_argument("output_path", type=str, nargs="?", default=default_name, help=f"输出 Excel 路径，默认 {default_name}")
    args = parser.parse_args()

    if not os.path.isabs(args.output_path):
        output_path = os.path.join(project_root, args.output_path)
    else:
        output_path = args.output_path

    print("=" * 60)
    print("导出用户清单")
    print("=" * 60)
    try:
        export_users(output_path)
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
