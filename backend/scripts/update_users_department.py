"""
从配置好的 Excel 更新用户部门及负责内容

Excel 需包含 users 工作表，列：id 或 username, department.code 或 department.name, responsible_for
按 id 或 username 匹配用户，将 department、responsible_for 更新到用户。

执行方式：
  cd backend && python -m scripts.update_users_department users_export_configured.xlsx [--dry-run]
"""
import sys
import os
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "backend"))

from app.database import load_env_with_fallback

if not os.getenv("DATABASE_URL"):
    load_env_with_fallback()

import openpyxl
from app.database import SessionLocal
from app.models.user import User
from app.models.department import Department


def _col_index(ws, header_name):
    headers = [cell.value for cell in ws[1]]
    for idx, h in enumerate(headers, start=1):
        if h and str(h).strip().lower() == str(header_name).strip().lower():
            return idx
    return None


def _cell_value(row, col_1based):
    if col_1based is None or col_1based < 1:
        return None
    idx = col_1based - 1
    if idx >= len(row):
        return None
    cell = row[idx]
    return getattr(cell, "value", cell)


def update_users_department(excel_path: str, dry_run: bool = False):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws_name = "users" if "users" in wb.sheetnames else wb.active.title
    ws = wb[ws_name]

    col_id = _col_index(ws, "id")
    col_username = _col_index(ws, "username")
    col_dept_code = _col_index(ws, "department.code")
    col_dept_name = _col_index(ws, "department.name")
    col_responsible_for = _col_index(ws, "responsible_for")
    has_dept = col_dept_code or col_dept_name
    has_any = has_dept or col_responsible_for
    if not has_any:
        print("⚠️ 未找到 department.code、department.name 或 responsible_for 列")
        return
    if not col_id and not col_username:
        print("⚠️ 未找到 id 或 username 列")
        return

    db = SessionLocal()
    try:
        dept_by_code = {d.code: d for d in db.query(Department).all()}
        dept_by_name = {d.name: d for d in db.query(Department).all()}
        updated = 0
        for row_idx in range(2, ws.max_row + 1):
            row = list(ws[row_idx])
            user_id = _cell_value(row, col_id)
            username = _cell_value(row, col_username)
            dept_code = _cell_value(row, col_dept_code) if col_dept_code else None
            dept_name = _cell_value(row, col_dept_name) if col_dept_name else None
            dept_val = (dept_code or dept_name)
            dept_val = str(dept_val).strip() if dept_val else None
            responsible_for_val = _cell_value(row, col_responsible_for) if col_responsible_for else None
            responsible_for_val = str(responsible_for_val).strip() if responsible_for_val else None
            department = None
            if dept_val:
                department = dept_by_code.get(dept_val) or dept_by_name.get(dept_val)
                if not department:
                    print(f"  第 {row_idx} 行: 未找到部门 '{dept_val}'")
                    continue

            user = None
            if user_id is not None:
                try:
                    user = db.query(User).filter(User.id == int(float(user_id))).first()
                except (TypeError, ValueError):
                    pass
            if not user and username:
                username = str(username).strip()
                if username:
                    user = db.query(User).filter(User.username == username).first()
            if not user:
                continue

            changed = False
            if has_dept:
                target_dept_id = department.id if department else None
                if getattr(user, "department_id", None) != target_dept_id:
                    user.department_id = target_dept_id
                    changed = True
            if col_responsible_for and getattr(user, "responsible_for", None) != responsible_for_val:
                user.responsible_for = responsible_for_val
                changed = True
            if changed:
                updated += 1
                parts = []
                if has_dept:
                    parts.append(f"部门:{department.code if department else '(空)'}")
                if col_responsible_for:
                    parts.append(f"负责:{responsible_for_val or '(空)'}")
                print(f"  ✅ {user.username} (id={user.id}) -> {', '.join(parts)}")

        if dry_run:
            db.rollback()
            print(f"\n【dry-run】将更新 {updated} 个用户的部门，已回滚")
        else:
            db.commit()
            print(f"\n✅ 已更新 {updated} 个用户的部门")
    finally:
        db.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="从 Excel 更新用户部门")
    parser.add_argument("excel_path", type=str, help="配置好的 Excel 文件路径")
    parser.add_argument("--dry-run", action="store_true", help="试运行，不提交")
    args = parser.parse_args()

    if not os.path.isabs(args.excel_path):
        excel_path = os.path.join(project_root, args.excel_path)
    else:
        excel_path = args.excel_path

    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        sys.exit(1)

    print("=" * 60)
    print("更新用户部门")
    print("=" * 60)
    try:
        update_users_department(excel_path, dry_run=args.dry_run)
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
