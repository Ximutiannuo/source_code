"""
导入 WorkSteps.xlsx 到 workstep_defines 表
识别绿色底纹单元格，设置 is_key_quantity = True
"""
import sys
import os
from pathlib import Path

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "backend"))

import openpyxl
from openpyxl.styles.colors import RGB
from sqlalchemy.orm import Session
from app.database import SessionLocal
from app.models.workstep import WorkStepDefine
from datetime import datetime, timezone
from decimal import Decimal


def get_cell_color(cell):
    """获取单元格背景色，识别绿色底纹"""
    if not cell.fill:
        return None
    
    fill = cell.fill
    if not hasattr(fill, 'patternType') or fill.patternType == 'none':
        return None
    
    if hasattr(fill, 'start_color') and fill.start_color:
        # 优先检查主题色（Excel 中绿色底纹通常使用主题色）
        if hasattr(fill.start_color, 'theme') and fill.start_color.theme is not None:
            return f"theme_{fill.start_color.theme}"
        # 其次检查索引色
        elif hasattr(fill.start_color, 'indexed') and fill.start_color.indexed is not None:
            return f"index_{fill.start_color.indexed}"
        # 最后检查 RGB 值（排除 '00000000' 这种无效值）
        elif hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
            rgb_val = fill.start_color.rgb
            # 如果是字符串且不是 '00000000'，才返回
            if isinstance(rgb_val, str) and rgb_val != '00000000':
                return rgb_val
            # 如果是 RGB 对象，转换为字符串
            elif isinstance(rgb_val, RGB):
                rgb_str = str(rgb_val)
                if rgb_str != '00000000':
                    return rgb_str
    
    return None


def is_green_background(cell):
    """判断是否为绿色底纹（关键数据）"""
    color = get_cell_color(cell)
    if not color:
        return False
    
    # openpyxl 的 RGB 对象实际上是一个字符串，格式为 "AARRGGBB" 或 "RRGGBB"
    rgb_str = None
    
    if isinstance(color, RGB):
        # RGB 对象转换为字符串
        rgb_str = str(color)
    elif isinstance(color, str):
        rgb_str = color
    
    if rgb_str:
        # 处理不同的格式
        # 格式1: "FFC6EFCE" (8位，前两位是透明度)
        # 格式2: "C6EFCE" (6位)
        # 格式3: "00000000" (黑色，8位)
        
        # 去掉前导的 "FF" 或 "00"（透明度）
        if len(rgb_str) == 8:
            rgb_hex = rgb_str[2:]  # 去掉前两位透明度
        elif len(rgb_str) == 6:
            rgb_hex = rgb_str
        else:
            rgb_hex = None
        
        if rgb_hex and len(rgb_hex) == 6:
            try:
                # 转换为 RGB 值
                r = int(rgb_hex[0:2], 16)
                g = int(rgb_hex[2:4], 16)
                b = int(rgb_hex[4:6], 16)
                
                # 检查是否为绿色（浅绿色：G值高，R和B值较低）
                # 典型的浅绿色 RGB(198, 239, 206) 或类似的
                if g > 200 and r < 220 and b < 220:
                    return True
            except (ValueError, IndexError):
                pass
    
    # 检查主题色6（Excel 中绿色底纹使用主题色6）
    if isinstance(color, str) and "theme_6" in color:
        return True
    
    # 检查索引色（根据之前的分析，索引色6可能是绿色）
    if isinstance(color, str) and "index_6" in color:
        return True
    
    return False


def import_worksteps_from_excel(file_path: str, db: Session):
    """从 WorkSteps.xlsx 导入工作步骤定义"""
    print(f"\n正在导入工作步骤定义: {file_path}")
    
    if not os.path.exists(file_path):
        print(f"  ✗ 错误: 文件不存在: {file_path}")
        return 0, 0, [f"文件不存在: {file_path}"]
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"  工作表: {ws.title}")
        print(f"  总行数: {ws.max_row}, 总列数: {ws.max_column}")
        
        # 读取表头（第1-2行）
        row1_headers = []
        row2_headers = []
        
        for col in range(1, ws.max_column + 1):
            cell1 = ws.cell(row=1, column=col)
            cell2 = ws.cell(row=2, column=col)
            row1_headers.append(cell1.value if cell1.value else "")
            row2_headers.append(cell2.value if cell2.value else "")
        
        # 找到关键列的位置
        level2_code_col = None  # Level II Code
        level2_disc_col = None  # Level II Discipline
        level3_code_col = None  # Level III Code
        level3_group_col = None  # Level III Work Group
        unit_col = None  # Unit
        weight_type_col = None  # Weight by work Steps
        
        for i, (h1, h2) in enumerate(zip(row1_headers, row2_headers)):
            if h1 == "Level II" and h2 == "Code":
                level2_code_col = i + 1
            elif h1 == "Level II" and h2 == "Discipline":
                level2_disc_col = i + 1
            elif h1 == "Level III" and h2 == "Code":
                level3_code_col = i + 1
            elif h1 == "Level III" and h2 == "Work Group":
                level3_group_col = i + 1
            elif h1 == "Unit":
                unit_col = i + 1
            elif h1 == "Weight by work Steps":
                weight_type_col = i + 1
        
        if not level3_code_col:
            print("  ✗ 错误: 找不到 Level III Code 列")
            return 0, 0, ["找不到 Level III Code 列"]
        
        # Workstep 列从第7列开始（列G，索引7）
        workstep_start_col = 7
        
        imported = 0
        updated = 0
        errors = []
        
        # 从第3行开始读取数据
        current_work_package = None
        current_discipline = None
        
        for row in range(3, ws.max_row + 1):
            # 读取 Level III Code（工作包）
            level3_code = ws.cell(row=row, column=level3_code_col).value
            level3_group = ws.cell(row=row, column=level3_group_col).value if level3_group_col else None
            level2_code = ws.cell(row=row, column=level2_code_col).value if level2_code_col else None
            level2_disc = ws.cell(row=row, column=level2_disc_col).value if level2_disc_col else None
            weight_type = ws.cell(row=row, column=weight_type_col).value if weight_type_col else None
            
            # 如果是描述行（包含 Level III Code）
            if level3_code and weight_type == "Work Step Descriptions":
                current_work_package = str(level3_code).strip()
                current_discipline = str(level2_disc).strip() if level2_disc else None
                
                # 读取工作步骤描述和权重
                for col in range(workstep_start_col, ws.max_column + 1):
                    workstep_desc = ws.cell(row=row, column=col).value
                    if not workstep_desc or str(workstep_desc).strip() == "":
                        continue
                    
                    # 检查下一行（权重行）的对应权重
                    weight_value = None
                    if row + 1 <= ws.max_row:
                        next_weight_type = ws.cell(row=row + 1, column=weight_type_col).value if weight_type_col else None
                        if next_weight_type == "Work Step Weight %":
                            weight_value = ws.cell(row=row + 1, column=col).value
                    
                    # 检查是否为关键数量（绿色底纹）
                    cell = ws.cell(row=row, column=col)
                    is_key = is_green_background(cell)
                    
                    # 保存到数据库
                    workstep_desc_str = str(workstep_desc).strip()
                    
                    if workstep_desc_str and workstep_desc_str.lower() != "not applicable":
                        try:
                            # 检查是否已存在
                            existing = db.query(WorkStepDefine).filter(
                                WorkStepDefine.work_package == current_work_package,
                                WorkStepDefine.work_step_description == workstep_desc_str
                            ).first()
                            
                            weight_decimal = None
                            if weight_value is not None:
                                try:
                                    weight_decimal = Decimal(str(weight_value))
                                except:
                                    pass
                            
                            if existing:
                                # 更新
                                existing.work_step_weight = weight_decimal
                                existing.is_key_quantity = is_key
                                existing.sort_order = col - workstep_start_col + 1
                                existing.updated_at = datetime.now(timezone.utc)
                                updated += 1
                            else:
                                # 新建
                                new_workstep = WorkStepDefine(
                                    work_package=current_work_package,
                                    work_step_description=workstep_desc_str,
                                    work_step_weight=weight_decimal,
                                    is_key_quantity=is_key,
                                    sort_order=col - workstep_start_col + 1,
                                    is_active=True
                                )
                                db.add(new_workstep)
                                imported += 1
                        except Exception as e:
                            error_msg = f"行{row}列{col}: {str(e)}"
                            errors.append(error_msg)
                            print(f"  ✗ {error_msg}")
        
        db.commit()
        print(f"  ✓ 导入完成: 新增 {imported} 条, 更新 {updated} 条")
        if errors:
            print(f"  ⚠ 错误 {len(errors)} 条")
        
        wb.close()
        return imported, updated, errors
        
    except Exception as e:
        db.rollback()
        error_msg = f"导入失败: {str(e)}"
        print(f"  ✗ {error_msg}")
        import traceback
        traceback.print_exc()
        return 0, 0, [error_msg]


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='导入 WorkSteps.xlsx 到 workstep_defines 表')
    parser.add_argument('--file', type=str, default='original system/WorkSteps.xlsx',
                        help='WorkSteps.xlsx 文件路径')
    parser.add_argument('--clear', action='store_true', help='清空现有数据（谨慎使用）')
    
    args = parser.parse_args()
    
    file_path = Path(project_root) / args.file
    
    db = SessionLocal()
    try:
        if args.clear:
            print("⚠ 警告: 将清空 workstep_defines 表的所有数据")
            confirm = input("确认清空? (yes/no): ")
            if confirm.lower() == 'yes':
                db.query(WorkStepDefine).delete()
                db.commit()
                print("✓ 已清空 workstep_defines 表")
            else:
                print("取消清空操作")
                return
        
        imported, updated, errors = import_worksteps_from_excel(str(file_path), db)
        
        print(f"\n导入结果:")
        print(f"  新增: {imported} 条")
        print(f"  更新: {updated} 条")
        if errors:
            print(f"  错误: {len(errors)} 条")
            for error in errors[:10]:  # 只显示前10个错误
                print(f"    - {error}")
        
    except Exception as e:
        print(f"✗ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    import argparse
    main()

